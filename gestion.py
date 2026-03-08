import streamlit as st
import pandas as pd
import os
from datetime import date, timedelta
import gspread
from google.oauth2.service_account import Credentials
from streamlit_option_menu import option_menu
from streamlit_calendar import calendar
import base64
import plotly.graph_objects as go
import plotly.express as px
import calendar as cal_module
import unicodedata

def normalizar(texto):
    """Quita tildes y convierte a mayúsculas para comparaciones robustas."""
    if not isinstance(texto, str):
        return ""
    return unicodedata.normalize('NFD', texto).encode('ascii', 'ignore').decode('ascii').upper()

def mask_forma(serie, palabra):
    """Filtra una Serie por forma de pago ignorando tildes y mayúsculas."""
    palabra_norm = normalizar(palabra)
    return serie.fillna('-').apply(normalizar).str.contains(palabra_norm, na=False)

# --- 1. CONFIGURACIÓN Y CONEXIÓN ---
st.set_page_config(page_title="CHACAGEST - GESTIÓN TOTAL", page_icon="🚛", layout="wide")


LOGO_B64 = "data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2wCEAAkGBxISERUQEBMVEhUWEhUVEBYREhAXFRUXGBIWFxUVFRUYHSggGBslHxUVITEiJSkrLi4wGB8zODMsOCgtLisBCgoKDQ0NGQ8PFzcZFx8rKy0rMis3Ky0rLTcrNys3LTcrLSstLSstMC0tMDIrNystNy03LTY3Ny8rLSs3LTcrK//AABEIALcBEwMBIgACEQEDEQH/xAAcAAEAAgIDAQAAAAAAAAAAAAAABgcBBQMECAL/xABHEAACAQICBQcHCQUHBQAAAAAAAQIDEQQhBQYxQWEHEhMiUXGBMkJTVJHR0hQVFiNSk5ShwReSorHhM0NicnOC8EVjhKOz/8QAGgEBAQADAQEAAAAAAAAAAAAAAAECAwYEBf/EACQRAQACAgEDAwUAAAAAAAAAAAABAgMRBBIxUQUhIhMjQmFx/9oADAMBAAIRAxEAPwC8QAAAAAAAAAAAAAAAAAAAAAAAAcGLxcKUJVaslCEIuU5S2JJXbIsuU/RPra+6r/ABMCOaz6UcbUabae2bTat2RvxNbiuVHRag3DEqUrdVKnXze7zSHVddMFKTlKum27t8yr8IEi+WVfS1PvJ+8fLKvpKn3k/eR6GtmCbSVdXbslzal29iSXN25oli0FXtfmZWvnKCy9pYR1PllX0lT7yfvHy2r6Sp95P3nC0fVGk5yUYq7bshI3Or9OrVqXlUqOEc5deeb3LaS86mjsEqVNQj/ufa97O2RQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjmvms8dH4Sdd2dR9TDxfnVGnbLsWbfBAVxy3623a0ZQnkmpYtx35XhSfDZJr/AClRHJXrynOVSpJynKTlOT2uTd233s4woAbLV3QtTG4mnhaPlTl1pboRWcpvglfv2ATvkW1S+UV3jq0fqqLtRT2Tq73xjFfm+BbWs2keZHoovrS8q26PZ4nawmFo4HCxpU1zYUoKMFvk7b+1t5kOxNeU5Octrbb/AEQRxko1W0bZdPJZvKF+ze/E0uh8A61RR81Zza7OzvZO4RSVlklsRZkZRkAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbA+Kskk23ZJNtvYkt7PM3KRrW9I4tzg/qad4YdXyaum6lu2TS8Eix+W3W5UqXzdRf1lWN67XmUnfq98rPwXEo0AAAoegeR7VJ4TDPE1l9fXSdms6dPbGHe/KfeuwrHkt1dp4rFqriGlQotSkpf3k79SHdezfcu0vbTOmYKnzaUlKUsuq/JW9hGq1j0j0s+ZHyIP2y3s1EYtuyzb2foYJFqvo676aSyV1BPe97A3OhcAqNNR855zfHs8DYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGajWnT1PA4WpiquyK6sd85vKEFxba8Lm2Z535XdbvlmKeHpSvQoNxVtlSp58+5eSvF70BDNK6RqYmtUxFZ86dSTlJ3duCXBKyXBHUAChzYPCyq1I0qavKclGPf/Tb4HCWNyc6C5kPldRdaatRvug9sv938gJPoTRkcNQjRhuXWe+UnnKV/y8DvgykZaR2dG4N1aigu+T7F2k9oUVCKhHJJJIjeDx2EwNPnYqvRoynm+kqQi+EUm7vwOhjOVbRUNleVX/SpVH+bSTMRNwVrV5aNHryadeXfCMfybON8teC9DX9kPeBZwKvfLZg/QV//AF+8x+2zB+r1/wCD3gWiCrHy3YT1av7afvMftvwvq1f20veBagKpfLhhvVa/71L3mHy44b1Sv+/R94Frgqd8uOH9Ur/v0fefP7csP6nW8alH3gW0DQ6E03XxNCniI4V01Ujzoxq1YqaV3ZtJO11n4gDfAAAAAAAAAAAAAAAAAxc1Gs+sNHA4eWIrvJZQj505box4gRblf1teCwvQUZWr104xtthT8+pbt81cXfceeTY6w6bq43ETxVd9abyS2Qir82EeCXtua4AAZpwcmoxTk27RS2tvYkFbnVLQrxVdQf8AZw61V8L2UfFlwxikrLJJWXYkv6Gk0Ho+ngML9ZJRdudWl/its422L+pBtZ9b6mIbp0r06Ozb1p8Zdi4FRLdO67UKF4UvrprLqvqJ8Ze4g+k9bcXW21Ojj9mkuavb5X5mjsBs0zNttybbb8pttt97ebMAEUAAAwZAAAAAAAJfyY6qPH4tc9PoKLU67z62+FK/a9/YvAiuDws6tSFKlHnTnJRgu1t5LgeodRdWoaPwkMNHrS8qtO1nOpLOT7lsS7EgjexjZWSVlkrZWW5GDlAAAxJgZBpsdrXgKNSVKvi8PSqRtzoVK9KMldJq8W7rJp+JwfTjRnr+E/E0fiAkAI/9ONGev4T8TR+IfTjRnr+E/E0fiAkAI/8ATjRnr+E/E0fiH040Z6/hPxNH4gJACP8A040Z6/hPxNH4jq4nlG0VTV3jaUrei59T/wCaYEqMc4rXSXLRgIJ9BCtXe7qKnF+M3f8AIgOsfKxj8QnCi1hIPL6q7qtf6j2eCT4gW/rnrzhdHRfSS6Sra8KNNpzff9lcWefNa9Zq+kK3TYh7LqlCN+ZTT3RvteWb2s09Sbk3KTcm3eTk2232tvNvifIUAAAnHJ9oZK+NrWjGF+icu1K0p9y2eJGdX9ESxVaNKOSunUl9mN8337Uu8kuvml4wjHAYfqwhFdLb+Gn+r8AjT626xyxdS0bqjF9SOznf45Lt7Ow0AAUAAAEl1a1Ex+OtKjS5lPfVrNwp27VleXgn4Fl6E5FMPFJ4uvOtLfGlanC/B5yftAo8xc9O4Tk40VTtbB058avOqX71NtP2Gyjqno9KywWFS4Yah8IHlC4uesfotgPUsL+GofCPotgPUsL+GofCB5OuLnrH6LYD1LC/hqHwj6LYD1LC/hqHwgeTrhtHrH6LYD1LC/hqHwj6LYD1LC/hqHwgVnyIao/9TrRzd44VP2Tq97zSff2lxJHHh6EacYwhFRjFJRjFKMYpKyUYrJLgjlCAAAMj+u2skMBhJ4idnLyaMPt1H5K7ltfBM30pJK7aSW2+482cqGtrx+Lag/qKLcKGflNPrVPHdwsBE8di51qk61WXOnObnNva2/8An5HCAFAAAAAAAAAAAAAA58DhJ1qkaVOLlKTskv5t7lxOxofRFXE1OZRi39qTuox/zMtPV7V+lgqbs+dNpupUllxaXZFAazoqeisFJq0q0978+o1kv8kezhxKyq1XJuUm5Sk7yb2tva3xNzrdpr5VXck/q4XjSW6185d79xpAgAYuFclGlKclCEXOUpKMYxTbbbskki7dQeSmnSUcRpGKq1cnCi86dPeuevPlw2LvOxyRairD0447Ew+vqK9KMl/YwezLdOS9iaRZyCMRikkkklutsPqwAAAAAAAAAAAAAAAAAFY8tGtvyagsFRlatXT6Rp506W/xlsXBMoax6D03yTUMXXqYmvisS51Jc550rRW6MVzcklZHTXIlgvWMR7aXwAUQC+FyJ4H02I9tL4T6XIpgfTYj96n8IFCgvv8AYtgPS1/34fCJcjej0rupXtxqR9wFCAvhckei/S1fvY+4z+yTRXpan30RuGXTbwoYF9rkm0T6Wp9+jK5KdEfbm/8AyBuF6Z8KDMSkltaXeeg6eoGg6OclGX+pXlJfzO/o+Wh6ErYalQ5/munSUpZf42m14sm4X6WSY3FfZ5/wGgcTWt0dGdnslKLjG3beVrruJfobk+SfPxVTnf8Abp7P909/giwsXiHUm5y3/ktyRsMLq/WqQU04JNXSk5J29jMmto8LhYUoKnTioRWxRSSIjyi6d5kPklN9aavVs9kL+Twvb8ic6zYd4HDTxVaVPmxyilKV5SeUIJc3a3+pQmNxUqtSVWbvKcnKT/52ZLwA4QARQnnJHqk8bilXqq9ChJSnfZOpthDuTzfdxIbovR9TEVoUKMedOpJRgt1+19iWbfcepdVNBU8DhaeFpZ82PXlZXnN+VN8W/wBAjbxRkAAAAAAAAAAAAAAAAAAAAAAAAAAdXSODjWpSpT2SVmdowyd1iZidwonTGBnh606M73i8nd9aPmyXgdPnPtftZafKDoHpqPTU1epTV8lnKO9d+8qo+dmpNLfp3HpvJpyMMTr5R7Szd9oMA1bfQ6Y8FiUaDwXRw50l1pbeC3LhlZ+JqtB4HpJ86Xkx28XuXh+hKacHJqKV23ZJb7ns42PfzlznrfN19ik/1sNBaP6apn5Mc5fovEm6SS4HU0VgVRpqC27ZPte8hXK/rf8AI8N8noytXrppW206eyU+Dd7Lx7GetzKtuVvW75biuhpSvQoSlGNtlSpsnPilnFdqu95AwAoASnk51UekcWqc0+gp2niGr5xv1YX7ZNNdyYFj8iWqHRU/nGvG1SrHm4dPbGk9su+WXgl2lrJGKdNRSjFJJJKKSsklkkkfQQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfM43Kg140G8NX50V9XUvKHB+dH9fEuBmq1j0VHE0JUpZO14P7MtzNWWnXXT3+ncyeNmi34z3UifdGk5SUY7W7L/nYZxNCVOcqc1aUZNSXFG+1fwPNj0stsl1eEdt/E8OPHNraddzOZTBg+pve+zY4TDqnBQjsX5ve2SvVXR39/Jdqp/rI02isC61RQ3bZvsRPKcVFKKVkkkuCPpxGo1Dhb3te02t3l1NMaSp4ajUxFZ82FOLlJ78tyW9vYu88s6yacq43EzxVbJzfVis1CK8mC7l7Xdk95bNbunrfIKL+qou9drz6u6PdH+b4Z1eGAAAr7oUpTlGEE5SlJRgltcpO0Uu92PT2oWq0dH4SNFWdST59eSt1ptbO5K0VwXErrkP1S50npOtHKLccKpLbJXU6q7s4p95dSCMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGGjJ8zkkm3kkrtgQbXnV6E6tPEJpX6tWO+Vlk1+a8ToJbku5I72mMf01Ry81ZQXDt72dKE3Fpp2azTJWsRMy25M+S9K0tO4r2TbQeA6Gmr+VLOfuNFym61rR+Dbg109VOGHXY2s6jXZFZ+w6PzrX9LP2mp0vgKWKkp4qCrSiubF1M7LsXZ/RGWmpRs5Nttttttttttt5tt9pguV6r4L1en4J+8k+geTzR7p8+thKUnLOKaeS9u8g85m71N1dnj8XTw0E1FvnVp7oU1m5d7yS4tHoT9nmivUqP7r95s9C6t4TCc54XD06LnbnuEbN22JsDuaPwcKNKFGlFQhTioQitiilZI7IAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA02scqjiqdON+dnN3isuzNgARr5qrfY/ih7x81VvsfxQ95kFRj5qrfY/ih7x81VvsfxQ95kFkdvRWhZyqLpY2is3nHPPJZMmKAMVZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAf/2Q=="
COL_CLIENTES         = ["Razón Social", "CUIT / CUIL / DNI *", "Email", "Teléfono", "Dirección Fiscal", "Localidad", "Provincia", "Condición IVA", "Condición de Venta"]
COL_VIAJES           = ["Fecha Carga", "Cliente", "Fecha Viaje", "Origen", "Destino", "Patente / Móvil", "Importe", "Tipo Comp", "Nro Comp Asoc"]
COL_PRESUPUESTOS     = ["Fecha Emisión", "Cliente", "Vencimiento", "Detalle", "Tipo Móvil", "Importe"]
COL_TESORERIA        = ["Fecha", "Tipo", "Caja/Banco", "Forma", "Concepto", "Cliente/Proveedor", "Monto", "Ref AFIP"]
COL_PROVEEDORES      = ["Razón Social", "CUIT/DNI", "Cuenta de Gastos", "Categoría IVA", "CBU", "Alias"]
COL_COMPRAS          = ["Fecha", "Proveedor", "Punto Venta", "Tipo Factura", "Neto 21", "Neto 10.5", "Ret IVA", "Ret Ganancia", "Ret IIBB", "No Gravados", "Total"]
# Cheques emitidos por la empresa (pagos con cheque propio)
COL_CHEQ_EMITIDOS    = ["Fecha Emisión", "Nro Cheque", "Tipo", "Banco", "Beneficiario", "Importe", "Fecha Vencimiento", "Estado", "Fecha Conciliación", "Observaciones"]
# Cheques de terceros recibidos en cobranzas (cartera)
COL_CHEQ_CARTERA     = ["Fecha Recepción", "Nro Cheque", "Tipo", "Banco Librador", "Librador", "Importe", "Fecha Vencimiento", "Estado", "Destino", "Fecha Aplicación", "Observaciones"]

# Módulo de Facturación
COL_FACTURAS = ["Fecha", "Tipo", "Punto Venta", "Numero", "Cliente", "CUIT Cliente", "Condicion IVA",
                "Detalle", "Neto", "IVA", "No Gravado", "Total", "Estado", "Comp Asoc Tipo", "Comp Asoc Nro", "Observaciones"]

def conectar_google():
    """Devuelve la conexión a Google Sheets, reutilizando la del session_state si ya existe."""
    nombre_planilla = "Base_Chacagest"
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

    # Reutilizar conexión existente para no agotar cuota de lecturas
    if "gsheets_conn" in st.session_state and st.session_state.gsheets_conn is not None:
        return st.session_state.gsheets_conn

    try:
        if "gcp_service_account" in st.secrets:
            creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scope)
        else:
            creds = Credentials.from_service_account_file("llave_google.json", scopes=scope)
        client = gspread.authorize(creds)
        sh = client.open(nombre_planilla)
        st.session_state.gsheets_conn = sh   # cachear para toda la sesión
        return sh
    except Exception as e:
        st.session_state.gsheets_conn = None
        st.error(f"Error de conexión: {e}")
        return None

def cargar_datos():
    try:
        sh = conectar_google()
        if sh is None: return None, None, None, None, None, None

        ws_c   = sh.worksheet("clientes")
        datos_c = ws_c.get_all_records()
        df_c   = pd.DataFrame(datos_c) if datos_c else pd.DataFrame(columns=COL_CLIENTES)

        ws_v   = sh.worksheet("viajes")
        datos_v = ws_v.get_all_records()
        df_v   = pd.DataFrame(datos_v) if datos_v else pd.DataFrame(columns=COL_VIAJES)
        df_v['Importe'] = pd.to_numeric(df_v['Importe'], errors='coerce').fillna(0)

        try:
            ws_p   = sh.worksheet("presupuestos")
            datos_p = ws_p.get_all_records()
            df_p   = pd.DataFrame(datos_p) if datos_p else pd.DataFrame(columns=COL_PRESUPUESTOS)
            df_p['Importe'] = pd.to_numeric(df_p['Importe'], errors='coerce').fillna(0)
        except:
            df_p = pd.DataFrame(columns=COL_PRESUPUESTOS)

        try:
            ws_t   = sh.worksheet("tesoreria")
            datos_t = ws_t.get_all_records()
            df_t   = pd.DataFrame(datos_t) if datos_t else pd.DataFrame(columns=COL_TESORERIA)
            df_t['Monto'] = pd.to_numeric(df_t['Monto'], errors='coerce').fillna(0)
            # Compatibilidad: si la hoja existente no tiene columna "Forma", la agrega vacía
            if 'Forma' not in df_t.columns:
                df_t.insert(3, 'Forma', '-')
        except:
            df_t = pd.DataFrame(columns=COL_TESORERIA)

        try:
            ws_prov   = sh.worksheet("proveedores")
            datos_prov = ws_prov.get_all_records()
            df_prov   = pd.DataFrame(datos_prov) if datos_prov else pd.DataFrame(columns=COL_PROVEEDORES)
            for col in ["CBU", "Alias"]:
                if col not in df_prov.columns:
                    df_prov[col] = "-"
        except:
            df_prov = pd.DataFrame(columns=COL_PROVEEDORES)

        try:
            ws_com   = sh.worksheet("compras")
            datos_com = ws_com.get_all_records()
            df_com   = pd.DataFrame(datos_com) if datos_com else pd.DataFrame(columns=COL_COMPRAS)
            for c in ["Neto 21", "Neto 10.5", "Ret IVA", "Ret Ganancia", "Ret IIBB", "No Gravados", "Total"]:
                df_com[c] = pd.to_numeric(df_com[c], errors='coerce').fillna(0)
        except:
            df_com = pd.DataFrame(columns=COL_COMPRAS)

        try:
            ws_ce    = sh.worksheet("cheques_emitidos")
            datos_ce = ws_ce.get_all_records()
            df_ce    = pd.DataFrame(datos_ce) if datos_ce else pd.DataFrame(columns=COL_CHEQ_EMITIDOS)
            df_ce['Importe'] = pd.to_numeric(df_ce['Importe'], errors='coerce').fillna(0)
        except:
            df_ce = pd.DataFrame(columns=COL_CHEQ_EMITIDOS)

        try:
            ws_cc    = sh.worksheet("cheques_cartera")
            datos_cc = ws_cc.get_all_records()
            df_cc    = pd.DataFrame(datos_cc) if datos_cc else pd.DataFrame(columns=COL_CHEQ_CARTERA)
            df_cc['Importe'] = pd.to_numeric(df_cc['Importe'], errors='coerce').fillna(0)
        except:
            df_cc = pd.DataFrame(columns=COL_CHEQ_CARTERA)

        try:
            ws_fac    = sh.worksheet("facturas")
            datos_fac = ws_fac.get_all_records()
            df_fac    = pd.DataFrame(datos_fac) if datos_fac else pd.DataFrame(columns=COL_FACTURAS)
            for col in ["Neto", "IVA", "No Gravado", "Total"]:
                df_fac[col] = pd.to_numeric(df_fac[col], errors='coerce').fillna(0)
        except:
            df_fac = pd.DataFrame(columns=COL_FACTURAS)

        return df_c, df_v, df_p, df_t, df_prov, df_com, df_ce, df_cc, df_fac
    except:
        return None, None, None, None, None, None, None, None, None

def guardar_datos(nombre_hoja, df, reintentos=3):
    """Guarda un DataFrame en Google Sheets. Reintenta hasta 3 veces ante errores de red."""
    import time
    ultimo_error = None
    for intento in range(reintentos):
        try:
            sh = conectar_google()
            if sh is None: return False
            try:
                ws = sh.worksheet(nombre_hoja)
            except gspread.exceptions.WorksheetNotFound:
                ws = sh.add_worksheet(title=nombre_hoja, rows=2000, cols=25)
            df_save = df.fillna("-").copy()
            datos   = [df_save.columns.values.tolist()] + df_save.astype(str).values.tolist()
            ws.clear()
            ws.update(datos)
            return True
        except Exception as e:
            ultimo_error = e
            # Limpiar conexión cacheada para forzar reconexión en el próximo intento
            st.session_state.gsheets_conn = None
            if intento < reintentos - 1:
                time.sleep(1.5)
    st.error(f"❌ Error al guardar '{nombre_hoja}' tras {reintentos} intentos: {ultimo_error}")
    return False


def guardar_tesoreria_y_compras():
    """Guarda tesoreria y compras atomicamente. Muestra error claro si alguna falla."""
    ok_t = guardar_datos("tesoreria", st.session_state.tesoreria)
    ok_c = guardar_datos("compras",   st.session_state.compras)
    if not ok_t:
        st.error("CRITICO: No se pudo guardar TESORERIA. Revisa la conexion y vuelve a intentar.")
    if not ok_c:
        st.error("CRITICO: No se pudo guardar COMPRAS (cuenta corriente proveedor). Revisa la conexion.")
    return ok_t and ok_c


def guardar_tesoreria_rerun(msg_key=None, msg_texto=None):
    """Guarda tesoreria. Si falla, muestra error y NO hace rerun. Si ok, guarda msg y hace rerun."""
    ok = guardar_datos("tesoreria", st.session_state.tesoreria)
    if ok:
        if msg_key and msg_texto:
            st.session_state[msg_key] = msg_texto
        st.rerun()
    else:
        st.error("❌ No se pudo guardar el movimiento. Revisá la conexión e intentá de nuevo.")

# =========================================================
# --- FUNCIONES PARA REPORTES HTML PROFESIONALES ---
# =========================================================

def generar_html_resumen(cliente, df, saldo):
    tabla_html = df.to_html(index=False, classes='tabla')
    return f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; color: #333; padding: 20px; }}
            .header-table {{ width: 100%; border-bottom: 4px solid #5e2d61; margin-bottom: 20px; }}
            .empresa-name {{ color: #5e2d61; font-size: 26px; font-weight: bold; margin: 0; }}
            .sub-title {{ color: #666; font-size: 14px; margin-top: 5px; }}
            .tabla {{ width: 100%; border-collapse: collapse; margin-top: 20px; background-color: white; }}
            .tabla th {{ background-color: #f8f9fa; color: #5e2d61; border-bottom: 2px solid #5e2d61; padding: 12px; text-align: left; font-size: 13px; }}
            .tabla td {{ border-bottom: 1px solid #eee; padding: 10px; font-size: 12px; }}
            .footer-resumen {{ margin-top: 30px; padding: 15px; background: #5e2d61; color: white; border-radius: 8px; text-align: right; }}
            .total-label {{ font-size: 14px; opacity: 0.9; }}
            .total-monto {{ font-size: 22px; font-weight: bold; display: block; }}
        </style>
    </head>
    <body>
        <table class="header-table">
            <tr>
                <td style="vertical-align:middle;">
                    <img src="{LOGO_B64}" style="height:60px; margin-right:15px; vertical-align:middle;">
                    <span style="vertical-align:middle;">
                        <p class="empresa-name">CHACABUCO NOROESTE TOUR S.R.L.</p>
                        <p class="sub-title">Desde 1996 viajando con vos | CHACAGEST Software System</p>
                    </span>
                </td>
                <td style="text-align: right; vertical-align:middle;">
                    <p><b>ESTADO DE CUENTA</b><br>Emisión: {{date.today()}}</p>
                </td>
            </tr>
        </table>
        <div style="margin-bottom: 20px;">
            <p><b>CLIENTE:</b> {cliente}</p>
        </div>
        {tabla_html}
        <div class="footer-resumen">
            <span class="total-label">SALDO TOTAL PENDIENTE</span>
            <span class="total-monto">$ {saldo:,.2f}</span>
        </div>
    </body>
    </html>
    """

def generar_html_cta_cte_general(tipo, df_resumen, fecha_emision):
    """Genera HTML para PDF de cuenta corriente general (clientes o proveedores)."""
    filas_html = ""
    total_general = 0
    for _, row in df_resumen.iterrows():
        nombre = row.get('Cliente', row.get('Proveedor', '-'))
        saldo  = round(float(row.get('Importe', row.get('Total', 0))), 2)
        if abs(saldo) <= 0.01:   # excluir saldos en cero o centavos residuales
            continue
        total_general += saldo
        color = "#e74c3c" if saldo < 0 else "#2ecc71"
        filas_html += f"""
        <tr>
            <td>{nombre}</td>
            <td style="text-align:right;font-weight:bold;color:{color};">$ {saldo:,.2f}</td>
        </tr>"""
    color_total = "#e74c3c" if total_general < 0 else "#2ecc71"
    return f"""<html><head><style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #333; padding: 24px; font-size: 13px; }}
        .logo-header {{ width:100%; border-bottom:4px solid #5e2d61; margin-bottom:20px; padding-bottom:10px; }}
        .empresa {{ color:#5e2d61; font-size:22px; font-weight:bold; }}
        .sub {{ color:#888; font-size:12px; }}
        h2 {{ color:#5e2d61; margin-top:0; }}
        table {{ width:100%; border-collapse:collapse; margin-top:16px; }}
        th {{ background:#5e2d61; color:white; padding:10px; text-align:left; font-size:12px; }}
        td {{ border-bottom:1px solid #eee; padding:9px 10px; }}
        tr:nth-child(even) td {{ background:#fafafa; }}
        .total-box {{ margin-top:20px; background:#5e2d61; color:white; border-radius:8px;
                      padding:14px 20px; display:flex; justify-content:space-between; align-items:center; }}
        .total-num {{ font-size:22px; font-weight:bold; color:{color_total}; }}
        .fecha {{ color:#888; font-size:12px; text-align:right; }}
    </style></head><body>
    <table class="logo-header"><tr>
        <td style="vertical-align:middle;">
            <img src="{LOGO_B64}" style="height:55px;vertical-align:middle;margin-right:12px;">
            <span style="vertical-align:middle;">
                <div class="empresa">CHACABUCO NOROESTE TOUR S.R.L.</div>
                <div class="sub">Desde 1996 viajando con vos | CHACAGEST Software System</div>
            </span>
        </td>
        <td style="text-align:right;vertical-align:middle;">
            <div class="fecha"><b>Emisión:</b> {fecha_emision}</div>
        </td>
    </tr></table>
    <h2>📊 Estado General de {tipo}</h2>
    <table>
        <tr><th>{tipo[:-1] if tipo.endswith('s') else tipo}</th><th style="text-align:right;">Saldo</th></tr>
        {filas_html}
    </table>
    <div class="total-box">
        <span style="font-size:14px;">TOTAL GENERAL</span>
        <span class="total-num">$ {total_general:,.2f}</span>
    </div>
    </body></html>"""


    return f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; }}
            .recibo-box {{ border: 3px double #5e2d61; padding: 30px; position: relative; }}
            .header-recibo {{ display: flex; justify-content: space-between; border-bottom: 1px solid #ccc; padding-bottom: 15px; }}
            .monto-destacado {{ font-size: 28px; color: #5e2d61; font-weight: bold; background: #f0f2f6; padding: 10px 20px; border-radius: 5px; }}
            .cuerpo {{ margin-top: 30px; line-height: 2.0; font-size: 16px; }}
            .firma-box {{ margin-top: 60px; border-top: 1px solid #333; width: 200px; text-align: center; float: right; font-size: 12px; }}
            .afip-ref {{ color: #777; font-size: 12px; font-style: italic; }}
        </style>
    </head>
    <body>
        <div class="recibo-box">
            <div class="header-recibo">
                <div style="display:flex;align-items:center;gap:12px;">
                    <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2wCEAAkGBxISERUQEBMVEhUWEhUVEBYREhAXFRUXGBIWFxUVFRUYHSggGBslHxUVITEiJSkrLi4wGB8zODMsOCgtLisBCgoKDQ0NGQ8PFzcZFx8rKy0rMis3Ky0rLTcrNys3LTcrLSstLSstMC0tMDIrNystNy03LTY3Ny8rLSs3LTcrK//AABEIALcBEwMBIgACEQEDEQH/xAAcAAEAAgIDAQAAAAAAAAAAAAAABgcBBQMECAL/xABHEAACAQICBQcHCQUHBQAAAAAAAQIDEQQhBQYxQWEHEhMiUXGBMkJTVJHR0hQVFiNSk5ShwReSorHhM0NicnOC8EVjhKOz/8QAGgEBAQADAQEAAAAAAAAAAAAAAAECAwYEBf/EACQRAQACAgEDAwUAAAAAAAAAAAABAgMRBBIxUQUhIhMjQmFx/9oADAMBAAIRAxEAPwC8QAAAAAAAAAAAAAAAAAAAAAAAAcGLxcKUJVaslCEIuU5S2JJXbIsuU/RPra+6r/ABMCOaz6UcbUabae2bTat2RvxNbiuVHRag3DEqUrdVKnXze7zSHVddMFKTlKum27t8yr8IEi+WVfS1PvJ+8fLKvpKn3k/eR6GtmCbSVdXbslzal29iSXN25oli0FXtfmZWvnKCy9pYR1PllX0lT7yfvHy2r6Sp95P3nC0fVGk5yUYq7bshI3Or9OrVqXlUqOEc5deeb3LaS86mjsEqVNQj/ufa97O2RQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjmvms8dH4Sdd2dR9TDxfnVGnbLsWbfBAVxy3623a0ZQnkmpYtx35XhSfDZJr/AClRHJXrynOVSpJynKTlOT2uTd233s4woAbLV3QtTG4mnhaPlTl1pboRWcpvglfv2ATvkW1S+UV3jq0fqqLtRT2Tq73xjFfm+BbWs2keZHoovrS8q26PZ4nawmFo4HCxpU1zYUoKMFvk7b+1t5kOxNeU5Octrbb/AEQRxko1W0bZdPJZvKF+ze/E0uh8A61RR81Zza7OzvZO4RSVlklsRZkZRkAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbA+Kskk23ZJNtvYkt7PM3KRrW9I4tzg/qad4YdXyaum6lu2TS8Eix+W3W5UqXzdRf1lWN67XmUnfq98rPwXEo0AAAoegeR7VJ4TDPE1l9fXSdms6dPbGHe/KfeuwrHkt1dp4rFqriGlQotSkpf3k79SHdezfcu0vbTOmYKnzaUlKUsuq/JW9hGq1j0j0s+ZHyIP2y3s1EYtuyzb2foYJFqvo676aSyV1BPe97A3OhcAqNNR855zfHs8DYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGajWnT1PA4WpiquyK6sd85vKEFxba8Lm2Z535XdbvlmKeHpSvQoNxVtlSp58+5eSvF70BDNK6RqYmtUxFZ86dSTlJ3duCXBKyXBHUAChzYPCyq1I0qavKclGPf/Tb4HCWNyc6C5kPldRdaatRvug9sv938gJPoTRkcNQjRhuXWe+UnnKV/y8DvgykZaR2dG4N1aigu+T7F2k9oUVCKhHJJJIjeDx2EwNPnYqvRoynm+kqQi+EUm7vwOhjOVbRUNleVX/SpVH+bSTMRNwVrV5aNHryadeXfCMfybON8teC9DX9kPeBZwKvfLZg/QV//AF+8x+2zB+r1/wCD3gWiCrHy3YT1av7afvMftvwvq1f20veBagKpfLhhvVa/71L3mHy44b1Sv+/R94Frgqd8uOH9Ur/v0fefP7csP6nW8alH3gW0DQ6E03XxNCniI4V01Ujzoxq1YqaV3ZtJO11n4gDfAAAAAAAAAAAAAAAAAxc1Gs+sNHA4eWIrvJZQj505box4gRblf1teCwvQUZWr104xtthT8+pbt81cXfceeTY6w6bq43ETxVd9abyS2Qir82EeCXtua4AAZpwcmoxTk27RS2tvYkFbnVLQrxVdQf8AZw61V8L2UfFlwxikrLJJWXYkv6Gk0Ho+ngML9ZJRdudWl/its422L+pBtZ9b6mIbp0r06Ozb1p8Zdi4FRLdO67UKF4UvrprLqvqJ8Ze4g+k9bcXW21Ojj9mkuavb5X5mjsBs0zNttybbb8pttt97ebMAEUAAAwZAAAAAAAJfyY6qPH4tc9PoKLU67z62+FK/a9/YvAiuDws6tSFKlHnTnJRgu1t5LgeodRdWoaPwkMNHrS8qtO1nOpLOT7lsS7EgjexjZWSVlkrZWW5GDlAAAxJgZBpsdrXgKNSVKvi8PSqRtzoVK9KMldJq8W7rJp+JwfTjRnr+E/E0fiAkAI/9ONGev4T8TR+IfTjRnr+E/E0fiAkAI/8ATjRnr+E/E0fiH040Z6/hPxNH4gJACP8A040Z6/hPxNH4jq4nlG0VTV3jaUrei59T/wCaYEqMc4rXSXLRgIJ9BCtXe7qKnF+M3f8AIgOsfKxj8QnCi1hIPL6q7qtf6j2eCT4gW/rnrzhdHRfSS6Sra8KNNpzff9lcWefNa9Zq+kK3TYh7LqlCN+ZTT3RvteWb2s09Sbk3KTcm3eTk2232tvNvifIUAAAnHJ9oZK+NrWjGF+icu1K0p9y2eJGdX9ESxVaNKOSunUl9mN8337Uu8kuvml4wjHAYfqwhFdLb+Gn+r8AjT626xyxdS0bqjF9SOznf45Lt7Ow0AAUAAAEl1a1Ex+OtKjS5lPfVrNwp27VleXgn4Fl6E5FMPFJ4uvOtLfGlanC/B5yftAo8xc9O4Tk40VTtbB058avOqX71NtP2Gyjqno9KywWFS4Yah8IHlC4uesfotgPUsL+GofCPotgPUsL+GofCB5OuLnrH6LYD1LC/hqHwj6LYD1LC/hqHwgeTrhtHrH6LYD1LC/hqHwj6LYD1LC/hqHwgVnyIao/9TrRzd44VP2Tq97zSff2lxJHHh6EacYwhFRjFJRjFKMYpKyUYrJLgjlCAAAMj+u2skMBhJ4idnLyaMPt1H5K7ltfBM30pJK7aSW2+482cqGtrx+Lag/qKLcKGflNPrVPHdwsBE8di51qk61WXOnObnNva2/8An5HCAFAAAAAAAAAAAAAA58DhJ1qkaVOLlKTskv5t7lxOxofRFXE1OZRi39qTuox/zMtPV7V+lgqbs+dNpupUllxaXZFAazoqeisFJq0q0978+o1kv8kezhxKyq1XJuUm5Sk7yb2tva3xNzrdpr5VXck/q4XjSW6185d79xpAgAYuFclGlKclCEXOUpKMYxTbbbskki7dQeSmnSUcRpGKq1cnCi86dPeuevPlw2LvOxyRairD0447Ew+vqK9KMl/YwezLdOS9iaRZyCMRikkkklutsPqwAAAAAAAAAAAAAAAAAFY8tGtvyagsFRlatXT6Rp506W/xlsXBMoax6D03yTUMXXqYmvisS51Jc550rRW6MVzcklZHTXIlgvWMR7aXwAUQC+FyJ4H02I9tL4T6XIpgfTYj96n8IFCgvv8AYtgPS1/34fCJcjej0rupXtxqR9wFCAvhckei/S1fvY+4z+yTRXpan30RuGXTbwoYF9rkm0T6Wp9+jK5KdEfbm/8AyBuF6Z8KDMSkltaXeeg6eoGg6OclGX+pXlJfzO/o+Wh6ErYalQ5/munSUpZf42m14sm4X6WSY3FfZ5/wGgcTWt0dGdnslKLjG3beVrruJfobk+SfPxVTnf8Abp7P909/giwsXiHUm5y3/ktyRsMLq/WqQU04JNXSk5J29jMmto8LhYUoKnTioRWxRSSIjyi6d5kPklN9aavVs9kL+Twvb8ic6zYd4HDTxVaVPmxyilKV5SeUIJc3a3+pQmNxUqtSVWbvKcnKT/52ZLwA4QARQnnJHqk8bilXqq9ChJSnfZOpthDuTzfdxIbovR9TEVoUKMedOpJRgt1+19iWbfcepdVNBU8DhaeFpZ82PXlZXnN+VN8W/wBAjbxRkAAAAAAAAAAAAAAAAAAAAAAAAAAdXSODjWpSpT2SVmdowyd1iZidwonTGBnh606M73i8nd9aPmyXgdPnPtftZafKDoHpqPTU1epTV8lnKO9d+8qo+dmpNLfp3HpvJpyMMTr5R7Szd9oMA1bfQ6Y8FiUaDwXRw50l1pbeC3LhlZ+JqtB4HpJ86Xkx28XuXh+hKacHJqKV23ZJb7ns42PfzlznrfN19ik/1sNBaP6apn5Mc5fovEm6SS4HU0VgVRpqC27ZPte8hXK/rf8AI8N8noytXrppW206eyU+Dd7Lx7GetzKtuVvW75biuhpSvQoSlGNtlSpsnPilnFdqu95AwAoASnk51UekcWqc0+gp2niGr5xv1YX7ZNNdyYFj8iWqHRU/nGvG1SrHm4dPbGk9su+WXgl2lrJGKdNRSjFJJJKKSsklkkkfQQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfM43Kg140G8NX50V9XUvKHB+dH9fEuBmq1j0VHE0JUpZO14P7MtzNWWnXXT3+ncyeNmi34z3UifdGk5SUY7W7L/nYZxNCVOcqc1aUZNSXFG+1fwPNj0stsl1eEdt/E8OPHNraddzOZTBg+pve+zY4TDqnBQjsX5ve2SvVXR39/Jdqp/rI02isC61RQ3bZvsRPKcVFKKVkkkuCPpxGo1Dhb3te02t3l1NMaSp4ajUxFZ82FOLlJ78tyW9vYu88s6yacq43EzxVbJzfVis1CK8mC7l7Xdk95bNbunrfIKL+qou9drz6u6PdH+b4Z1eGAAAr7oUpTlGEE5SlJRgltcpO0Uu92PT2oWq0dH4SNFWdST59eSt1ptbO5K0VwXErrkP1S50npOtHKLccKpLbJXU6q7s4p95dSCMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGGjJ8zkkm3kkrtgQbXnV6E6tPEJpX6tWO+Vlk1+a8ToJbku5I72mMf01Ry81ZQXDt72dKE3Fpp2azTJWsRMy25M+S9K0tO4r2TbQeA6Gmr+VLOfuNFym61rR+Dbg109VOGHXY2s6jXZFZ+w6PzrX9LP2mp0vgKWKkp4qCrSiubF1M7LsXZ/RGWmpRs5Nttttttttttt5tt9pguV6r4L1en4J+8k+geTzR7p8+thKUnLOKaeS9u8g85m71N1dnj8XTw0E1FvnVp7oU1m5d7yS4tHoT9nmivUqP7r95s9C6t4TCc54XD06LnbnuEbN22JsDuaPwcKNKFGlFQhTioQitiilZI7IAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA02scqjiqdON+dnN3isuzNgARr5qrfY/ih7x81VvsfxQ95kFRj5qrfY/ih7x81VvsfxQ95kFkdvRWhZyqLpY2is3nHPPJZMmKAMVZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAf/2Q==" style="height:65px;vertical-align:middle;margin-right:12px;">
                    <div>
                        <b style="font-size: 20px;">CHACABUCO NOROESTE TOUR S.R.L.</b><br>
                        <span>CUIT 30-71114824-4 - C.P. 6740 - Chacabuco, Bs. As.</span>
                    </div>
                </div>
                <div class="monto-destacado">$ {abs(data['Monto']):,.2f}</div>
            </div>
            <h2 style="text-align: center; text-decoration: underline;">RECIBO DE PAGO</h2>
            <div class="cuerpo">
                Recibimos de <b>{data['Cliente/Proveedor']}</b> la cantidad de pesos
                <span style="text-transform: uppercase;"><b>{abs(data['Monto']):,.2f}</b></span>
                en concepto de: <b>{data['Concepto']}</b>.<br>
                Realizado mediante: <b>{data['Caja/Banco']}</b>.<br>
                <span class="afip-ref">En Concepto de: {data['Ref AFIP']}</span>
            </div>
            <div style="margin-top: 40px;"><b>FECHA:</b> {data['Fecha']}</div>
            <div class="firma-box">Firma y Sello Responsable</div>
            <div style="clear: both;"></div>
            <p style="text-align: center; font-size: 9px; color: #aaa; margin-top: 50px;">Generado por CHACAGEST.</p>
        </div>
    </body>
    </html>
    """

def generar_html_orden_pago(data):
    return f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; padding: 20px; color: #333; }}
            .op-container {{ border: 2px solid #d35400; border-radius: 10px; padding: 30px; background-color: #fff; }}
            .header-op {{ border-bottom: 2px solid #d35400; padding-bottom: 15px; margin-bottom: 25px; }}
            .titulo-doc {{ font-size: 24px; font-weight: bold; text-align: center; margin: 20px 0; color: #444; }}
            .monto-op {{ background: #fff4e6; border: 1px dashed #d35400; padding: 15px; font-size: 22px; font-weight: bold; text-align: center; color: #d35400; margin: 20px 0; }}
            .detalle-table {{ width: 100%; margin-top: 20px; line-height: 1.8; }}
        </style>
    </head>
    <body>
        <div class="op-container">
            <div class="header-op" style="display:flex;align-items:center;gap:14px;">
                <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2wCEAAkGBxISERUQEBMVEhUWEhUVEBYREhAXFRUXGBIWFxUVFRUYHSggGBslHxUVITEiJSkrLi4wGB8zODMsOCgtLisBCgoKDQ0NGQ8PFzcZFx8rKy0rMis3Ky0rLTcrNys3LTcrLSstLSstMC0tMDIrNystNy03LTY3Ny8rLSs3LTcrK//AABEIALcBEwMBIgACEQEDEQH/xAAcAAEAAgIDAQAAAAAAAAAAAAAABgcBBQMECAL/xABHEAACAQICBQcHCQUHBQAAAAAAAQIDEQQhBQYxQWEHEhMiUXGBMkJTVJHR0hQVFiNSk5ShwReSorHhM0NicnOC8EVjhKOz/8QAGgEBAQADAQEAAAAAAAAAAAAAAAECAwYEBf/EACQRAQACAgEDAwUAAAAAAAAAAAABAgMRBBIxUQUhIhMjQmFx/9oADAMBAAIRAxEAPwC8QAAAAAAAAAAAAAAAAAAAAAAAAcGLxcKUJVaslCEIuU5S2JJXbIsuU/RPra+6r/ABMCOaz6UcbUabae2bTat2RvxNbiuVHRag3DEqUrdVKnXze7zSHVddMFKTlKum27t8yr8IEi+WVfS1PvJ+8fLKvpKn3k/eR6GtmCbSVdXbslzal29iSXN25oli0FXtfmZWvnKCy9pYR1PllX0lT7yfvHy2r6Sp95P3nC0fVGk5yUYq7bshI3Or9OrVqXlUqOEc5deeb3LaS86mjsEqVNQj/ufa97O2RQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjmvms8dH4Sdd2dR9TDxfnVGnbLsWbfBAVxy3623a0ZQnkmpYtx35XhSfDZJr/AClRHJXrynOVSpJynKTlOT2uTd233s4woAbLV3QtTG4mnhaPlTl1pboRWcpvglfv2ATvkW1S+UV3jq0fqqLtRT2Tq73xjFfm+BbWs2keZHoovrS8q26PZ4nawmFo4HCxpU1zYUoKMFvk7b+1t5kOxNeU5Octrbb/AEQRxko1W0bZdPJZvKF+ze/E0uh8A61RR81Zza7OzvZO4RSVlklsRZkZRkAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbA+Kskk23ZJNtvYkt7PM3KRrW9I4tzg/qad4YdXyaum6lu2TS8Eix+W3W5UqXzdRf1lWN67XmUnfq98rPwXEo0AAAoegeR7VJ4TDPE1l9fXSdms6dPbGHe/KfeuwrHkt1dp4rFqriGlQotSkpf3k79SHdezfcu0vbTOmYKnzaUlKUsuq/JW9hGq1j0j0s+ZHyIP2y3s1EYtuyzb2foYJFqvo676aSyV1BPe97A3OhcAqNNR855zfHs8DYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGajWnT1PA4WpiquyK6sd85vKEFxba8Lm2Z535XdbvlmKeHpSvQoNxVtlSp58+5eSvF70BDNK6RqYmtUxFZ86dSTlJ3duCXBKyXBHUAChzYPCyq1I0qavKclGPf/Tb4HCWNyc6C5kPldRdaatRvug9sv938gJPoTRkcNQjRhuXWe+UnnKV/y8DvgykZaR2dG4N1aigu+T7F2k9oUVCKhHJJJIjeDx2EwNPnYqvRoynm+kqQi+EUm7vwOhjOVbRUNleVX/SpVH+bSTMRNwVrV5aNHryadeXfCMfybON8teC9DX9kPeBZwKvfLZg/QV//AF+8x+2zB+r1/wCD3gWiCrHy3YT1av7afvMftvwvq1f20veBagKpfLhhvVa/71L3mHy44b1Sv+/R94Frgqd8uOH9Ur/v0fefP7csP6nW8alH3gW0DQ6E03XxNCniI4V01Ujzoxq1YqaV3ZtJO11n4gDfAAAAAAAAAAAAAAAAAxc1Gs+sNHA4eWIrvJZQj505box4gRblf1teCwvQUZWr104xtthT8+pbt81cXfceeTY6w6bq43ETxVd9abyS2Qir82EeCXtua4AAZpwcmoxTk27RS2tvYkFbnVLQrxVdQf8AZw61V8L2UfFlwxikrLJJWXYkv6Gk0Ho+ngML9ZJRdudWl/its422L+pBtZ9b6mIbp0r06Ozb1p8Zdi4FRLdO67UKF4UvrprLqvqJ8Ze4g+k9bcXW21Ojj9mkuavb5X5mjsBs0zNttybbb8pttt97ebMAEUAAAwZAAAAAAAJfyY6qPH4tc9PoKLU67z62+FK/a9/YvAiuDws6tSFKlHnTnJRgu1t5LgeodRdWoaPwkMNHrS8qtO1nOpLOT7lsS7EgjexjZWSVlkrZWW5GDlAAAxJgZBpsdrXgKNSVKvi8PSqRtzoVK9KMldJq8W7rJp+JwfTjRnr+E/E0fiAkAI/9ONGev4T8TR+IfTjRnr+E/E0fiAkAI/8ATjRnr+E/E0fiH040Z6/hPxNH4gJACP8A040Z6/hPxNH4jq4nlG0VTV3jaUrei59T/wCaYEqMc4rXSXLRgIJ9BCtXe7qKnF+M3f8AIgOsfKxj8QnCi1hIPL6q7qtf6j2eCT4gW/rnrzhdHRfSS6Sra8KNNpzff9lcWefNa9Zq+kK3TYh7LqlCN+ZTT3RvteWb2s09Sbk3KTcm3eTk2232tvNvifIUAAAnHJ9oZK+NrWjGF+icu1K0p9y2eJGdX9ESxVaNKOSunUl9mN8337Uu8kuvml4wjHAYfqwhFdLb+Gn+r8AjT626xyxdS0bqjF9SOznf45Lt7Ow0AAUAAAEl1a1Ex+OtKjS5lPfVrNwp27VleXgn4Fl6E5FMPFJ4uvOtLfGlanC/B5yftAo8xc9O4Tk40VTtbB058avOqX71NtP2Gyjqno9KywWFS4Yah8IHlC4uesfotgPUsL+GofCPotgPUsL+GofCB5OuLnrH6LYD1LC/hqHwj6LYD1LC/hqHwgeTrhtHrH6LYD1LC/hqHwj6LYD1LC/hqHwgVnyIao/9TrRzd44VP2Tq97zSff2lxJHHh6EacYwhFRjFJRjFKMYpKyUYrJLgjlCAAAMj+u2skMBhJ4idnLyaMPt1H5K7ltfBM30pJK7aSW2+482cqGtrx+Lag/qKLcKGflNPrVPHdwsBE8di51qk61WXOnObnNva2/8An5HCAFAAAAAAAAAAAAAA58DhJ1qkaVOLlKTskv5t7lxOxofRFXE1OZRi39qTuox/zMtPV7V+lgqbs+dNpupUllxaXZFAazoqeisFJq0q0978+o1kv8kezhxKyq1XJuUm5Sk7yb2tva3xNzrdpr5VXck/q4XjSW6185d79xpAgAYuFclGlKclCEXOUpKMYxTbbbskki7dQeSmnSUcRpGKq1cnCi86dPeuevPlw2LvOxyRairD0447Ew+vqK9KMl/YwezLdOS9iaRZyCMRikkkklutsPqwAAAAAAAAAAAAAAAAAFY8tGtvyagsFRlatXT6Rp506W/xlsXBMoax6D03yTUMXXqYmvisS51Jc550rRW6MVzcklZHTXIlgvWMR7aXwAUQC+FyJ4H02I9tL4T6XIpgfTYj96n8IFCgvv8AYtgPS1/34fCJcjej0rupXtxqR9wFCAvhckei/S1fvY+4z+yTRXpan30RuGXTbwoYF9rkm0T6Wp9+jK5KdEfbm/8AyBuF6Z8KDMSkltaXeeg6eoGg6OclGX+pXlJfzO/o+Wh6ErYalQ5/munSUpZf42m14sm4X6WSY3FfZ5/wGgcTWt0dGdnslKLjG3beVrruJfobk+SfPxVTnf8Abp7P909/giwsXiHUm5y3/ktyRsMLq/WqQU04JNXSk5J29jMmto8LhYUoKnTioRWxRSSIjyi6d5kPklN9aavVs9kL+Twvb8ic6zYd4HDTxVaVPmxyilKV5SeUIJc3a3+pQmNxUqtSVWbvKcnKT/52ZLwA4QARQnnJHqk8bilXqq9ChJSnfZOpthDuTzfdxIbovR9TEVoUKMedOpJRgt1+19iWbfcepdVNBU8DhaeFpZ82PXlZXnN+VN8W/wBAjbxRkAAAAAAAAAAAAAAAAAAAAAAAAAAdXSODjWpSpT2SVmdowyd1iZidwonTGBnh606M73i8nd9aPmyXgdPnPtftZafKDoHpqPTU1epTV8lnKO9d+8qo+dmpNLfp3HpvJpyMMTr5R7Szd9oMA1bfQ6Y8FiUaDwXRw50l1pbeC3LhlZ+JqtB4HpJ86Xkx28XuXh+hKacHJqKV23ZJb7ns42PfzlznrfN19ik/1sNBaP6apn5Mc5fovEm6SS4HU0VgVRpqC27ZPte8hXK/rf8AI8N8noytXrppW206eyU+Dd7Lx7GetzKtuVvW75biuhpSvQoSlGNtlSpsnPilnFdqu95AwAoASnk51UekcWqc0+gp2niGr5xv1YX7ZNNdyYFj8iWqHRU/nGvG1SrHm4dPbGk9su+WXgl2lrJGKdNRSjFJJJKKSsklkkkfQQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfM43Kg140G8NX50V9XUvKHB+dH9fEuBmq1j0VHE0JUpZO14P7MtzNWWnXXT3+ncyeNmi34z3UifdGk5SUY7W7L/nYZxNCVOcqc1aUZNSXFG+1fwPNj0stsl1eEdt/E8OPHNraddzOZTBg+pve+zY4TDqnBQjsX5ve2SvVXR39/Jdqp/rI02isC61RQ3bZvsRPKcVFKKVkkkuCPpxGo1Dhb3te02t3l1NMaSp4ajUxFZ82FOLlJ78tyW9vYu88s6yacq43EzxVbJzfVis1CK8mC7l7Xdk95bNbunrfIKL+qou9drz6u6PdH+b4Z1eGAAAr7oUpTlGEE5SlJRgltcpO0Uu92PT2oWq0dH4SNFWdST59eSt1ptbO5K0VwXErrkP1S50npOtHKLccKpLbJXU6q7s4p95dSCMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGGjJ8zkkm3kkrtgQbXnV6E6tPEJpX6tWO+Vlk1+a8ToJbku5I72mMf01Ry81ZQXDt72dKE3Fpp2azTJWsRMy25M+S9K0tO4r2TbQeA6Gmr+VLOfuNFym61rR+Dbg109VOGHXY2s6jXZFZ+w6PzrX9LP2mp0vgKWKkp4qCrSiubF1M7LsXZ/RGWmpRs5Nttttttttttt5tt9pguV6r4L1en4J+8k+geTzR7p8+thKUnLOKaeS9u8g85m71N1dnj8XTw0E1FvnVp7oU1m5d7yS4tHoT9nmivUqP7r95s9C6t4TCc54XD06LnbnuEbN22JsDuaPwcKNKFGlFQhTioQitiilZI7IAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA02scqjiqdON+dnN3isuzNgARr5qrfY/ih7x81VvsfxQ95kFRj5qrfY/ih7x81VvsfxQ95kFkdvRWhZyqLpY2is3nHPPJZMmKAMVZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAf/2Q==" style="height:65px;vertical-align:middle;margin-right:12px;">
                <div>
                    <h2 style="margin:0; color: #d35400;">CHACABUCO NOROESTE TOUR S.R.L.</h2>
                    <small>Desde 1996, viajando con vos | CHACAGEST Software System</small>
                </div>
            </div>
            <div class="titulo-doc">ORDEN DE PAGO A PROVEEDOR</div>
            <table class="detalle-table">
                <tr><td><b>BENEFICIARIO:</b></td><td>{data['Proveedor']}</td></tr>
                <tr><td><b>FECHA:</b></td><td>{data['Fecha']}</td></tr>
                <tr><td><b>CONCEPTO:</b></td><td>{data['Concepto']}</td></tr>
                <tr><td><b>REFERENCIA:</b></td><td>{data['Ref AFIP']}</td></tr>
            </table>
            <div class="monto-op">TOTAL PAGADO: $ {abs(data['Monto']):,.2f}</div>
            <div style="margin-top: 60px; border-top: 1px solid #333; width: 220px; text-align: center; float: right;">Recibí conforme</div>
            <div style="clear: both;"></div>
        </div>
    </body>
    </html>
    """

def generar_html_recibo(data):
    return f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; padding: 20px; color: #333; }}
            .rec-container {{ border: 2px solid #5e2d61; border-radius: 10px; padding: 30px; background-color: #fff; }}
            .header-rec {{ border-bottom: 2px solid #5e2d61; padding-bottom: 15px; margin-bottom: 25px; display:flex; align-items:center; gap:14px; }}
            .titulo-doc {{ font-size: 24px; font-weight: bold; text-align: center; margin: 20px 0; color: #5e2d61; letter-spacing: 2px; }}
            .monto-rec {{ background: #f3eaf4; border: 1px dashed #5e2d61; padding: 15px; font-size: 26px; font-weight: bold; text-align: center; color: #5e2d61; margin: 20px 0; border-radius: 8px; }}
            .detalle-table {{ width: 100%; margin-top: 20px; line-height: 2.0; border-collapse: collapse; }}
            .detalle-table td {{ padding: 4px 8px; }}
            .detalle-table td:first-child {{ font-weight: bold; width: 180px; color: #555; }}
            .firma-box {{ margin-top: 60px; display: flex; justify-content: space-between; }}
            .firma-linea {{ border-top: 1px solid #333; width: 200px; text-align: center; padding-top: 6px; font-size: 13px; color: #555; }}
            .footer-rec {{ margin-top: 30px; font-size: 11px; color: #aaa; text-align: center; border-top: 1px solid #eee; padding-top: 12px; }}
        </style>
    </head>
    <body>
        <div class="rec-container">
            <div class="header-rec">
                <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2wCEAAkGBxISERUQEBMVEhUWEhUVEBYREhAXFRUXGBIWFxUVFRUYHSggGBslHxUVITEiJSkrLi4wGB8zODMsOCgtLisBCgoKDQ0NGQ8PFzcZFx8rKy0rMis3Ky0rLTcrNys3LTcrLSstLSstMC0tMDIrNystNy03LTY3Ny8rLSs3LTcrK//AABEIALcBEwMBIgACEQEDEQH/xAAcAAEAAgIDAQAAAAAAAAAAAAAABgcBBQMECAL/xABHEAACAQICBQcHCQUHBQAAAAAAAQIDEQQhBQYxQWEHEhMiUXGBMkJTVJHR0hQVFiNSk5ShwReSorHhM0NicnOC8EVjhKOz/8QAGgEBAQADAQEAAAAAAAAAAAAAAAECAwYEBf/EACQRAQACAgEDAwUAAAAAAAAAAAABAgMRBBIxUQUhIhMjQmFx/9oADAMBAAIRAxEAPwC8QAAAAAAAAAAAAAAAAAAAAAAAAcGLxcKUJVaslCEIuU5S2JJXbIsuU/RPra+6r/ABMCOaz6UcbUabae2bTat2RvxNbiuVHRag3DEqUrdVKnXze7zSHVddMFKTlKum27t8yr8IEi+WVfS1PvJ+8fLKvpKn3k/eR6GtmCbSVdXbslzal29iSXN25oli0FXtfmZWvnKCy9pYR1PllX0lT7yfvHy2r6Sp95P3nC0fVGk5yUYq7bshI3Or9OrVqXlUqOEc5deeb3LaS86mjsEqVNQj/ufa97O2RQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjmvms8dH4Sdd2dR9TDxfnVGnbLsWbfBAVxy3623a0ZQnkmpYtx35XhSfDZJr/AClRHJXrynOVSpJynKTlOT2uTd233s4woAbLV3QtTG4mnhaPlTl1pboRWcpvglfv2ATvkW1S+UV3jq0fqqLtRT2Tq73xjFfm+BbWs2keZHoovrS8q26PZ4nawmFo4HCxpU1zYUoKMFvk7b+1t5kOxNeU5Octrbb/AEQRxko1W0bZdPJZvKF+ze/E0uh8A61RR81Zza7OzvZO4RSVlklsRZkZRkAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbA+Kskk23ZJNtvYkt7PM3KRrW9I4tzg/qad4YdXyaum6lu2TS8Eix+W3W5UqXzdRf1lWN67XmUnfq98rPwXEo0AAAoegeR7VJ4TDPE1l9fXSdms6dPbGHe/KfeuwrHkt1dp4rFqriGlQotSkpf3k79SHdezfcu0vbTOmYKnzaUlKUsuq/JW9hGq1j0j0s+ZHyIP2y3s1EYtuyzb2foYJFqvo676aSyV1BPe97A3OhcAqNNR855zfHs8DYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGajWnT1PA4WpiquyK6sd85vKEFxba8Lm2Z535XdbvlmKeHpSvQoNxVtlSp58+5eSvF70BDNK6RqYmtUxFZ86dSTlJ3duCXBKyXBHUAChzYPCyq1I0qavKclGPf/Tb4HCWNyc6C5kPldRdaatRvug9sv938gJPoTRkcNQjRhuXWe+UnnKV/y8DvgykZaR2dG4N1aigu+T7F2k9oUVCKhHJJJIjeDx2EwNPnYqvRoynm+kqQi+EUm7vwOhjOVbRUNleVX/SpVH+bSTMRNwVrV5aNHryadeXfCMfybON8teC9DX9kPeBZwKvfLZg/QV//AF+8x+2zB+r1/wCD3gWiCrHy3YT1av7afvMftvwvq1f20veBagKpfLhhvVa/71L3mHy44b1Sv+/R94Frgqd8uOH9Ur/v0fefP7csP6nW8alH3gW0DQ6E03XxNCniI4V01Ujzoxq1YqaV3ZtJO11n4gDfAAAAAAAAAAAAAAAAAxc1Gs+sNHA4eWIrvJZQj505box4gRblf1teCwvQUZWr104xtthT8+pbt81cXfceeTY6w6bq43ETxVd9abyS2Qir82EeCXtua4AAZpwcmoxTk27RS2tvYkFbnVLQrxVdQf8AZw61V8L2UfFlwxikrLJJWXYkv6Gk0Ho+ngML9ZJRdudWl/its422L+pBtZ9b6mIbp0r06Ozb1p8Zdi4FRLdO67UKF4UvrprLqvqJ8Ze4g+k9bcXW21Ojj9mkuavb5X5mjsBs0zNttybbb8pttt97ebMAEUAAAwZAAAAAAAJfyY6qPH4tc9PoKLU67z62+FK/a9/YvAiuDws6tSFKlHnTnJRgu1t5LgeodRdWoaPwkMNHrS8qtO1nOpLOT7lsS7EgjexjZWSVlkrZWW5GDlAAAxJgZBpsdrXgKNSVKvi8PSqRtzoVK9KMldJq8W7rJp+JwfTjRnr+E/E0fiAkAI/9ONGev4T8TR+IfTjRnr+E/E0fiAkAI/8ATjRnr+E/E0fiH040Z6/hPxNH4gJACP8A040Z6/hPxNH4jq4nlG0VTV3jaUrei59T/wCaYEqMc4rXSXLRgIJ9BCtXe7qKnF+M3f8AIgOsfKxj8QnCi1hIPL6q7qtf6j2eCT4gW/rnrzhdHRfSS6Sra8KNNpzff9lcWefNa9Zq+kK3TYh7LqlCN+ZTT3RvteWb2s09Sbk3KTcm3eTk2232tvNvifIUAAAnHJ9oZK+NrWjGF+icu1K0p9y2eJGdX9ESxVaNKOSunUl9mN8337Uu8kuvml4wjHAYfqwhFdLb+Gn+r8AjT626xyxdS0bqjF9SOznf45Lt7Ow0AAUAAAEl1a1Ex+OtKjS5lPfVrNwp27VleXgn4Fl6E5FMPFJ4uvOtLfGlanC/B5yftAo8xc9O4Tk40VTtbB058avOqX71NtP2Gyjqno9KywWFS4Yah8IHlC4uesfotgPUsL+GofCPotgPUsL+GofCB5OuLnrH6LYD1LC/hqHwj6LYD1LC/hqHwgeTrhtHrH6LYD1LC/hqHwj6LYD1LC/hqHwgVnyIao/9TrRzd44VP2Tq97zSff2lxJHHh6EacYwhFRjFJRjFKMYpKyUYrJLgjlCAAAMj+u2skMBhJ4idnLyaMPt1H5K7ltfBM30pJK7aSW2+482cqGtrx+Lag/qKLcKGflNPrVPHdwsBE8di51qk61WXOnObnNva2/8An5HCAFAAAAAAAAAAAAAA58DhJ1qkaVOLlKTskv5t7lxOxofRFXE1OZRi39qTuox/zMtPV7V+lgqbs+dNpupUllxaXZFAazoqeisFJq0q0978+o1kv8kezhxKyq1XJuUm5Sk7yb2tva3xNzrdpr5VXck/q4XjSW6185d79xpAgAYuFclGlKclCEXOUpKMYxTbbbskki7dQeSmnSUcRpGKq1cnCi86dPeuevPlw2LvOxyRairD0447Ew+vqK9KMl/YwezLdOS9iaRZyCMRikkkklutsPqwAAAAAAAAAAAAAAAAAFY8tGtvyagsFRlatXT6Rp506W/xlsXBMoax6D03yTUMXXqYmvisS51Jc550rRW6MVzcklZHTXIlgvWMR7aXwAUQC+FyJ4H02I9tL4T6XIpgfTYj96n8IFCgvv8AYtgPS1/34fCJcjej0rupXtxqR9wFCAvhckei/S1fvY+4z+yTRXpan30RuGXTbwoYF9rkm0T6Wp9+jK5KdEfbm/8AyBuF6Z8KDMSkltaXeeg6eoGg6OclGX+pXlJfzO/o+Wh6ErYalQ5/munSUpZf42m14sm4X6WSY3FfZ5/wGgcTWt0dGdnslKLjG3beVrruJfobk+SfPxVTnf8Abp7P909/giwsXiHUm5y3/ktyRsMLq/WqQU04JNXSk5J29jMmto8LhYUoKnTioRWxRSSIjyi6d5kPklN9aavVs9kL+Twvb8ic6zYd4HDTxVaVPmxyilKV5SeUIJc3a3+pQmNxUqtSVWbvKcnKT/52ZLwA4QARQnnJHqk8bilXqq9ChJSnfZOpthDuTzfdxIbovR9TEVoUKMedOpJRgt1+19iWbfcepdVNBU8DhaeFpZ82PXlZXnN+VN8W/wBAjbxRkAAAAAAAAAAAAAAAAAAAAAAAAAAdXSODjWpSpT2SVmdowyd1iZidwonTGBnh606M73i8nd9aPmyXgdPnPtftZafKDoHpqPTU1epTV8lnKO9d+8qo+dmpNLfp3HpvJpyMMTr5R7Szd9oMA1bfQ6Y8FiUaDwXRw50l1pbeC3LhlZ+JqtB4HpJ86Xkx28XuXh+hKacHJqKV23ZJb7ns42PfzlznrfN19ik/1sNBaP6apn5Mc5fovEm6SS4HU0VgVRpqC27ZPte8hXK/rf8AI8N8noytXrppW206eyU+Dd7Lx7GetzKtuVvW75biuhpSvQoSlGNtlSpsnPilnFdqu95AwAoASnk51UekcWqc0+gp2niGr5xv1YX7ZNNdyYFj8iWqHRU/nGvG1SrHm4dPbGk9su+WXgl2lrJGKdNRSjFJJJKKSsklkkkfQQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfM43Kg140G8NX50V9XUvKHB+dH9fEuBmq1j0VHE0JUpZO14P7MtzNWWnXXT3+ncyeNmi34z3UifdGk5SUY7W7L/nYZxNCVOcqc1aUZNSXFG+1fwPNj0stsl1eEdt/E8OPHNraddzOZTBg+pve+zY4TDqnBQjsX5ve2SvVXR39/Jdqp/rI02isC61RQ3bZvsRPKcVFKKVkkkuCPpxGo1Dhb3te02t3l1NMaSp4ajUxFZ82FOLlJ78tyW9vYu88s6yacq43EzxVbJzfVis1CK8mC7l7Xdk95bNbunrfIKL+qou9drz6u6PdH+b4Z1eGAAAr7oUpTlGEE5SlJRgltcpO0Uu92PT2oWq0dH4SNFWdST59eSt1ptbO5K0VwXErrkP1S50npOtHKLccKpLbJXU6q7s4p95dSCMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGGjJ8zkkm3kkrtgQbXnV6E6tPEJpX6tWO+Vlk1+a8ToJbku5I72mMf01Ry81ZQXDt72dKE3Fpp2azTJWsRMy25M+S9K0tO4r2TbQeA6Gmr+VLOfuNFym61rR+Dbg109VOGHXY2s6jXZFZ+w6PzrX9LP2mp0vgKWKkp4qCrSiubF1M7LsXZ/RGWmpRs5Nttttttttttt5tt9pguV6r4L1en4J+8k+geTzR7p8+thKUnLOKaeS9u8g85m71N1dnj8XTw0E1FvnVp7oU1m5d7yS4tHoT9nmivUqP7r95s9C6t4TCc54XD06LnbnuEbN22JsDuaPwcKNKFGlFQhTioQitiilZI7IAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA02scqjiqdON+dnN3isuzNgARr5qrfY/ih7x81VvsfxQ95kFRj5qrfY/ih7x81VvsfxQ95kFkdvRWhZyqLpY2is3nHPPJZMmKAMVZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAf/2Q==" style="height:65px;vertical-align:middle;">
                <div>
                    <h2 style="margin:0; color:#5e2d61;">CHACABUCO NOROESTE TOUR S.R.L.</h2>
                    <small>Desde 1996, viajando con vos | CHACAGEST Software System</small>
                </div>
            </div>
            <div class="titulo-doc">RECIBO DE COBRANZA</div>
            <table class="detalle-table">
                <tr><td>CLIENTE:</td><td>{data['Cliente/Proveedor']}</td></tr>
                <tr><td>FECHA:</td><td>{data['Fecha']}</td></tr>
                <tr><td>CONCEPTO:</td><td>{data['Concepto']}</td></tr>
                <tr><td>FORMA DE PAGO:</td><td>{data['Caja/Banco']}</td></tr>
                <tr><td>Nº REFERENCIA:</td><td>{data['Ref AFIP']}</td></tr>
            </table>
            <div class="monto-rec">TOTAL RECIBIDO: $ {abs(data['Monto']):,.2f}</div>
            <div class="firma-box">
                <div class="firma-linea">Firma del cliente</div>
                <div class="firma-linea">Firma y sello empresa</div>
            </div>
            <div class="footer-rec">
                CHACABUCO NOROESTE TOUR S.R.L. | Chacabuco, Buenos Aires | CHACAGEST Software System
            </div>
        </div>
    </body>
    </html>
    """

def generar_html_presupuesto(p_data):
    return f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Trebuchet MS', sans-serif; padding: 30px; }}
            .main-border {{ border: 1px solid #ddd; border-top: 10px solid #f39c12; padding: 40px; box-shadow: 0 0 10px #eee; }}
            .header-p {{ display: flex; justify-content: space-between; margin-bottom: 40px; border-bottom: 1px solid #eee; padding-bottom: 20px; }}
            .label-presu {{ background: #f39c12; color: white; padding: 5px 15px; font-weight: bold; border-radius: 3px; }}
            .box-detalle {{ background: #fafafa; border: 1px solid #eee; padding: 20px; margin: 20px 0; min-height: 120px; }}
            .total-p {{ font-size: 24px; text-align: right; color: #333; border-top: 2px solid #333; padding-top: 10px; margin-top: 20px; }}
            .leyenda-box {{ margin-top: 30px; padding: 15px; border: 1px solid #f39c12; border-radius: 5px; background-color: #fffaf0; font-size: 13px; color: #555; }}
            .footer-p {{ margin-top: 40px; font-size: 11px; color: #888; text-align: center; border-top: 1px solid #eee; padding-top: 15px; }}
        </style>
    </head>
    <body>
        <div class="main-border">
            <div class="header-p">
                <div style="display:flex;align-items:center;gap:14px;">
                    <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2wCEAAkGBxISERUQEBMVEhUWEhUVEBYREhAXFRUXGBIWFxUVFRUYHSggGBslHxUVITEiJSkrLi4wGB8zODMsOCgtLisBCgoKDQ0NGQ8PFzcZFx8rKy0rMis3Ky0rLTcrNys3LTcrLSstLSstMC0tMDIrNystNy03LTY3Ny8rLSs3LTcrK//AABEIALcBEwMBIgACEQEDEQH/xAAcAAEAAgIDAQAAAAAAAAAAAAAABgcBBQMECAL/xABHEAACAQICBQcHCQUHBQAAAAAAAQIDEQQhBQYxQWEHEhMiUXGBMkJTVJHR0hQVFiNSk5ShwReSorHhM0NicnOC8EVjhKOz/8QAGgEBAQADAQEAAAAAAAAAAAAAAAECAwYEBf/EACQRAQACAgEDAwUAAAAAAAAAAAABAgMRBBIxUQUhIhMjQmFx/9oADAMBAAIRAxEAPwC8QAAAAAAAAAAAAAAAAAAAAAAAAcGLxcKUJVaslCEIuU5S2JJXbIsuU/RPra+6r/ABMCOaz6UcbUabae2bTat2RvxNbiuVHRag3DEqUrdVKnXze7zSHVddMFKTlKum27t8yr8IEi+WVfS1PvJ+8fLKvpKn3k/eR6GtmCbSVdXbslzal29iSXN25oli0FXtfmZWvnKCy9pYR1PllX0lT7yfvHy2r6Sp95P3nC0fVGk5yUYq7bshI3Or9OrVqXlUqOEc5deeb3LaS86mjsEqVNQj/ufa97O2RQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjmvms8dH4Sdd2dR9TDxfnVGnbLsWbfBAVxy3623a0ZQnkmpYtx35XhSfDZJr/AClRHJXrynOVSpJynKTlOT2uTd233s4woAbLV3QtTG4mnhaPlTl1pboRWcpvglfv2ATvkW1S+UV3jq0fqqLtRT2Tq73xjFfm+BbWs2keZHoovrS8q26PZ4nawmFo4HCxpU1zYUoKMFvk7b+1t5kOxNeU5Octrbb/AEQRxko1W0bZdPJZvKF+ze/E0uh8A61RR81Zza7OzvZO4RSVlklsRZkZRkAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbA+Kskk23ZJNtvYkt7PM3KRrW9I4tzg/qad4YdXyaum6lu2TS8Eix+W3W5UqXzdRf1lWN67XmUnfq98rPwXEo0AAAoegeR7VJ4TDPE1l9fXSdms6dPbGHe/KfeuwrHkt1dp4rFqriGlQotSkpf3k79SHdezfcu0vbTOmYKnzaUlKUsuq/JW9hGq1j0j0s+ZHyIP2y3s1EYtuyzb2foYJFqvo676aSyV1BPe97A3OhcAqNNR855zfHs8DYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGajWnT1PA4WpiquyK6sd85vKEFxba8Lm2Z535XdbvlmKeHpSvQoNxVtlSp58+5eSvF70BDNK6RqYmtUxFZ86dSTlJ3duCXBKyXBHUAChzYPCyq1I0qavKclGPf/Tb4HCWNyc6C5kPldRdaatRvug9sv938gJPoTRkcNQjRhuXWe+UnnKV/y8DvgykZaR2dG4N1aigu+T7F2k9oUVCKhHJJJIjeDx2EwNPnYqvRoynm+kqQi+EUm7vwOhjOVbRUNleVX/SpVH+bSTMRNwVrV5aNHryadeXfCMfybON8teC9DX9kPeBZwKvfLZg/QV//AF+8x+2zB+r1/wCD3gWiCrHy3YT1av7afvMftvwvq1f20veBagKpfLhhvVa/71L3mHy44b1Sv+/R94Frgqd8uOH9Ur/v0fefP7csP6nW8alH3gW0DQ6E03XxNCniI4V01Ujzoxq1YqaV3ZtJO11n4gDfAAAAAAAAAAAAAAAAAxc1Gs+sNHA4eWIrvJZQj505box4gRblf1teCwvQUZWr104xtthT8+pbt81cXfceeTY6w6bq43ETxVd9abyS2Qir82EeCXtua4AAZpwcmoxTk27RS2tvYkFbnVLQrxVdQf8AZw61V8L2UfFlwxikrLJJWXYkv6Gk0Ho+ngML9ZJRdudWl/its422L+pBtZ9b6mIbp0r06Ozb1p8Zdi4FRLdO67UKF4UvrprLqvqJ8Ze4g+k9bcXW21Ojj9mkuavb5X5mjsBs0zNttybbb8pttt97ebMAEUAAAwZAAAAAAAJfyY6qPH4tc9PoKLU67z62+FK/a9/YvAiuDws6tSFKlHnTnJRgu1t5LgeodRdWoaPwkMNHrS8qtO1nOpLOT7lsS7EgjexjZWSVlkrZWW5GDlAAAxJgZBpsdrXgKNSVKvi8PSqRtzoVK9KMldJq8W7rJp+JwfTjRnr+E/E0fiAkAI/9ONGev4T8TR+IfTjRnr+E/E0fiAkAI/8ATjRnr+E/E0fiH040Z6/hPxNH4gJACP8A040Z6/hPxNH4jq4nlG0VTV3jaUrei59T/wCaYEqMc4rXSXLRgIJ9BCtXe7qKnF+M3f8AIgOsfKxj8QnCi1hIPL6q7qtf6j2eCT4gW/rnrzhdHRfSS6Sra8KNNpzff9lcWefNa9Zq+kK3TYh7LqlCN+ZTT3RvteWb2s09Sbk3KTcm3eTk2232tvNvifIUAAAnHJ9oZK+NrWjGF+icu1K0p9y2eJGdX9ESxVaNKOSunUl9mN8337Uu8kuvml4wjHAYfqwhFdLb+Gn+r8AjT626xyxdS0bqjF9SOznf45Lt7Ow0AAUAAAEl1a1Ex+OtKjS5lPfVrNwp27VleXgn4Fl6E5FMPFJ4uvOtLfGlanC/B5yftAo8xc9O4Tk40VTtbB058avOqX71NtP2Gyjqno9KywWFS4Yah8IHlC4uesfotgPUsL+GofCPotgPUsL+GofCB5OuLnrH6LYD1LC/hqHwj6LYD1LC/hqHwgeTrhtHrH6LYD1LC/hqHwj6LYD1LC/hqHwgVnyIao/9TrRzd44VP2Tq97zSff2lxJHHh6EacYwhFRjFJRjFKMYpKyUYrJLgjlCAAAMj+u2skMBhJ4idnLyaMPt1H5K7ltfBM30pJK7aSW2+482cqGtrx+Lag/qKLcKGflNPrVPHdwsBE8di51qk61WXOnObnNva2/8An5HCAFAAAAAAAAAAAAAA58DhJ1qkaVOLlKTskv5t7lxOxofRFXE1OZRi39qTuox/zMtPV7V+lgqbs+dNpupUllxaXZFAazoqeisFJq0q0978+o1kv8kezhxKyq1XJuUm5Sk7yb2tva3xNzrdpr5VXck/q4XjSW6185d79xpAgAYuFclGlKclCEXOUpKMYxTbbbskki7dQeSmnSUcRpGKq1cnCi86dPeuevPlw2LvOxyRairD0447Ew+vqK9KMl/YwezLdOS9iaRZyCMRikkkklutsPqwAAAAAAAAAAAAAAAAAFY8tGtvyagsFRlatXT6Rp506W/xlsXBMoax6D03yTUMXXqYmvisS51Jc550rRW6MVzcklZHTXIlgvWMR7aXwAUQC+FyJ4H02I9tL4T6XIpgfTYj96n8IFCgvv8AYtgPS1/34fCJcjej0rupXtxqR9wFCAvhckei/S1fvY+4z+yTRXpan30RuGXTbwoYF9rkm0T6Wp9+jK5KdEfbm/8AyBuF6Z8KDMSkltaXeeg6eoGg6OclGX+pXlJfzO/o+Wh6ErYalQ5/munSUpZf42m14sm4X6WSY3FfZ5/wGgcTWt0dGdnslKLjG3beVrruJfobk+SfPxVTnf8Abp7P909/giwsXiHUm5y3/ktyRsMLq/WqQU04JNXSk5J29jMmto8LhYUoKnTioRWxRSSIjyi6d5kPklN9aavVs9kL+Twvb8ic6zYd4HDTxVaVPmxyilKV5SeUIJc3a3+pQmNxUqtSVWbvKcnKT/52ZLwA4QARQnnJHqk8bilXqq9ChJSnfZOpthDuTzfdxIbovR9TEVoUKMedOpJRgt1+19iWbfcepdVNBU8DhaeFpZ82PXlZXnN+VN8W/wBAjbxRkAAAAAAAAAAAAAAAAAAAAAAAAAAdXSODjWpSpT2SVmdowyd1iZidwonTGBnh606M73i8nd9aPmyXgdPnPtftZafKDoHpqPTU1epTV8lnKO9d+8qo+dmpNLfp3HpvJpyMMTr5R7Szd9oMA1bfQ6Y8FiUaDwXRw50l1pbeC3LhlZ+JqtB4HpJ86Xkx28XuXh+hKacHJqKV23ZJb7ns42PfzlznrfN19ik/1sNBaP6apn5Mc5fovEm6SS4HU0VgVRpqC27ZPte8hXK/rf8AI8N8noytXrppW206eyU+Dd7Lx7GetzKtuVvW75biuhpSvQoSlGNtlSpsnPilnFdqu95AwAoASnk51UekcWqc0+gp2niGr5xv1YX7ZNNdyYFj8iWqHRU/nGvG1SrHm4dPbGk9su+WXgl2lrJGKdNRSjFJJJKKSsklkkkfQQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfM43Kg140G8NX50V9XUvKHB+dH9fEuBmq1j0VHE0JUpZO14P7MtzNWWnXXT3+ncyeNmi34z3UifdGk5SUY7W7L/nYZxNCVOcqc1aUZNSXFG+1fwPNj0stsl1eEdt/E8OPHNraddzOZTBg+pve+zY4TDqnBQjsX5ve2SvVXR39/Jdqp/rI02isC61RQ3bZvsRPKcVFKKVkkkuCPpxGo1Dhb3te02t3l1NMaSp4ajUxFZ82FOLlJ78tyW9vYu88s6yacq43EzxVbJzfVis1CK8mC7l7Xdk95bNbunrfIKL+qou9drz6u6PdH+b4Z1eGAAAr7oUpTlGEE5SlJRgltcpO0Uu92PT2oWq0dH4SNFWdST59eSt1ptbO5K0VwXErrkP1S50npOtHKLccKpLbJXU6q7s4p95dSCMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGGjJ8zkkm3kkrtgQbXnV6E6tPEJpX6tWO+Vlk1+a8ToJbku5I72mMf01Ry81ZQXDt72dKE3Fpp2azTJWsRMy25M+S9K0tO4r2TbQeA6Gmr+VLOfuNFym61rR+Dbg109VOGHXY2s6jXZFZ+w6PzrX9LP2mp0vgKWKkp4qCrSiubF1M7LsXZ/RGWmpRs5Nttttttttttt5tt9pguV6r4L1en4J+8k+geTzR7p8+thKUnLOKaeS9u8g85m71N1dnj8XTw0E1FvnVp7oU1m5d7yS4tHoT9nmivUqP7r95s9C6t4TCc54XD06LnbnuEbN22JsDuaPwcKNKFGlFQhTioQitiilZI7IAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA02scqjiqdON+dnN3isuzNgARr5qrfY/ih7x81VvsfxQ95kFRj5qrfY/ih7x81VvsfxQ95kFkdvRWhZyqLpY2is3nHPPJZMmKAMVZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAf/2Q==" style="height:65px;vertical-align:middle;margin-right:12px;">
                    <div>
                        <h1 style="margin:0; color:#444;">CHACABUCO NOROESTE TOUR S.R.L.</h1>
                        <small>VIAJES ESPECIALES - TURISMO - TRASLADOS PERSONALES</small>
                    </div>
                </div>
                <div style="text-align: right;">
                    <span class="label-presu">PRESUPUESTO</span><br>
                    <p style="margin-top:10px;"><b>Fecha:</b> {p_data['Fecha Emisión']}</p>
                </div>
            </div>
            <p><b>CLIENTE:</b> {p_data['Cliente']}</p>
            <p><b>TIPO DE UNIDAD:</b> {p_data['Tipo Móvil']}</p>
            <div class="box-detalle">
                <b>DETALLE DEL SERVICIO:</b><br><br>
                {str(p_data['Detalle']).replace(chr(10), '<br>')}
            </div>
            <div class="total-p">
                TOTAL: <span style="color: #d35400;">$ {p_data['Importe']:,.2f}</span>
            </div>
            <div class="leyenda-box">
                • La seña para la reserva es del 30%.<br>
                • Los gastos de los choferes (hospedaje y comida) estarán a cargo del contratante. En caso de que la empresa tenga que hacerse responsable de los mismos, el presente presupuesto deberá ser reformulado.
            </div>
            <p style="margin: 20px 0; font-size: 13px;"><b>Validez de la oferta:</b> Hasta el {p_data['Vencimiento']}</p>
            <div class="footer-p">
                CHACABUCO NOROESTE TOUR S.R.L. | Chacabuco, Buenos Aires | CHACAGEST Software System
            </div>
        </div>
    </body>
    </html>
    """

def generar_html_factura(data):
    """Genera HTML profesional de Factura / NC / ND con logo de empresa."""
    tipo      = data.get("tipo", "FACTURA")
    letra     = data.get("letra", "B")
    pv        = str(data.get("punto_venta", "0001")).zfill(4)
    nro       = str(data.get("numero", "00000001")).zfill(8)
    fecha     = data.get("fecha", "")
    cliente   = data.get("cliente", "")
    cuit_cli  = data.get("cuit_cliente", "")
    cond_iva  = data.get("condicion_iva", "")
    dir_cli   = data.get("direccion_cliente", "")
    items     = data.get("items", [])
    neto      = data.get("neto", 0.0)
    iva_monto = data.get("iva_monto", 0.0)
    no_grav   = data.get("no_gravado", 0.0)
    total     = data.get("total", 0.0)
    obs       = data.get("observaciones", "")
    comp_asoc = data.get("comp_asoc", "")
    logo_b64  = data.get("logo_b64", "")
    responsable = data.get("responsable", "")

    color_tipo = {"FACTURA": "#5e2d61", "NOTA DE CREDITO": "#27ae60", "NOTA DE DEBITO": "#c0392b"}.get(tipo, "#5e2d61")
    label_tipo = {"FACTURA": "FACTURA", "NOTA DE CREDITO": "NOTA DE CRÉDITO", "NOTA DE DEBITO": "NOTA DE DÉBITO"}.get(tipo, tipo)

    filas_items = ""
    for it in items:
        filas_items += f"""
        <tr>
            <td style='padding:7px 10px;border-bottom:1px solid #f0f0f0;'>{it.get('descripcion','')}</td>
            <td style='padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;'>{it.get('cantidad',1)}</td>
            <td style='padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:right;'>${float(it.get('precio_unitario',0)):,.2f}</td>
            <td style='padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:center;'>{it.get('alicuota','21%')}</td>
            <td style='padding:7px 10px;border-bottom:1px solid #f0f0f0;text-align:right;font-weight:bold;'>${float(it.get('subtotal',0)):,.2f}</td>
        </tr>"""

    comp_asoc_html = f"<tr><td><b>Comprobante Asociado:</b></td><td>{comp_asoc}</td></tr>" if comp_asoc else ""
    obs_html = f"<div style='margin-top:16px;padding:12px;background:#fafafa;border-radius:6px;font-size:12px;color:#555;border-left:3px solid {color_tipo};'><b>Observaciones:</b> {obs}</div>" if obs else ""
    logo_html = f"<img src='{logo_b64}' style='height:70px;vertical-align:middle;margin-right:14px;'>" if logo_b64 else ""

    return f"""<!DOCTYPE html>
<html lang='es'>
<head>
  <meta charset='UTF-8'>
  <style>
    * {{ box-sizing: border-box; margin:0; padding:0; }}
    body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #333; background: #fff; font-size:13px; padding:30px; }}
    .doc-wrapper {{ max-width:820px; margin:0 auto; border:1px solid #ddd; border-top:8px solid {color_tipo}; padding:30px; }}
    .header {{ display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:20px; padding-bottom:16px; border-bottom:1px solid #eee; }}
    .empresa-info h2 {{ color:#333; font-size:18px; margin-bottom:4px; }}
    .empresa-info small {{ color:#777; font-size:11px; }}
    .tipo-badge {{ text-align:center; border:3px solid {color_tipo}; border-radius:8px; padding:12px 24px; min-width:160px; }}
    .tipo-badge .letra {{ font-size:52px; font-weight:900; color:{color_tipo}; line-height:1; }}
    .tipo-badge .tipo-label {{ font-size:11px; font-weight:bold; color:{color_tipo}; letter-spacing:1px; margin-top:4px; }}
    .tipo-badge .nro-comp {{ font-size:13px; font-weight:bold; color:#333; margin-top:6px; }}
    .datos-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:20px; }}
    .datos-box {{ background:#f8f9fa; border-radius:6px; padding:14px; }}
    .datos-box h4 {{ font-size:10px; font-weight:bold; color:{color_tipo}; text-transform:uppercase; letter-spacing:1px; margin-bottom:8px; border-bottom:1px solid #eee; padding-bottom:6px; }}
    .datos-box table {{ width:100%; border-collapse:collapse; }}
    .datos-box td {{ padding:3px 0; font-size:12px; }}
    .datos-box td:first-child {{ color:#777; width:42%; }}
    .items-table {{ width:100%; border-collapse:collapse; margin-bottom:20px; }}
    .items-table thead tr {{ background:{color_tipo}; color:white; }}
    .items-table thead th {{ padding:9px 10px; text-align:left; font-size:12px; font-weight:bold; }}
    .items-table thead th:nth-child(2), .items-table thead th:nth-child(3), .items-table thead th:nth-child(4), .items-table thead th:nth-child(5) {{ text-align:center; }}
    .items-table thead th:nth-child(5) {{ text-align:right; }}
    .totales-box {{ display:flex; justify-content:flex-end; margin-bottom:20px; }}
    .totales-inner {{ min-width:260px; }}
    .totales-inner table {{ width:100%; border-collapse:collapse; }}
    .totales-inner td {{ padding:5px 10px; font-size:13px; }}
    .totales-inner td:last-child {{ text-align:right; font-weight:bold; }}
    .total-final td {{ background:{color_tipo}; color:white; font-size:16px; font-weight:900; border-radius:4px; padding:10px 14px; }}
    .footer-doc {{ text-align:center; font-size:10px; color:#aaa; margin-top:20px; border-top:1px solid #eee; padding-top:12px; }}
    @media print {{
      body {{ padding:0; }}
      .doc-wrapper {{ border:none; padding:20px; }}
    }}
  </style>
</head>
<body>
  <div class='doc-wrapper'>
    <div class='header'>
      <div class='empresa-info'>
        {logo_html}
        <div style='display:inline-block;vertical-align:middle;'>
          <h2>CHACABUCO NOROESTE TOUR S.R.L.</h2>
          <small>VIAJES ESPECIALES · TURISMO · TRASLADOS PERSONALES<br>
          CUIT: 30-71234567-9 · IVA Responsable Inscripto<br>
          Chacabuco, Buenos Aires</small>
        </div>
      </div>
      <div class='tipo-badge'>
        <div class='letra'>{letra}</div>
        <div class='tipo-label'>{label_tipo}</div>
        <div class='nro-comp'>{pv}-{nro}</div>
      </div>
    </div>

    <div class='datos-grid'>
      <div class='datos-box'>
        <h4>Datos del Comprobante</h4>
        <table>
          <tr><td>Fecha:</td><td><b>{fecha}</b></td></tr>
          <tr><td>Punto de Venta:</td><td>{pv}</td></tr>
          <tr><td>Número:</td><td>{nro}</td></tr>
          {comp_asoc_html}
          <tr><td>Emitido por:</td><td>{responsable}</td></tr>
        </table>
      </div>
      <div class='datos-box'>
        <h4>Datos del Cliente</h4>
        <table>
          <tr><td>Razón Social:</td><td><b>{cliente}</b></td></tr>
          <tr><td>CUIT/DNI:</td><td>{cuit_cli}</td></tr>
          <tr><td>Condición IVA:</td><td>{cond_iva}</td></tr>
          <tr><td>Dirección:</td><td>{dir_cli}</td></tr>
        </table>
      </div>
    </div>

    <table class='items-table'>
      <thead>
        <tr>
          <th style='width:45%;'>Descripción</th>
          <th style='width:10%;text-align:center;'>Cant.</th>
          <th style='width:17%;text-align:right;'>Precio Unit.</th>
          <th style='width:12%;text-align:center;'>IVA %</th>
          <th style='width:16%;text-align:right;'>Subtotal</th>
        </tr>
      </thead>
      <tbody>
        {filas_items}
      </tbody>
    </table>

    <div class='totales-box'>
      <div class='totales-inner'>
        <table>
          <tr><td>Subtotal Neto:</td><td>$ {neto:,.2f}</td></tr>
          <tr><td>IVA:</td><td>$ {iva_monto:,.2f}</td></tr>
          {'<tr><td>No Gravado:</td><td>$ ' + f'{no_grav:,.2f}' + '</td></tr>' if no_grav else ''}
          <tr class='total-final'><td>TOTAL:</td><td>$ {total:,.2f}</td></tr>
        </table>
      </div>
    </div>

    {obs_html}

    <div class='footer-doc'>
      Generado por CHACAGEST · {fecha} · {label_tipo} {letra} {pv}-{nro} — Chacabuco Noroeste Tour S.R.L.<br>
      <small>Este comprobante no reemplaza a los comprobantes oficiales emitidos ante AFIP</small>
    </div>
  </div>
</body>
</html>"""


def generar_html_cierre_caja(data):
    # Construir tabla de movimientos del día
    df = data['movimientos']
    filas_html = ""
    for _, row in df.iterrows():
        color_fila = "#fff" if row['Monto'] >= 0 else "#fff8f8"
        signo = "+" if row['Monto'] >= 0 else ""
        filas_html += f"""
        <tr style="background:{color_fila};">
            <td>{row['Fecha']}</td>
            <td>{row['Tipo']}</td>
            <td>{row.get('Forma','-')}</td>
            <td>{row['Concepto']}</td>
            <td>{row['Cliente/Proveedor']}</td>
            <td style="text-align:right;font-weight:bold;color:{'#27ae60' if row['Monto']>=0 else '#e74c3c'};">{signo}$ {row['Monto']:,.2f}</td>
        </tr>"""

    # Subtotales por forma
    FORMAS = ["EFECTIVO", "TRANSFERENCIA", "TARJETA DE CREDITO", "DÓLARES", "OTROS"]
    ICONOS = {"EFECTIVO":"💵","TRANSFERENCIA":"🏦","TARJETA DE CREDITO":"💳","DÓLARES":"💲","OTROS":"📋"}
    subtotales_html = ""
    for f in FORMAS:
        mask = mask_forma(df['Forma'], f.replace("DÓLARES","DOLARES").replace("TARJETA DE CREDITO","TARJETA"))
        sub = df[mask]['Monto'].sum()
        if sub != 0:
            color_s = "#27ae60" if sub >= 0 else "#e74c3c"
            signo_s = "+" if sub >= 0 else ""
            subtotales_html += f"""
            <tr>
                <td style="padding:8px 12px;">{ICONOS.get(f,'💰')} {f}</td>
                <td style="text-align:right;padding:8px 12px;font-weight:bold;color:{color_s};">{signo_s}$ {sub:,.2f}</td>
            </tr>"""

    total = data['total']
    color_total = "#27ae60" if total >= 0 else "#e74c3c"

    return f"""<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <style>
        @media print {{ .no-print {{ display: none; }} body {{ margin: 0; }} }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #333; padding: 30px; background: #fff; }}
        .header-cierre {{ display: flex; justify-content: space-between; align-items: flex-start;
                          border-bottom: 4px solid #5e2d61; padding-bottom: 20px; margin-bottom: 25px; }}
        .empresa-nombre {{ font-size: 22px; font-weight: bold; color: #5e2d61; margin: 0; }}
        .empresa-sub {{ font-size: 12px; color: #888; margin-top: 4px; }}
        .badge-cierre {{ background: #5e2d61; color: white; padding: 8px 20px; border-radius: 6px;
                         font-size: 16px; font-weight: bold; text-align: center; }}
        .info-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; margin-bottom: 25px; }}
        .info-box {{ background: #f8f9fa; border-radius: 8px; padding: 14px; border-left: 4px solid #f39c12; }}
        .info-label {{ font-size: 11px; color: #888; font-weight: bold; text-transform: uppercase; }}
        .info-value {{ font-size: 16px; font-weight: bold; color: #333; margin-top: 4px; }}
        table.mov {{ width: 100%; border-collapse: collapse; font-size: 12px; margin-bottom: 25px; }}
        table.mov th {{ background: #5e2d61; color: white; padding: 9px 10px; text-align: left; }}
        table.mov td {{ padding: 7px 10px; border-bottom: 1px solid #eee; }}
        .subtotales-tabla {{ width: 100%; border-collapse: collapse; max-width: 380px; margin-left: auto; }}
        .subtotales-tabla td {{ border: 1px solid #ddd; }}
        .subtotales-header {{ background: #f0f2f6; font-weight: bold; font-size: 13px; padding: 10px 12px; text-transform: uppercase; letter-spacing: 1px; }}
        .total-box {{ background: {color_total}; color: white; padding: 16px 24px; border-radius: 10px;
                      text-align: right; margin-top: 20px; }}
        .total-label {{ font-size: 13px; opacity: 0.9; }}
        .total-monto {{ font-size: 32px; font-weight: bold; display: block; margin-top: 4px; }}
        .firmas {{ display: grid; grid-template-columns: 1fr 1fr; gap: 40px; margin-top: 60px; }}
        .firma-box {{ border-top: 2px solid #333; padding-top: 10px; text-align: center; font-size: 12px; color: #555; }}
        .footer-cierre {{ margin-top: 40px; text-align: center; font-size: 10px; color: #bbb;
                           border-top: 1px solid #eee; padding-top: 12px; }}
        .btn-imprimir {{ background: #5e2d61; color: white; border: none; padding: 12px 28px;
                          font-size: 15px; font-weight: bold; border-radius: 8px; cursor: pointer;
                          display: block; margin: 0 auto 24px auto; }}
        .btn-imprimir:hover {{ background: #4a2350; }}
    </style>
</head>
<body>
    <button class="btn-imprimir no-print" onclick="window.print()">🖨️ Imprimir / Guardar PDF</button>

    <div class="header-cierre">
        <div style="display:flex;align-items:center;gap:14px;">
            <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/2wCEAAkGBxISERUQEBMVEhUWEhUVEBYREhAXFRUXGBIWFxUVFRUYHSggGBslHxUVITEiJSkrLi4wGB8zODMsOCgtLisBCgoKDQ0NGQ8PFzcZFx8rKy0rMis3Ky0rLTcrNys3LTcrLSstLSstMC0tMDIrNystNy03LTY3Ny8rLSs3LTcrK//AABEIALcBEwMBIgACEQEDEQH/xAAcAAEAAgIDAQAAAAAAAAAAAAAABgcBBQMECAL/xABHEAACAQICBQcHCQUHBQAAAAAAAQIDEQQhBQYxQWEHEhMiUXGBMkJTVJHR0hQVFiNSk5ShwReSorHhM0NicnOC8EVjhKOz/8QAGgEBAQADAQEAAAAAAAAAAAAAAAECAwYEBf/EACQRAQACAgEDAwUAAAAAAAAAAAABAgMRBBIxUQUhIhMjQmFx/9oADAMBAAIRAxEAPwC8QAAAAAAAAAAAAAAAAAAAAAAAAcGLxcKUJVaslCEIuU5S2JJXbIsuU/RPra+6r/ABMCOaz6UcbUabae2bTat2RvxNbiuVHRag3DEqUrdVKnXze7zSHVddMFKTlKum27t8yr8IEi+WVfS1PvJ+8fLKvpKn3k/eR6GtmCbSVdXbslzal29iSXN25oli0FXtfmZWvnKCy9pYR1PllX0lT7yfvHy2r6Sp95P3nC0fVGk5yUYq7bshI3Or9OrVqXlUqOEc5deeb3LaS86mjsEqVNQj/ufa97O2RQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjmvms8dH4Sdd2dR9TDxfnVGnbLsWbfBAVxy3623a0ZQnkmpYtx35XhSfDZJr/AClRHJXrynOVSpJynKTlOT2uTd233s4woAbLV3QtTG4mnhaPlTl1pboRWcpvglfv2ATvkW1S+UV3jq0fqqLtRT2Tq73xjFfm+BbWs2keZHoovrS8q26PZ4nawmFo4HCxpU1zYUoKMFvk7b+1t5kOxNeU5Octrbb/AEQRxko1W0bZdPJZvKF+ze/E0uh8A61RR81Zza7OzvZO4RSVlklsRZkZRkAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAbA+Kskk23ZJNtvYkt7PM3KRrW9I4tzg/qad4YdXyaum6lu2TS8Eix+W3W5UqXzdRf1lWN67XmUnfq98rPwXEo0AAAoegeR7VJ4TDPE1l9fXSdms6dPbGHe/KfeuwrHkt1dp4rFqriGlQotSkpf3k79SHdezfcu0vbTOmYKnzaUlKUsuq/JW9hGq1j0j0s+ZHyIP2y3s1EYtuyzb2foYJFqvo676aSyV1BPe97A3OhcAqNNR855zfHs8DYAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGajWnT1PA4WpiquyK6sd85vKEFxba8Lm2Z535XdbvlmKeHpSvQoNxVtlSp58+5eSvF70BDNK6RqYmtUxFZ86dSTlJ3duCXBKyXBHUAChzYPCyq1I0qavKclGPf/Tb4HCWNyc6C5kPldRdaatRvug9sv938gJPoTRkcNQjRhuXWe+UnnKV/y8DvgykZaR2dG4N1aigu+T7F2k9oUVCKhHJJJIjeDx2EwNPnYqvRoynm+kqQi+EUm7vwOhjOVbRUNleVX/SpVH+bSTMRNwVrV5aNHryadeXfCMfybON8teC9DX9kPeBZwKvfLZg/QV//AF+8x+2zB+r1/wCD3gWiCrHy3YT1av7afvMftvwvq1f20veBagKpfLhhvVa/71L3mHy44b1Sv+/R94Frgqd8uOH9Ur/v0fefP7csP6nW8alH3gW0DQ6E03XxNCniI4V01Ujzoxq1YqaV3ZtJO11n4gDfAAAAAAAAAAAAAAAAAxc1Gs+sNHA4eWIrvJZQj505box4gRblf1teCwvQUZWr104xtthT8+pbt81cXfceeTY6w6bq43ETxVd9abyS2Qir82EeCXtua4AAZpwcmoxTk27RS2tvYkFbnVLQrxVdQf8AZw61V8L2UfFlwxikrLJJWXYkv6Gk0Ho+ngML9ZJRdudWl/its422L+pBtZ9b6mIbp0r06Ozb1p8Zdi4FRLdO67UKF4UvrprLqvqJ8Ze4g+k9bcXW21Ojj9mkuavb5X5mjsBs0zNttybbb8pttt97ebMAEUAAAwZAAAAAAAJfyY6qPH4tc9PoKLU67z62+FK/a9/YvAiuDws6tSFKlHnTnJRgu1t5LgeodRdWoaPwkMNHrS8qtO1nOpLOT7lsS7EgjexjZWSVlkrZWW5GDlAAAxJgZBpsdrXgKNSVKvi8PSqRtzoVK9KMldJq8W7rJp+JwfTjRnr+E/E0fiAkAI/9ONGev4T8TR+IfTjRnr+E/E0fiAkAI/8ATjRnr+E/E0fiH040Z6/hPxNH4gJACP8A040Z6/hPxNH4jq4nlG0VTV3jaUrei59T/wCaYEqMc4rXSXLRgIJ9BCtXe7qKnF+M3f8AIgOsfKxj8QnCi1hIPL6q7qtf6j2eCT4gW/rnrzhdHRfSS6Sra8KNNpzff9lcWefNa9Zq+kK3TYh7LqlCN+ZTT3RvteWb2s09Sbk3KTcm3eTk2232tvNvifIUAAAnHJ9oZK+NrWjGF+icu1K0p9y2eJGdX9ESxVaNKOSunUl9mN8337Uu8kuvml4wjHAYfqwhFdLb+Gn+r8AjT626xyxdS0bqjF9SOznf45Lt7Ow0AAUAAAEl1a1Ex+OtKjS5lPfVrNwp27VleXgn4Fl6E5FMPFJ4uvOtLfGlanC/B5yftAo8xc9O4Tk40VTtbB058avOqX71NtP2Gyjqno9KywWFS4Yah8IHlC4uesfotgPUsL+GofCPotgPUsL+GofCB5OuLnrH6LYD1LC/hqHwj6LYD1LC/hqHwgeTrhtHrH6LYD1LC/hqHwj6LYD1LC/hqHwgVnyIao/9TrRzd44VP2Tq97zSff2lxJHHh6EacYwhFRjFJRjFKMYpKyUYrJLgjlCAAAMj+u2skMBhJ4idnLyaMPt1H5K7ltfBM30pJK7aSW2+482cqGtrx+Lag/qKLcKGflNPrVPHdwsBE8di51qk61WXOnObnNva2/8An5HCAFAAAAAAAAAAAAAA58DhJ1qkaVOLlKTskv5t7lxOxofRFXE1OZRi39qTuox/zMtPV7V+lgqbs+dNpupUllxaXZFAazoqeisFJq0q0978+o1kv8kezhxKyq1XJuUm5Sk7yb2tva3xNzrdpr5VXck/q4XjSW6185d79xpAgAYuFclGlKclCEXOUpKMYxTbbbskki7dQeSmnSUcRpGKq1cnCi86dPeuevPlw2LvOxyRairD0447Ew+vqK9KMl/YwezLdOS9iaRZyCMRikkkklutsPqwAAAAAAAAAAAAAAAAAFY8tGtvyagsFRlatXT6Rp506W/xlsXBMoax6D03yTUMXXqYmvisS51Jc550rRW6MVzcklZHTXIlgvWMR7aXwAUQC+FyJ4H02I9tL4T6XIpgfTYj96n8IFCgvv8AYtgPS1/34fCJcjej0rupXtxqR9wFCAvhckei/S1fvY+4z+yTRXpan30RuGXTbwoYF9rkm0T6Wp9+jK5KdEfbm/8AyBuF6Z8KDMSkltaXeeg6eoGg6OclGX+pXlJfzO/o+Wh6ErYalQ5/munSUpZf42m14sm4X6WSY3FfZ5/wGgcTWt0dGdnslKLjG3beVrruJfobk+SfPxVTnf8Abp7P909/giwsXiHUm5y3/ktyRsMLq/WqQU04JNXSk5J29jMmto8LhYUoKnTioRWxRSSIjyi6d5kPklN9aavVs9kL+Twvb8ic6zYd4HDTxVaVPmxyilKV5SeUIJc3a3+pQmNxUqtSVWbvKcnKT/52ZLwA4QARQnnJHqk8bilXqq9ChJSnfZOpthDuTzfdxIbovR9TEVoUKMedOpJRgt1+19iWbfcepdVNBU8DhaeFpZ82PXlZXnN+VN8W/wBAjbxRkAAAAAAAAAAAAAAAAAAAAAAAAAAdXSODjWpSpT2SVmdowyd1iZidwonTGBnh606M73i8nd9aPmyXgdPnPtftZafKDoHpqPTU1epTV8lnKO9d+8qo+dmpNLfp3HpvJpyMMTr5R7Szd9oMA1bfQ6Y8FiUaDwXRw50l1pbeC3LhlZ+JqtB4HpJ86Xkx28XuXh+hKacHJqKV23ZJb7ns42PfzlznrfN19ik/1sNBaP6apn5Mc5fovEm6SS4HU0VgVRpqC27ZPte8hXK/rf8AI8N8noytXrppW206eyU+Dd7Lx7GetzKtuVvW75biuhpSvQoSlGNtlSpsnPilnFdqu95AwAoASnk51UekcWqc0+gp2niGr5xv1YX7ZNNdyYFj8iWqHRU/nGvG1SrHm4dPbGk9su+WXgl2lrJGKdNRSjFJJJKKSsklkkkfQQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAfM43Kg140G8NX50V9XUvKHB+dH9fEuBmq1j0VHE0JUpZO14P7MtzNWWnXXT3+ncyeNmi34z3UifdGk5SUY7W7L/nYZxNCVOcqc1aUZNSXFG+1fwPNj0stsl1eEdt/E8OPHNraddzOZTBg+pve+zY4TDqnBQjsX5ve2SvVXR39/Jdqp/rI02isC61RQ3bZvsRPKcVFKKVkkkuCPpxGo1Dhb3te02t3l1NMaSp4ajUxFZ82FOLlJ78tyW9vYu88s6yacq43EzxVbJzfVis1CK8mC7l7Xdk95bNbunrfIKL+qou9drz6u6PdH+b4Z1eGAAAr7oUpTlGEE5SlJRgltcpO0Uu92PT2oWq0dH4SNFWdST59eSt1ptbO5K0VwXErrkP1S50npOtHKLccKpLbJXU6q7s4p95dSCMgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAGGjJ8zkkm3kkrtgQbXnV6E6tPEJpX6tWO+Vlk1+a8ToJbku5I72mMf01Ry81ZQXDt72dKE3Fpp2azTJWsRMy25M+S9K0tO4r2TbQeA6Gmr+VLOfuNFym61rR+Dbg109VOGHXY2s6jXZFZ+w6PzrX9LP2mp0vgKWKkp4qCrSiubF1M7LsXZ/RGWmpRs5Nttttttttttt5tt9pguV6r4L1en4J+8k+geTzR7p8+thKUnLOKaeS9u8g85m71N1dnj8XTw0E1FvnVp7oU1m5d7yS4tHoT9nmivUqP7r95s9C6t4TCc54XD06LnbnuEbN22JsDuaPwcKNKFGlFQhTioQitiilZI7IAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA02scqjiqdON+dnN3isuzNgARr5qrfY/ih7x81VvsfxQ95kFRj5qrfY/ih7x81VvsfxQ95kFkdvRWhZyqLpY2is3nHPPJZMmKAMVZAAAAAAAAAAAAAAAAAAAAAAAAAAAAAf/2Q==" style="height:65px;vertical-align:middle;margin-right:12px;">
            <div>
                <p class="empresa-nombre">CHACABUCO NOROESTE TOUR S.R.L.</p>
                <p class="empresa-sub">Desde 1996 viajando con vos | CHACAGEST Software System</p>
            </div>
        </div>
        <div class="badge-cierre">RENDICIÓN DE CAJA</div>
    </div>

    <div class="info-grid">
        <div class="info-box">
            <div class="info-label">📅 Fecha de Cierre</div>
            <div class="info-value">{data['fecha_cierre']}</div>
        </div>
        <div class="info-box">
            <div class="info-label">🏦 Caja</div>
            <div class="info-value">{data['caja']}</div>
        </div>
        <div class="info-box">
            <div class="info-label">👤 Responsable</div>
            <div class="info-value">{data['responsable']}</div>
        </div>
    </div>

    <h3 style="color:#5e2d61;border-bottom:2px solid #f39c12;padding-bottom:6px;">📋 Movimientos del Período</h3>
    <table class="mov">
        <thead>
            <tr>
                <th>Fecha</th><th>Tipo</th><th>Forma</th><th>Concepto</th><th>Cliente / Proveedor</th><th style="text-align:right;">Monto</th>
            </tr>
        </thead>
        <tbody>
            {filas_html if filas_html else '<tr><td colspan="6" style="text-align:center;padding:20px;color:#aaa;">Sin movimientos en el período</td></tr>'}
        </tbody>
    </table>

    <table class="subtotales-tabla">
        <tr><td colspan="2" class="subtotales-header">Resumen por Forma de Pago</td></tr>
        {subtotales_html if subtotales_html else '<tr><td colspan="2" style="padding:8px 12px;color:#aaa;">Sin movimientos</td></tr>'}
    </table>

    <div style="margin-top:20px;border-radius:10px;overflow:hidden;border:2px solid #5e2d61;">
        <div style="background:#5e2d61;color:white;padding:10px 16px;font-weight:bold;font-size:13px;letter-spacing:1px;">
            💰 RESUMEN DE RENDICIÓN
        </div>
        <div style="display:flex;gap:0;">
            <div style="flex:1;padding:16px 20px;border-right:1px solid #eee;background:#f0fff4;">
                <div style="font-size:11px;color:#888;font-weight:bold;">💵 EFECTIVO DISPONIBLE (PESOS)</div>
                <div style="font-size:26px;font-weight:bold;color:#27ae60;margin-top:6px;">$ {data.get('efectivo_disponible', 0):,.2f}</div>
            </div>
            <div style="flex:1;padding:16px 20px;background:#fffaf0;">
                <div style="font-size:11px;color:#888;font-weight:bold;">💲 DÓLARES DISPONIBLES</div>
                <div style="font-size:26px;font-weight:bold;color:#d4a017;margin-top:6px;">USD {data.get('dolares_disponibles', 0):,.2f}</div>
            </div>
        </div>
        <div style="background:#fdf2f8;border-top:2px solid #5e2d61;padding:18px 20px;">
            <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:16px;">
                <div>
                    <div style="font-size:11px;color:#5e2d61;font-weight:bold;text-transform:uppercase;">📤 Monto Rendido ({data.get('tipo_rendicion','EFECTIVO')})</div>
                    <div style="font-size:30px;font-weight:bold;color:#c0392b;margin-top:6px;">
                        {'$' if data.get('tipo_rendicion','EFECTIVO') == 'EFECTIVO' else 'USD'} {data.get('monto_rendicion', 0):,.2f}
                    </div>
                </div>
                <div style="font-size:32px;color:#ccc;">→</div>
                <div style="text-align:right;">
                    <div style="font-size:11px;color:#5e2d61;font-weight:bold;text-transform:uppercase;">
                        {'✅ Queda en Caja' if data.get('saldo_restante', 0) >= 0 else '⚠️ Queda en Caja (excede disponible)'}
                    </div>
                    <div style="font-size:30px;font-weight:bold;margin-top:6px;color:{'#27ae60' if data.get('saldo_restante', 0) >= 0 else '#e74c3c'};">
                        {'$' if data.get('tipo_rendicion','EFECTIVO') == 'EFECTIVO' else 'USD'} {data.get('saldo_restante', 0):,.2f}
                    </div>
                </div>
            </div>
        </div>
    </div>

    {"<div style='background:#fff8e1;border:1px solid #f39c12;border-radius:8px;padding:14px;margin-top:20px;font-size:13px;'><b>📝 Observaciones:</b><br><br>" + data['observaciones'] + "</div>" if data.get('observaciones') else ""}

    <div class="firmas">
        <div class="firma-box">
            <p style="margin:4px 0;">Responsable de Caja</p>
            <p style="margin:4px 0;font-weight:bold;">{data['responsable']}</p>
        </div>
        <div class="firma-box">
            <p style="margin:4px 0;">Supervisión / Administración</p>
            <p style="margin:4px 0;color:#aaa;">____________________________</p>
        </div>
    </div>

    <div class="footer-cierre">
        Generado por CHACAGEST · {data['fecha_cierre']} · Rendición de Caja — Chacabuco Noroeste Tour S.R.L.
    </div>
</body>
</html>"""

# --- 2. SISTEMA DE USUARIOS Y ROLES ---
# ─────────────────────────────────────────────────────────────────────────────
# USUARIOS DEL SISTEMA
# Estructura: "usuario": {"password": "...", "rol": "admin"/"operador", "caja": "NOMBRE CAJA"}
# rol "admin"   → acceso total (Dashboard, todas las cajas, todo el sistema)
# rol "operador"→ acceso a su caja propia, carga viajes/gastos/clientes/proveedores, sin Dashboard
#
# Para agregar un nuevo usuario: copiar uno de los bloques de operador y editar.
# Las cajas de operadores deben coincidir con los nombres en opc_cajas (más abajo).
# ─────────────────────────────────────────────────────────────────────────────
USUARIOS = {
    "admin": {
        "password": "chaca2026",
        "rol": "admin",
        "caja": None,           # Admin no tiene caja asignada, ve todas
        "nombre": "Administrador"
    },
    "coti": {
        "password": "coti2026",
        "rol": "operador",
        "caja": "CAJA COTI",    # Solo puede ver y operar CAJA COTI
        "nombre": "Coti"
    },
    "tato": {
        "password": "tato2026",
        "rol": "operador",
        "caja": "CAJA TATO",    # Solo puede ver y operar CAJA TATO
        "nombre": "Tato"
    },
    "mel": {
        "password": "congo2026",
        "rol": "operador",
        "caja": "CAJA JUNIN",    # Solo puede ver y operar CAJA JUNIN
        "nombre": "Mel"
    },
    # ── Para agregar más usuarios, descomentá y editá el bloque: ──
    # "nuevo_usuario": {
    #     "password": "password123",
    #     "rol": "operador",
    #     "caja": "CAJA NUEVO",
    #     "nombre": "Nombre Visible"
    # },
}

if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
    st.session_state.usuario_actual = None
    st.session_state.rol_actual = None
    st.session_state.caja_propia = None
    st.session_state.nombre_usuario = None

if not st.session_state.autenticado:
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        try: st.image("logo_path.png", width=250)
        except: st.title("🚛 CHACAGEST")
        u = st.text_input("Usuario")
        p = st.text_input("Contraseña", type="password")
        if st.button("INGRESAR"):
            u_lower = u.strip().lower()
            if u_lower in USUARIOS and USUARIOS[u_lower]["password"] == p.strip():
                datos_usuario = USUARIOS[u_lower]
                st.session_state.autenticado    = True
                st.session_state.usuario_actual = u_lower
                st.session_state.rol_actual     = datos_usuario["rol"]
                st.session_state.caja_propia    = datos_usuario["caja"]
                st.session_state.nombre_usuario = datos_usuario["nombre"]
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos")
    st.stop()

# Helpers de rol
es_admin    = st.session_state.rol_actual == "admin"
es_operador = st.session_state.rol_actual == "operador"
caja_propia = st.session_state.caja_propia

# --- 3. INICIALIZACIÓN ---

# Cuentas de gastos: se pueden gestionar desde el sistema
CUENTAS_GASTOS_DEFAULT = [
    "COMBUSTIBLE", "REPARACION", "REPUESTO", "SERVICIO LUZ, GAS",
    "ALQUILERES", "COMPRA ART. LIMPIEZA", "HONORARIOS",
    "PUBLICIDAD", "GASTOS OFICINA", "GASTOS EN LIBRERIA",
    "CHOFERES EVENTUALES", "VARIOS"
]

if 'cuentas_gastos' not in st.session_state:
    st.session_state.cuentas_gastos = list(CUENTAS_GASTOS_DEFAULT)

if 'clientes' not in st.session_state or 'viajes' not in st.session_state:
    c, v, p, t, prov, com, ce, cc, fac = cargar_datos()
    st.session_state.clientes          = c    if c    is not None else pd.DataFrame(columns=COL_CLIENTES)
    st.session_state.viajes            = v    if v    is not None else pd.DataFrame(columns=COL_VIAJES)
    st.session_state.presupuestos      = p    if p    is not None else pd.DataFrame(columns=COL_PRESUPUESTOS)
    st.session_state.tesoreria         = t    if t    is not None else pd.DataFrame(columns=COL_TESORERIA)
    st.session_state.proveedores       = prov if prov is not None else pd.DataFrame(columns=COL_PROVEEDORES)
    st.session_state.compras           = com  if com  is not None else pd.DataFrame(columns=COL_COMPRAS)
    st.session_state.cheques_emitidos  = ce   if ce   is not None else pd.DataFrame(columns=COL_CHEQ_EMITIDOS)
    st.session_state.cheques_cartera   = cc   if cc   is not None else pd.DataFrame(columns=COL_CHEQ_CARTERA)
    st.session_state.facturas          = fac  if fac  is not None else pd.DataFrame(columns=COL_FACTURAS)

# --- 4. DISEÑO ---
st.markdown("""
    <style>
    [data-testid="stSidebarNav"] { display: none; }
    header { visibility: hidden; }
    h1, h2, h3 { color: #5e2d61 !important; }
    div.stButton > button {
        background: linear-gradient(to right, #f39c12, #d35400) !important;
        color: white !important;
        border-radius: 8px !important;
        border: none !important;
        font-weight: bold !important;
    }
    .stDataFrame { border: 1px solid #5e2d61; border-radius: 5px; }
    </style>
    """, unsafe_allow_html=True)

# --- 5. SIDEBAR ---
with st.sidebar:
    try: st.image("logo_path.png", use_container_width=True)
    except: pass
    st.markdown("---")

    # ── Badge de usuario logueado ──
    rol_badge = "🔑 Admin" if es_admin else f"👤 {st.session_state.nombre_usuario}"
    caja_badge = "" if es_admin else f" | 🏦 {caja_propia}"
    st.markdown(f"<div style='background:#5e2d61;color:white;padding:8px 12px;border-radius:8px;font-size:13px;font-weight:bold;margin-bottom:8px;'>{rol_badge}{caja_badge}</div>", unsafe_allow_html=True)
    st.markdown("---")

    # ── Menú principal: Admin ve todo, Operador no ve Dashboard ──
    if es_admin:
        opciones_menu = ["CALENDARIO", "DASHBOARD", "VENTAS", "COMPRAS", "FACTURACION", "TESORERIA", "CHEQUES"]
        iconos_menu   = ["calendar3", "bar-chart-line", "cart4", "bag-check", "receipt-cutoff", "safe", "bank2"]
    else:
        opciones_menu = ["CALENDARIO", "VENTAS", "COMPRAS", "FACTURACION", "TESORERIA", "CHEQUES"]
        iconos_menu   = ["calendar3", "cart4", "bag-check", "receipt-cutoff", "safe", "bank2"]

    menu_principal = option_menu(
        menu_title=None,
        options=opciones_menu,
        icons=iconos_menu,
        default_index=0,
        key="menu_p",
        styles={
            "container": {"padding": "0px", "background-color": "#f0f2f6"},
            "nav-link": {"font-size": "15px", "font-weight": "bold"},
            "nav-link-selected": {"background-color": "#5e2d61"},
        }
    )

    sel_sub = None
    if menu_principal == "VENTAS":
        st.markdown("<div style='margin-left: 20px; border-left: 2px solid #f39c12; padding-left: 10px;'>", unsafe_allow_html=True)
        opciones_ventas = ["CLIENTES", "CARGA VIAJE", "PRESUPUESTOS", "CTA CTE INDIVIDUAL", "CTA CTE GENERAL", "COMPROBANTES"]
        iconos_ventas   = ["people", "truck", "file-earmark-spreadsheet", "person-vcard", "globe", "file-text"]
        sel_sub = option_menu(
            menu_title=None,
            options=opciones_ventas,
            icons=iconos_ventas,
            default_index=0,
            key="menu_s",
            styles={
                "container": {"background-color": "transparent", "padding": "0px"},
                "nav-link": {"font-size": "13px", "text-align": "left", "margin": "2px"},
                "nav-link-selected": {"background-color": "#f39c12", "color": "white"},
            }
        )
        st.markdown("</div>", unsafe_allow_html=True)

    elif menu_principal == "COMPRAS":
        st.markdown("<div style='margin-left: 20px; border-left: 2px solid #f39c12; padding-left: 10px;'>", unsafe_allow_html=True)
        opciones_compras = ["CARGA PROVEEDOR", "CARGA GASTOS", "CTA CTE PROVEEDOR", "CTA CTE GENERAL PROV", "HISTORICO COMPRAS", "MAYOR DE CUENTAS"]
        iconos_compras   = ["person-plus", "receipt", "person-lines-fill", "globe", "clock-history", "journal-text"]
        sel_sub = option_menu(
            menu_title=None,
            options=opciones_compras,
            icons=iconos_compras,
            default_index=0,
            key="menu_c",
            styles={
                "container": {"background-color": "transparent", "padding": "0px"},
                "nav-link": {"font-size": "13px", "text-align": "left", "margin": "2px"},
                "nav-link-selected": {"background-color": "#f39c12", "color": "white"},
            }
        )
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("---")

    # ── Gestión de Cuentas de Gastos (solo admin) ──
    if es_admin:
        with st.expander("⚙️ Cuentas de Gastos", expanded=False):
            st.markdown("<small><b>Agregar nueva cuenta:</b></small>", unsafe_allow_html=True)
            nueva_cta = st.text_input("Nueva cuenta", key="nueva_cta_input", label_visibility="collapsed", placeholder="Ej: SEGUROS")
            if st.button("➕ Agregar", key="btn_agregar_cta"):
                nueva_cta_upper = nueva_cta.strip().upper()
                if nueva_cta_upper and nueva_cta_upper not in st.session_state.cuentas_gastos:
                    st.session_state.cuentas_gastos.append(nueva_cta_upper)
                    st.session_state.cuentas_gastos.sort()
                    st.rerun()
                elif nueva_cta_upper in st.session_state.cuentas_gastos:
                    st.warning("Ya existe esa cuenta.")
            st.markdown("<small><b>Cuentas actuales:</b></small>", unsafe_allow_html=True)
            for cta in sorted(st.session_state.cuentas_gastos):
                col_cta, col_del = st.columns([0.8, 0.2])
                col_cta.markdown(f"<small>📂 {cta}</small>", unsafe_allow_html=True)
                if col_del.button("🗑️", key=f"del_cta_{cta}", help=f"Eliminar {cta}"):
                    if len(st.session_state.cuentas_gastos) > 1:
                        st.session_state.cuentas_gastos.remove(cta)
                        st.rerun()
                    else:
                        st.warning("Debe quedar al menos una cuenta.")

    st.markdown("---")
    if st.button("🔄 Sincronizar"):
        with st.spinner("Sincronizando..."):
            c, v, p, t, prov, com, ce, cc, fac = cargar_datos()
            st.session_state.clientes         = c    if c    is not None else pd.DataFrame(columns=COL_CLIENTES)
            st.session_state.viajes           = v    if v    is not None else pd.DataFrame(columns=COL_VIAJES)
            st.session_state.presupuestos     = p    if p    is not None else pd.DataFrame(columns=COL_PRESUPUESTOS)
            st.session_state.tesoreria        = t    if t    is not None else pd.DataFrame(columns=COL_TESORERIA)
            st.session_state.proveedores      = prov if prov is not None else pd.DataFrame(columns=COL_PROVEEDORES)
            st.session_state.compras          = com  if com  is not None else pd.DataFrame(columns=COL_COMPRAS)
            st.session_state.cheques_emitidos = ce   if ce   is not None else pd.DataFrame(columns=COL_CHEQ_EMITIDOS)
            st.session_state.cheques_cartera  = cc   if cc   is not None else pd.DataFrame(columns=COL_CHEQ_CARTERA)
            st.session_state.facturas         = fac  if fac  is not None else pd.DataFrame(columns=COL_FACTURAS)
            st.rerun()

    if st.button("🚪 Cerrar Sesión"):
        for key in ["autenticado", "usuario_actual", "rol_actual", "caja_propia", "nombre_usuario"]:
            st.session_state[key] = False if key == "autenticado" else None
        st.rerun()

# ── Definición de sel ── SIEMPRE después del sidebar
if menu_principal in ["VENTAS", "COMPRAS"]:
    sel = sel_sub
else:
    sel = menu_principal

# ── Bloqueo de seguridad: si operador intenta acceder a DASHBOARD vía URL ──
if sel == "DASHBOARD" and es_operador:
    st.error("🚫 Acceso denegado. Solo el administrador puede ver el Dashboard.")
    st.stop()

# --- 6. MÓDULOS ---

# =============================================================
# DASHBOARD
# =============================================================
if sel == "DASHBOARD":
    st.header("📊 Dashboard de Control Financiero")

    MESES_NOMBRES = {
        1:"Enero", 2:"Febrero", 3:"Marzo",    4:"Abril",
        5:"Mayo",  6:"Junio",   7:"Julio",     8:"Agosto",
        9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"
    }
    MESES_ORDEN = list(range(1, 13))
    MESES_LABEL = [cal_module.month_abbr[m] for m in MESES_ORDEN]

    # ── Preparar INGRESOS (viajes + facturas) ──
    df_ing = st.session_state.viajes.copy()
    df_ing = df_ing[df_ing['Importe'] > 0].copy()
    df_ing['Fecha Viaje'] = pd.to_datetime(df_ing['Fecha Viaje'], errors='coerce')
    df_ing = df_ing.dropna(subset=['Fecha Viaje'])
    df_ing['Año'] = df_ing['Fecha Viaje'].dt.year
    df_ing['Mes'] = df_ing['Fecha Viaje'].dt.month

    # Sumar facturas emitidas (tipo FACTURA, excluyendo NC/ND)
    if 'facturas' in st.session_state and not st.session_state.facturas.empty:
        df_fac_dash = st.session_state.facturas.copy()
        df_fac_dash = df_fac_dash[df_fac_dash['Tipo'] == 'FACTURA'].copy()
        df_fac_dash['Total'] = pd.to_numeric(df_fac_dash['Total'], errors='coerce').fillna(0)
        df_fac_dash = df_fac_dash[df_fac_dash['Total'] > 0]
        df_fac_dash['Fecha Viaje'] = pd.to_datetime(df_fac_dash['Fecha'], errors='coerce')
        df_fac_dash = df_fac_dash.dropna(subset=['Fecha Viaje'])
        df_fac_dash['Año'] = df_fac_dash['Fecha Viaje'].dt.year
        df_fac_dash['Mes'] = df_fac_dash['Fecha Viaje'].dt.month
        df_fac_dash = df_fac_dash.rename(columns={'Total': 'Importe'})
        df_fac_dash = df_fac_dash[['Fecha Viaje', 'Año', 'Mes', 'Importe']]
        df_ing = pd.concat([df_ing, df_fac_dash], ignore_index=True)

    # ── Preparar GASTOS ──
    df_gas = st.session_state.compras.copy()
    df_gas = df_gas[df_gas['Total'] > 0].copy()
    df_gas['Fecha'] = pd.to_datetime(df_gas['Fecha'], errors='coerce')
    df_gas = df_gas.dropna(subset=['Fecha'])
    df_gas['Año'] = df_gas['Fecha'].dt.year
    df_gas['Mes'] = df_gas['Fecha'].dt.month

    # Enriquecer gastos con Cuenta de Gastos del proveedor
    if not st.session_state.proveedores.empty:
        df_gas = df_gas.merge(
            st.session_state.proveedores[['Razón Social', 'Cuenta de Gastos']],
            left_on='Proveedor', right_on='Razón Social', how='left'
        )
        df_gas['Cuenta de Gastos'] = df_gas['Cuenta de Gastos'].fillna('SIN CATEGORÍA')
    else:
        df_gas['Cuenta de Gastos'] = 'SIN CATEGORÍA'

    # ── Años disponibles ──
    años_ing  = set(df_ing['Año'].unique()) if not df_ing.empty else set()
    años_gas  = set(df_gas['Año'].unique()) if not df_gas.empty else set()
    años_disp = sorted(años_ing | años_gas, reverse=True)
    if not años_disp:
        años_disp = [date.today().year]

    # ── Selectores ──
    col_v1, col_v2, col_v3 = st.columns([1, 1, 2])
    vista   = col_v1.radio("Vista", ["Mensual", "Anual"], horizontal=True)
    año_sel = col_v2.selectbox("Año", años_disp)

    mes_sel = None
    if vista == "Mensual":
        mes_sel = col_v3.selectbox(
            "Mes",
            options=list(MESES_NOMBRES.keys()),
            format_func=lambda x: MESES_NOMBRES[x],
            index=date.today().month - 1
        )

    # ── Filtrar ──
    if vista == "Mensual":
        df_ing_f = df_ing[(df_ing['Año'] == año_sel) & (df_ing['Mes'] == mes_sel)]
        df_gas_f = df_gas[(df_gas['Año'] == año_sel) & (df_gas['Mes'] == mes_sel)]
        titulo_periodo = f"{MESES_NOMBRES[mes_sel]} {año_sel}"
    else:
        df_ing_f = df_ing[df_ing['Año'] == año_sel]
        df_gas_f = df_gas[df_gas['Año'] == año_sel]
        titulo_periodo = f"Año {año_sel}"

    total_ing = df_ing_f['Importe'].sum()
    total_gas = df_gas_f['Total'].sum()
    resultado = total_ing - total_gas
    margen    = (resultado / total_ing * 100) if total_ing > 0 else 0

    st.markdown(f"### Período: {titulo_periodo}")
    st.markdown("---")

    # ── KPIs ──
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("💰 Total Ingresos", f"$ {total_ing:,.0f}")
    k2.metric("💸 Total Gastos",   f"$ {total_gas:,.0f}")
    k3.metric(
        "📈 Resultado Neto",
        f"$ {resultado:,.0f}",
        delta=f"{'▲' if resultado >= 0 else '▼'} {abs(resultado):,.0f}",
        delta_color="normal" if resultado >= 0 else "inverse"
    )
    k4.metric("📊 Margen", f"{margen:.1f} %")

    st.markdown("---")

    COLORES = ["#5e2d61", "#f39c12", "#d35400", "#2ecc71",
               "#3498db", "#e74c3c", "#9b59b6", "#1abc9c"]

    col_g1, col_g2 = st.columns(2)

    # ── Gráfico 1: Ingresos vs Gastos ──
    with col_g1:
        if vista == "Anual":
            ing_mes = df_ing_f.groupby('Mes')['Importe'].sum().reindex(MESES_ORDEN, fill_value=0)
            gas_mes = df_gas_f.groupby('Mes')['Total'].sum().reindex(MESES_ORDEN, fill_value=0)
            fig1 = go.Figure()
            fig1.add_trace(go.Bar(name="Ingresos", x=MESES_LABEL, y=ing_mes.values,  marker_color="#5e2d61"))
            fig1.add_trace(go.Bar(name="Gastos",   x=MESES_LABEL, y=gas_mes.values,  marker_color="#f39c12"))
            fig1.update_layout(
                title=f"Ingresos vs Gastos — {año_sel}",
                barmode='group', plot_bgcolor='white',
                legend=dict(orientation="h", y=-0.2),
                yaxis_tickprefix="$", margin=dict(t=40, b=10)
            )
        else:
            dias_mes  = cal_module.monthrange(año_sel, mes_sel)[1]
            todos_dias = list(range(1, dias_mes + 1))
            df_ing_d  = df_ing_f.copy(); df_ing_d['Dia'] = df_ing_d['Fecha Viaje'].dt.day
            df_gas_d  = df_gas_f.copy(); df_gas_d['Dia'] = df_gas_d['Fecha'].dt.day
            ing_dia   = df_ing_d.groupby('Dia')['Importe'].sum().reindex(todos_dias, fill_value=0)
            gas_dia   = df_gas_d.groupby('Dia')['Total'].sum().reindex(todos_dias, fill_value=0)
            fig1 = go.Figure()
            fig1.add_trace(go.Bar(name="Ingresos", x=todos_dias, y=ing_dia.values, marker_color="#5e2d61"))
            fig1.add_trace(go.Bar(name="Gastos",   x=todos_dias, y=gas_dia.values, marker_color="#f39c12"))
            fig1.update_layout(
                title=f"Ingresos vs Gastos por Día — {MESES_NOMBRES[mes_sel]} {año_sel}",
                barmode='group', plot_bgcolor='white',
                legend=dict(orientation="h", y=-0.2),
                xaxis_title="Día", yaxis_tickprefix="$",
                margin=dict(t=40, b=10)
            )
        st.plotly_chart(fig1, use_container_width=True)

    # ── Gráfico 2: Torta de gastos por categoría ──
    with col_g2:
        if not df_gas_f.empty:
            gas_cat = df_gas_f.groupby('Cuenta de Gastos')['Total'].sum().reset_index()
            fig2 = px.pie(
                gas_cat, values='Total', names='Cuenta de Gastos',
                title=f"Gastos por Categoría — {titulo_periodo}",
                color_discrete_sequence=COLORES, hole=0.4
            )
            fig2.update_traces(textposition='inside', textinfo='percent+label')
            fig2.update_layout(legend=dict(orientation="h", y=-0.2), margin=dict(t=40, b=10))
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Sin gastos registrados para el período seleccionado.")

    col_g3, col_g4 = st.columns(2)

    # ── Gráfico 3: Tendencia (anual) o Top clientes (mensual) ──
    with col_g3:
        if vista == "Anual":
            ing_mes   = df_ing_f.groupby('Mes')['Importe'].sum().reindex(MESES_ORDEN, fill_value=0)
            gas_mes   = df_gas_f.groupby('Mes')['Total'].sum().reindex(MESES_ORDEN, fill_value=0)
            res_mes   = ing_mes - gas_mes
            fig3 = go.Figure()
            fig3.add_trace(go.Scatter(
                x=MESES_LABEL, y=res_mes.values,
                mode='lines+markers+text',
                text=[f"${v:,.0f}" for v in res_mes.values],
                textposition="top center",
                line=dict(color="#5e2d61", width=3),
                marker=dict(size=8, color="#f39c12"),
                fill='tozeroy', fillcolor='rgba(94,45,97,0.1)',
                name="Resultado Neto"
            ))
            fig3.add_hline(y=0, line_dash="dash", line_color="red", opacity=0.5)
            fig3.update_layout(
                title="Tendencia del Resultado Neto",
                plot_bgcolor='white', yaxis_tickprefix="$",
                margin=dict(t=40, b=10)
            )
            st.plotly_chart(fig3, use_container_width=True)
        else:
            if not df_ing_f.empty:
                top_cli = (df_ing_f.groupby('Cliente')['Importe'].sum()
                           .reset_index().sort_values('Importe').tail(5))
                fig3 = go.Figure(go.Bar(
                    x=top_cli['Importe'], y=top_cli['Cliente'],
                    orientation='h', marker_color="#5e2d61",
                    text=[f"$ {v:,.0f}" for v in top_cli['Importe']],
                    textposition='outside'
                ))
                fig3.update_layout(
                    title="Top 5 Clientes del Mes",
                    plot_bgcolor='white', xaxis_tickprefix="$",
                    margin=dict(t=40, b=10)
                )
                st.plotly_chart(fig3, use_container_width=True)
            else:
                st.info("Sin ingresos en el período.")

    # ── Gráfico 4: Mapa de calor (anual) o Barras categoría (mensual) ──
    with col_g4:
        if not df_gas_f.empty:
            if vista == "Anual":
                pivot = df_gas_f.pivot_table(
                    index='Cuenta de Gastos', columns='Mes',
                    values='Total', aggfunc='sum', fill_value=0
                )
                pivot.columns = [cal_module.month_abbr[m] for m in pivot.columns]
                fig4 = px.imshow(
                    pivot,
                    color_continuous_scale=[[0,"#fff4e6"],[0.5,"#f39c12"],[1,"#d35400"]],
                    title="Mapa de Gastos por Categoría y Mes",
                    text_auto=True, aspect="auto"
                )
                fig4.update_layout(margin=dict(t=40, b=10))
            else:
                gas_c = (df_gas_f.groupby('Cuenta de Gastos')['Total'].sum()
                         .reset_index().sort_values('Total'))
                fig4 = go.Figure(go.Bar(
                    x=gas_c['Total'], y=gas_c['Cuenta de Gastos'],
                    orientation='h', marker_color="#f39c12",
                    text=[f"$ {v:,.0f}" for v in gas_c['Total']],
                    textposition='outside'
                ))
                fig4.update_layout(
                    title="Gastos por Categoría del Mes",
                    plot_bgcolor='white', xaxis_tickprefix="$",
                    margin=dict(t=40, b=10)
                )
            st.plotly_chart(fig4, use_container_width=True)
        else:
            st.info("Sin gastos registrados en el período.")

    # ── Tabla resumen ──
    st.markdown("---")
    st.subheader("📋 Resumen por Categoría de Gasto")
    if not df_gas_f.empty:
        res_cat = df_gas_f.groupby('Cuenta de Gastos')['Total'].agg(
            Total='sum', Comprobantes='count'
        ).reset_index().sort_values('Total', ascending=False)
        res_cat['% del Total'] = (res_cat['Total'] / res_cat['Total'].sum() * 100).round(1).astype(str) + " %"
        res_cat['Total']       = res_cat['Total'].apply(lambda x: f"$ {x:,.2f}")
        res_cat.columns        = ['Categoría', 'Total Gastado', 'N° Comprobantes', '% del Total']
        st.dataframe(res_cat, use_container_width=True, hide_index=True)
    else:
        st.info("Sin datos de gastos para el período.")

    st.caption(f"Última actualización: {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')} hs")

# =============================================================
# CALENDARIO
# =============================================================
elif sel == "CALENDARIO":
    st.header("📅 Agenda de Viajes")
    if "viaje_ver" not in st.session_state:
        st.session_state.viaje_ver = None
    eventos = []
    df_solo_viajes = st.session_state.viajes[st.session_state.viajes['Importe'] > 0]
    for i, row in df_solo_viajes.iterrows():
        if str(row['Fecha Viaje']) != "-" and row['Origen'] != "AJUSTE":
            eventos.append({
                "id": str(i), "title": f"🚛 {row['Cliente']}", "start": str(row['Fecha Viaje']),
                "allDay": True, "backgroundColor": "#f39c12", "borderColor": "#d35400"
            })
    cal_options  = {"headerToolbar": {"left": "prev,next today", "center": "title", "right": "dayGridMonth"}, "locale": "es", "height": 600}
    custom_css   = ".fc-button-primary { background-color: #5e2d61 !important; border-color: #5e2d61 !important; } .fc-event { background-color: #f39c12 !important; } .fc-toolbar-title { color: #5e2d61 !important; }"
    res_cal      = calendar(events=eventos, options=cal_options, custom_css=custom_css, key="cal_final")
    if res_cal.get("eventClick"):
        st.session_state.viaje_ver = int(res_cal["eventClick"]["event"]["id"])
    if st.session_state.viaje_ver is not None:
        idx = st.session_state.viaje_ver
        if idx in st.session_state.viajes.index:
            v_det = st.session_state.viajes.loc[idx]
            if st.button("❌ Cerrar"): st.session_state.viaje_ver = None; st.rerun()
            st.markdown(f"""<div style="background-color: #f0f2f6; padding: 15px; border-left: 5px solid #f39c12; border-radius: 5px; margin-top: 20px;">
                <h4 style="color: #5e2d61; margin: 0;">Detalles</h4><p><b>Cliente:</b> {v_det['Cliente']}</p>
                <p><b>Ruta:</b> {v_det['Origen']} ➔ {v_det['Destino']}</p>
                <p><b>Importe:</b> $ {v_det['Importe']}</p></div>""", unsafe_allow_html=True)

elif sel == "CLIENTES":
    st.header("👤 Gestión de Clientes")
    if st.session_state.get("msg_cliente"):
        st.success(st.session_state.msg_cliente)
        st.session_state.msg_cliente = None
    with st.expander("➕ ALTA DE NUEVO CLIENTE", expanded=False):
        with st.form("f_cli", clear_on_submit=True):
            c1, c2 = st.columns(2)
            r      = c1.text_input("Razón Social *")
            cuit   = c2.text_input("CUIT *")
            mail   = c1.text_input("Email")
            tel    = c2.text_input("Teléfono")
            dir_f  = c1.text_input("Dirección Fiscal")
            loc    = c2.text_input("Localidad")
            prov   = c1.text_input("Provincia")
            c_iva  = c2.selectbox("Condición IVA", ["Responsable Inscripto", "Monotributo", "Exento", "Consumidor Final"])
            c_vta  = c1.selectbox("Condición de Venta", ["Cuenta Corriente", "Contado"])
            if st.form_submit_button("REGISTRAR CLIENTE"):
                if r and cuit:
                    nueva_fila = pd.DataFrame([[r, cuit, mail, tel, dir_f, loc, prov, c_iva, c_vta]], columns=COL_CLIENTES)
                    st.session_state.clientes = pd.concat([st.session_state.clientes, nueva_fila], ignore_index=True)
                    guardar_datos("clientes", st.session_state.clientes)
                    st.session_state.msg_cliente = f"✅ Cliente '{r}' registrado correctamente."
                    st.rerun()
                else:
                    st.warning("Completá Razón Social y CUIT para continuar.")
    st.subheader("📋 Base de Clientes")
    if not st.session_state.clientes.empty:
        for i, row in st.session_state.clientes.iterrows():
            with st.container():
                c_inf, c_ed, c_el = st.columns([0.7, 0.15, 0.15])
                c_inf.markdown(f"**{row['Razón Social']}** | CUIT: {row['CUIT / CUIL / DNI *']}")
                c_inf.caption(f"📍 {row['Localidad']} - {row['Provincia']} | 📞 {row['Teléfono']}")
                if c_ed.button("📝 Editar", key=f"edit_{i}"): st.session_state[f"edit_mode_{i}"] = True
                if c_el.button("🗑️", key=f"del_cli_{i}"):
                    tiene_viajes = not st.session_state.viajes[st.session_state.viajes['Cliente'] == row['Razón Social']].empty
                    if tiene_viajes: st.error("No se puede eliminar: tiene viajes asociados.")
                    else:
                        st.session_state.clientes = st.session_state.clientes.drop(i).reset_index(drop=True)
                        guardar_datos("clientes", st.session_state.clientes)
                        st.rerun()
                if st.session_state.get(f"edit_mode_{i}", False):
                    with st.form(f"f_edit_{i}"):
                        ce1, ce2 = st.columns(2)
                        n_rs   = ce1.text_input("Razón Social", value=row['Razón Social'])
                        n_cuit = ce2.text_input("CUIT", value=row['CUIT / CUIL / DNI *'])
                        n_mail = ce1.text_input("Email", value=row['Email'])
                        n_tel  = ce2.text_input("Teléfono", value=row['Teléfono'])
                        n_loc  = ce1.text_input("Localidad", value=row['Localidad'])
                        n_prov = ce2.text_input("Provincia", value=row['Provincia'])
                        be1, be2 = st.columns(2)
                        if be1.form_submit_button("✅ Guardar"):
                            st.session_state.clientes.loc[i] = [n_rs, n_cuit, n_mail, n_tel, row['Dirección Fiscal'], n_loc, n_prov, row['Condición IVA'], row['Condición de Venta']]
                            guardar_datos("clientes", st.session_state.clientes)
                            st.session_state[f"edit_mode_{i}"] = False
                            st.rerun()
                        if be2.form_submit_button("❌ Cancelar"): st.session_state[f"edit_mode_{i}"] = False; st.rerun()
                st.divider()
    else: st.info("No hay clientes registrados.")

elif sel == "CARGA VIAJE":
    st.header("🚛 Registro de Viaje")
    if st.session_state.get("msg_viaje"):
        st.success(st.session_state.msg_viaje)
        st.session_state.msg_viaje = None
    with st.form("f_v", clear_on_submit=True):
        cli  = st.selectbox("Seleccionar Cliente", st.session_state.clientes['Razón Social'].unique() if not st.session_state.clientes.empty else [""])
        c1, c2 = st.columns(2)
        f_v  = c1.date_input("Fecha")
        pat  = c2.text_input("Patente")
        orig = st.text_input("Origen")
        dest = st.text_input("Destino")
        imp  = st.number_input("Importe Neto $", min_value=0.0)
        cond = st.selectbox("Tipo de Pago", ["Cuenta Corriente", "Contado"])
        if st.form_submit_button("GUARDAR VIAJE"):
            if cli and imp > 0:
                nv = pd.DataFrame([[date.today(), cli, f_v, orig, dest, pat, imp, f"Factura ({cond})", "-"]], columns=COL_VIAJES)
                st.session_state.viajes = pd.concat([st.session_state.viajes, nv], ignore_index=True)
                guardar_datos("viajes", st.session_state.viajes)
                st.session_state.msg_viaje = f"✅ Viaje de '{cli}' registrado correctamente por $ {imp:,.2f}."
                st.rerun()
            else:
                st.warning("Seleccioná un cliente y completá el importe.")

elif sel == "PRESUPUESTOS":
    st.header("📝 Gestión de Presupuestos")
    tab_crear, tab_historial = st.tabs(["🆕 Crear Presupuesto", "📂 Historial y Descargas"])
    with tab_crear:
        if st.session_state.get("msg_presupuesto"):
            st.success(st.session_state.msg_presupuesto)
            st.session_state.msg_presupuesto = None
        with st.form("f_presu", clear_on_submit=True):
            c1, c2   = st.columns(2)
            p_cli    = c1.selectbox("Cliente", st.session_state.clientes['Razón Social'].unique() if not st.session_state.clientes.empty else [""])
            p_f_emi  = c2.date_input("Fecha Emisión", date.today())
            c3, c4   = st.columns(2)
            p_f_venc = c3.date_input("Fecha Vencimiento", date.today() + timedelta(days=7))
            p_movil  = c4.selectbox("Tipo de Móvil", ["Combi 19 asientos", "Minibus 24 asientos", "Micro 45 asientos", "Micro 60 asientos"])
            p_det    = st.text_area("Detalle del Presupuesto (Servicio, Ruta, Horarios...)")
            p_imp    = st.number_input("Importe Total $", min_value=0.0)
            if st.form_submit_button("GENERAR PRESUPUESTO"):
                if p_cli and p_imp > 0:
                    nuevo_p = pd.DataFrame([[p_f_emi, p_cli, p_f_venc, p_det, p_movil, p_imp]], columns=COL_PRESUPUESTOS)
                    st.session_state.presupuestos = pd.concat([st.session_state.presupuestos, nuevo_p], ignore_index=True)
                    guardar_datos("presupuestos", st.session_state.presupuestos)
                    st.session_state.msg_presupuesto = f"✅ Presupuesto para '{p_cli}' guardado por $ {p_imp:,.2f}."
                    st.rerun()
                else:
                    st.warning("Seleccioná cliente y completá el importe.")
    with tab_historial:
        if not st.session_state.presupuestos.empty:
            for i in reversed(st.session_state.presupuestos.index):
                row_p = st.session_state.presupuestos.loc[i]
                with st.container():
                    c_a, c_b, c_c = st.columns([0.6, 0.2, 0.2])
                    c_a.markdown(f"**{row_p['Cliente']}** | {row_p['Tipo Móvil']}")
                    c_a.caption(f"Emisión: {row_p['Fecha Emisión']} - Vence: {row_p['Vencimiento']}")
                    c_b.markdown(f"**$ {row_p['Importe']:,.2f}**")
                    html_p = generar_html_presupuesto(row_p)
                    c_c.download_button(label="📄 Descargar", data=html_p, file_name=f"Presupuesto_{row_p['Cliente']}_{row_p['Fecha Emisión']}.html", mime="text/html", key=f"dl_p_{i}")
                    if c_c.button("🗑️", key=f"del_presu_{i}"):
                        st.session_state.presupuestos = st.session_state.presupuestos.drop(i)
                        guardar_datos("presupuestos", st.session_state.presupuestos)
                        st.rerun()
                    st.divider()
        else: st.info("No hay presupuestos registrados.")

elif sel == "TESORERIA":
    st.header("💰 Tesorería")

    # ── Cajas disponibles según rol ──
    # Admin ve todas. Operador solo ve su caja asignada.
    TODAS_CAJAS = ["CAJA COTI", "CAJA TATO", "CAJA JUNIN", "BANCO GALICIA", "BANCO PROVINCIA", "TARJETA DE CREDITO", "BANCO SUPERVIELLE", "DOLAR CAJA COTI", "DOLAR CAJA TATO"]

    if es_admin:
        opc_cajas = TODAS_CAJAS
        FORMAS_PAGO = ["EFECTIVO", "TRANSFERENCIA", "TARJETA DE CREDITO", "DÓLARES"]
    else:
        opc_cajas   = [caja_propia]
        FORMAS_PAGO = ["EFECTIVO", "TRANSFERENCIA", "TARJETA DE CREDITO", "DÓLARES"]
        st.info(f"🏦 Operando en: **{caja_propia}**")

    # ── Tabs: Admin ve todos, Operador no ve Traspaso ni Orden de Pago
    #         pero SÍ ve "Pase de Efectivo" (puede pasar efectivo a otra caja) ──
    if es_admin:
        tab_ing, tab_egr, tab_cob, tab_cob_fac, tab_ver, tab_pase, tab_cierre, tab_tras, tab_op = st.tabs(
            ["📥 INGRESOS VARIOS", "📤 EGRESOS VARIOS", "🧾 COBRANZA VIAJE", "🧾 COBRANZA FACTURA", "📊 VER MOVIMIENTOS", "💱 PASE DE EFECTIVO", "📋 RENDICIÓN", "🔄 TRASPASO", "💸 ORDEN DE PAGO"]
        )
    else:
        tab_ing, tab_egr, tab_cob, tab_cob_fac, tab_ver, tab_pase, tab_cierre, tab_op = st.tabs(
            ["📥 INGRESOS VARIOS", "📤 EGRESOS VARIOS", "🧾 COBRANZA VIAJE", "🧾 COBRANZA FACTURA", "📊 MIS MOVIMIENTOS", "💱 PASE DE EFECTIVO", "📋 RENDICIÓN", "💸 ORDEN DE PAGO"]
        )
        tab_tras = None

    with tab_ing:
        if st.session_state.get("msg_ingreso"):
            st.success(st.session_state.msg_ingreso)
            st.session_state.msg_ingreso = None
        with st.form("f_ing_var", clear_on_submit=True):
            f   = st.date_input("Fecha", date.today())
            # Admin elige caja; operador tiene la suya fija
            if es_admin:
                cj  = st.selectbox("Caja Destino", opc_cajas)
            else:
                st.markdown(f"**Caja:** {caja_propia}")
                cj = caja_propia
            forma = st.selectbox("Forma de Ingreso", FORMAS_PAGO)
            con = st.text_input("Concepto")
            mon = st.number_input("Monto $", min_value=0.0)
            if st.form_submit_button("REGISTRAR INGRESO"):
                if mon > 0:
                    concepto_completo = con if con else "-"
                    nt = pd.DataFrame([[f, "INGRESO VARIO", cj, forma, concepto_completo, "Varios", mon, "-"]], columns=COL_TESORERIA)
                    st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, nt], ignore_index=True)
                    guardar_tesoreria_rerun("msg_ingreso", f"✅ Ingreso de $ {mon:,.2f} ({forma}) registrado en {cj}.")
                else:
                    st.warning("Ingresá un monto mayor a cero.")

    with tab_egr:
        if st.session_state.get("msg_egreso"):
            st.success(st.session_state.msg_egreso)
            st.session_state.msg_egreso = None
        with st.form("f_egr_var", clear_on_submit=True):
            f   = st.date_input("Fecha", date.today())
            if es_admin:
                cj  = st.selectbox("Caja Origen", opc_cajas)
            else:
                st.markdown(f"**Caja:** {caja_propia}")
                cj = caja_propia
            forma = st.selectbox("Forma de Egreso", FORMAS_PAGO)
            cuentas_opciones = sorted(st.session_state.get("cuentas_gastos", ["SIN CATEGORÍA"])) + ["SIN CATEGORÍA"]
            cuentas_opciones = sorted(set(cuentas_opciones))
            cuenta_gasto = st.selectbox("Cuenta de Gasto", cuentas_opciones)
            con = st.text_input("Concepto")
            mon = st.number_input("Monto $", min_value=0.0)
            if st.form_submit_button("REGISTRAR EGRESO"):
                if mon > 0:
                    concepto_completo = f"[{cuenta_gasto}] {con}" if con else f"[{cuenta_gasto}]"
                    nt = pd.DataFrame([[f, "EGRESO VARIO", cj, forma, concepto_completo, "Varios", -mon, "-"]], columns=COL_TESORERIA)
                    st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, nt], ignore_index=True)
                    guardar_tesoreria_rerun("msg_egreso", f"✅ Egreso de $ {mon:,.2f} ({forma}) — {cuenta_gasto} — registrado desde {cj}.")
                else:
                    st.warning("Ingresá un monto mayor a cero.")

    with tab_cob:
        if "html_recibo_ready" not in st.session_state: st.session_state.html_recibo_ready = None

        FORMAS_COBRO_VIAJE = FORMAS_PAGO + ["CHEQUE DE TERCEROS", "OTROS"]

        if not st.session_state.html_recibo_ready:
            # ── Selectbox de cliente y forma FUERA del form para mostrar viajes ──
            cob_col1, cob_col2 = st.columns(2)
            c_sel_prev = cob_col1.selectbox("Cliente", st.session_state.clientes['Razón Social'].unique() if not st.session_state.clientes.empty else [""], key="cob_cli_sel")
            forma_cob_prev = cob_col2.selectbox("Forma de Cobro", FORMAS_COBRO_VIAJE, key="cob_forma_sel")

            # ── VIAJES PENDIENTES DEL CLIENTE ──
            # Lógica: distribuir los pagos (negativos) contra los viajes (positivos)
            # ordenados por fecha, y solo mostrar los que tienen saldo > 0.
            monto_sugerido_cob = 0.0
            viajes_desc_sel = []
            if c_sel_prev and not st.session_state.viajes.empty:
                df_todos_cli = st.session_state.viajes[
                    st.session_state.viajes['Cliente'] == c_sel_prev
                ].copy()
                df_viajes_cli = df_todos_cli[df_todos_cli['Importe'] > 0].copy()
                pagos_cli_total = df_todos_cli[df_todos_cli['Importe'] < 0]['Importe'].sum()
                saldo_viajes_cli = df_viajes_cli['Importe'].sum() if not df_viajes_cli.empty else 0.0
                saldo_neto_cli = round(saldo_viajes_cli + pagos_cli_total, 2)

                # Calcular saldo pendiente por viaje (distribuir pagos FIFO)
                saldo_restante_pagos = abs(pagos_cli_total)
                viajes_con_saldo = []
                for idx, vrow in df_viajes_cli.sort_values('Fecha Viaje').iterrows():
                    imp = float(vrow['Importe'])
                    if saldo_restante_pagos >= imp:
                        saldo_restante_pagos -= imp
                        # Viaje totalmente pagado → no mostrar
                    else:
                        saldo_pendiente_v = imp - saldo_restante_pagos
                        saldo_restante_pagos = 0.0
                        viajes_con_saldo.append((idx, vrow, round(saldo_pendiente_v, 2)))

                if viajes_con_saldo:
                    st.markdown("---")
                    st.markdown(f"##### 🚛 Viajes pendientes de cobro — {c_sel_prev}")
                    ms1, ms2, ms3 = st.columns(3)
                    ms1.metric("Total Viajes", f"$ {saldo_viajes_cli:,.2f}")
                    ms2.metric("Ya Cobrado",   f"$ {abs(pagos_cli_total):,.2f}")
                    ms3.metric("Saldo Pendiente", f"$ {saldo_neto_cli:,.2f}")
                    st.markdown("---")
                    st.markdown("**Seleccioná los viajes a imputar:**")
                    viajes_chequeados = {}
                    for vidx, vrow, saldo_v in viajes_con_saldo:
                        fecha_v   = vrow.get('Fecha Viaje', '-')
                        origen_v  = vrow.get('Origen', '-')
                        destino_v = vrow.get('Destino', '-')
                        nro_comp_v = vrow.get('Nro Comp Asoc', '-')
                        patente_v  = vrow.get('Patente / Móvil', '-')
                        imp_orig   = float(vrow['Importe'])
                        if saldo_v < imp_orig:
                            label_v = f"📅 {fecha_v} | {origen_v} → {destino_v} | 🚐 {patente_v} | $ {saldo_v:,.2f} (saldo de $ {imp_orig:,.2f})"
                        else:
                            label_v = f"📅 {fecha_v} | {origen_v} → {destino_v} | 🚐 {patente_v} | $ {saldo_v:,.2f}"
                        if str(nro_comp_v) not in ['-', '', 'nan']:
                            label_v += f" | Comp: {nro_comp_v}"
                        chk_v = st.checkbox(label_v, key=f"chk_viaje_{vidx}", value=False)
                        viajes_chequeados[vidx] = (chk_v, saldo_v, label_v)

                    monto_sugerido_cob = sum(v for c, v, l in viajes_chequeados.values() if c)
                    viajes_desc_sel    = [l for c, v, l in viajes_chequeados.values() if c]

                    if monto_sugerido_cob > 0:
                        st.markdown(
                            f"<div style='background:#eafaf1;border:2px solid #27ae60;border-radius:8px;"
                            f"padding:12px 20px;margin:10px 0;display:flex;justify-content:space-between;'>"
                            f"<b>Total seleccionado a cobrar:</b>"
                            f"<b style='color:#27ae60;font-size:18px;'>$ {monto_sugerido_cob:,.2f}</b></div>",
                            unsafe_allow_html=True
                        )
                    st.markdown("---")
                elif not df_viajes_cli.empty:
                    st.success(f"✅ {c_sel_prev} no tiene viajes pendientes de cobro.")
                else:
                    st.info(f"No hay viajes registrados para {c_sel_prev}.")

            es_cheque = (forma_cob_prev == "CHEQUE DE TERCEROS")

            with st.form("f_cob", clear_on_submit=True):
                # ── Datos base ──
                fb1, fb2 = st.columns(2)
                if es_admin:
                    cj = fb1.selectbox("Caja Destino", opc_cajas)
                else:
                    cj = caja_propia
                    fb1.markdown(f"**Caja:** {caja_propia}")
                mon  = fb2.number_input("Monto $", min_value=0.0, value=float(round(monto_sugerido_cob, 2)), step=100.0, format="%.2f")
                if viajes_desc_sel:
                    st.markdown(f"*Viajes seleccionados: {len(viajes_desc_sel)}*")
                afip = st.text_input("Comprobante Asociado (AFIP/Recibo)")

                # ── Campos del cheque: aparecen solo si forma = CHEQUE DE TERCEROS ──
                if es_cheque:
                    st.markdown("---")
                    st.markdown("##### 🏦 Datos del Cheque Recibido")
                    ch1, ch2 = st.columns(2)
                    ch_nro       = ch1.text_input("Nro. de Cheque *")
                    ch_tipo      = ch2.selectbox("Tipo", ["COMÚN", "DIFERIDO", "ELECTRÓNICO"])
                    ch3, ch4     = st.columns(2)
                    ch_banco     = ch3.text_input("Banco Librador *")
                    ch_librador  = ch4.text_input("Librador (titular del cheque) *")
                    ch5, ch6     = st.columns(2)
                    ch_recibido  = ch5.text_input("Recibido de (quien entrega el cheque)")
                    ch_femision  = ch6.date_input("Fecha de Emisión", date.today())
                    ch7, ch8     = st.columns(2)
                    ch_fvenc     = ch7.date_input("Fecha de Vencimiento / Pago *", date.today() + timedelta(days=30))
                    ch_obs       = ch8.text_input("Observaciones")
                else:
                    ch_nro = ch_tipo = ch_banco = ch_librador = ""
                    ch_recibido = ch_obs = ""
                    ch_fvenc = date.today()
                    ch_femision = date.today()

                if st.form_submit_button("✅ GUARDAR COBRANZA"):
                    c_sel     = c_sel_prev
                    forma_cob = forma_cob_prev
                    if c_sel and mon > 0:
                        if es_cheque:
                            if not ch_nro or not ch_banco or not ch_librador:
                                st.warning("Completá Nro. de Cheque, Banco Librador y Librador para continuar.")
                            else:
                                # 1) Registrar en cheques_cartera
                                nuevo_cheq = pd.DataFrame([[
                                    str(date.today()), ch_nro, ch_tipo, ch_banco, ch_librador,
                                    mon, str(ch_fvenc), "EN CARTERA", "-", "-",
                                    f"Emisión:{ch_femision} | Recibido de:{ch_recibido if ch_recibido else ch_librador} | {ch_obs}"
                                ]], columns=COL_CHEQ_CARTERA)
                                st.session_state.cheques_cartera = pd.concat([st.session_state.cheques_cartera, nuevo_cheq], ignore_index=True)
                                guardar_datos("cheques_cartera", st.session_state.cheques_cartera)
                                # 2) Tesorería
                                nt = pd.DataFrame([[
                                    str(date.today()), "COBRANZA", cj, "CHEQUE TERCERO",
                                    f"Cobro Viaje - Cheque #{ch_nro} de {ch_librador}",
                                    c_sel, mon, afip
                                ]], columns=COL_TESORERIA)
                                st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, nt], ignore_index=True)
                                guardar_datos("tesoreria", st.session_state.tesoreria)
                                # 3) Viajes (cta cte cliente)
                                nv = pd.DataFrame([[
                                    date.today(), c_sel, date.today(), "PAGO", "TESORERIA",
                                    "-", -mon, "RECIBO", afip
                                ]], columns=COL_VIAJES)
                                st.session_state.viajes = pd.concat([st.session_state.viajes, nv], ignore_index=True)
                                guardar_datos("viajes", st.session_state.viajes)
                                # 4) Recibo
                                st.session_state.html_recibo_ready = generar_html_recibo({
                                    "Fecha": date.today(), "Cliente/Proveedor": c_sel,
                                    "Concepto": f"Cobro de Viaje — Cheque #{ch_nro} ({ch_tipo}) | Librador: {ch_librador} | Vence: {ch_fvenc}",
                                    "Caja/Banco": f"{cj} - CHEQUE DE TERCEROS",
                                    "Monto": mon, "Ref AFIP": afip
                                })
                                st.session_state.cli_ready = c_sel
                                st.rerun()
                        else:
                            nt = pd.DataFrame([[date.today(), "COBRANZA", cj, forma_cob, "Cobro Viaje", c_sel, mon, afip]], columns=COL_TESORERIA)
                            nv = pd.DataFrame([[date.today(), c_sel, date.today(), "PAGO", "TESORERIA", "-", -mon, "RECIBO", afip]], columns=COL_VIAJES)
                            st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, nt], ignore_index=True)
                            st.session_state.viajes    = pd.concat([st.session_state.viajes, nv], ignore_index=True)
                            guardar_datos("tesoreria", st.session_state.tesoreria)
                            guardar_datos("viajes", st.session_state.viajes)
                            st.session_state.html_recibo_ready = generar_html_recibo({
                                "Fecha": date.today(), "Cliente/Proveedor": c_sel,
                                "Concepto": "Cobro de Viaje", "Caja/Banco": f"{cj} - {forma_cob}",
                                "Monto": mon, "Ref AFIP": afip
                            })
                            st.session_state.cli_ready = c_sel
                            st.rerun()
                    else:
                        st.warning("Completá el cliente y el monto antes de continuar.")

        if st.session_state.html_recibo_ready:
            st.success(f"✅ Cobranza de '{st.session_state.cli_ready}' registrada con éxito.")
            st.download_button("🖨️ IMPRIMIR RECIBO PDF/HTML", st.session_state.html_recibo_ready, file_name=f"Recibo_{st.session_state.cli_ready}.html", mime="text/html")
            if st.button("Limpiar"): st.session_state.html_recibo_ready = None; st.rerun()

    # ── COBRANZA FACTURA ──
    with tab_cob_fac:
        if "html_recibo_fac_ready" not in st.session_state:
            st.session_state.html_recibo_fac_ready = None
        if "cli_fac_ready" not in st.session_state:
            st.session_state.cli_fac_ready = None

        if st.session_state.html_recibo_fac_ready:
            st.success(f"✅ Cobranza de factura de '{st.session_state.cli_fac_ready}' registrada con éxito.")
            st.download_button("🖨️ IMPRIMIR RECIBO", st.session_state.html_recibo_fac_ready,
                               file_name=f"ReciboCobFac_{st.session_state.cli_fac_ready}.html", mime="text/html")
            if st.button("🔄 Nueva cobranza", key="limpiar_cob_fac"):
                st.session_state.html_recibo_fac_ready = None
                st.session_state.cli_fac_ready = None
                st.rerun()
        else:
            st.markdown("Registrá el cobro de facturas emitidas. Seleccioná el cliente, marcá las facturas a cancelar e ingresá el medio de pago.")

            FORMAS_COB_FAC = ["EFECTIVO", "TRANSFERENCIA", "CHEQUE DE TERCEROS", "TARJETA DE CREDITO", "DÓLARES", "OTROS"]

            # ── Selectores FUERA del form para reactividad ──
            cf1, cf2 = st.columns(2)
            cli_fac_sel  = cf1.selectbox("Cliente", [""] + sorted(st.session_state.clientes['Razón Social'].tolist()) if not st.session_state.clientes.empty else [""], key="cob_fac_cli")
            forma_fac_sel = cf2.selectbox("Forma de Cobro", FORMAS_COB_FAC, key="cob_fac_forma")

            es_cheque_fac = (forma_fac_sel == "CHEQUE DE TERCEROS")
            es_transf_fac = (forma_fac_sel == "TRANSFERENCIA")

            # ── Facturas impagas del cliente ──
            facturas_sel_indices = []
            total_seleccionado = 0.0

            if cli_fac_sel and "facturas" in st.session_state and not st.session_state.facturas.empty:
                # Calcular saldo de cada factura cruzando con cobros ya registrados
                df_facs_cli = st.session_state.facturas[
                    (st.session_state.facturas['Cliente'] == cli_fac_sel) &
                    (st.session_state.facturas['Tipo'] == 'FACTURA')
                ].copy()

                # Cobros ya registrados para este cliente
                df_cobros_cli = st.session_state.tesoreria[
                    (st.session_state.tesoreria['Cliente/Proveedor'] == cli_fac_sel) &
                    (st.session_state.tesoreria['Tipo'] == 'COBRANZA FACTURA')
                ]
                total_cobrado_cli = df_cobros_cli['Monto'].sum()  # negativos = cobros

                # NC y ND
                df_nc_cli = st.session_state.facturas[
                    (st.session_state.facturas['Cliente'] == cli_fac_sel) &
                    (st.session_state.facturas['Tipo'].isin(['NOTA DE CREDITO', 'NOTA DE DEBITO']))
                ]
                ajuste_nc_nd = df_nc_cli.apply(
                    lambda r: -float(r['Total']) if r['Tipo'] == 'NOTA DE CREDITO' else float(r['Total']), axis=1
                ).sum() if not df_nc_cli.empty else 0.0

                total_facturado_cli = df_facs_cli['Total'].sum() if not df_facs_cli.empty else 0.0
                saldo_pendiente_cli = total_facturado_cli + ajuste_nc_nd + total_cobrado_cli

                if df_facs_cli.empty:
                    st.info(f"No hay facturas registradas para {cli_fac_sel}.")
                else:
                    st.markdown("---")
                    st.markdown(f"##### 📋 Facturas de {cli_fac_sel}")

                    # Métricas rápidas
                    sm1, sm2, sm3 = st.columns(3)
                    sm1.metric("Total Facturado", f"$ {total_facturado_cli:,.2f}")
                    sm2.metric("Total Cobrado",   f"$ {abs(total_cobrado_cli):,.2f}")
                    sm3.metric("Saldo Pendiente", f"$ {saldo_pendiente_cli:,.2f}")
                    st.markdown("---")

                    # Checkboxes de facturas impagas
                    st.markdown("**Seleccioná las facturas a cancelar:**")
                    facturas_checkeadas = {}
                    for _, frow in df_facs_cli.iterrows():
                        fidx = frow.name
                        label = f"FAC {frow.get('Punto Venta','')}-{frow.get('Numero','')} | {frow['Fecha']} | $ {float(frow['Total']):,.2f} | {frow.get('Detalle','')[:50]}"
                        estado_fac = str(frow.get('Estado', 'EMITIDA'))
                        ya_cobrada = (estado_fac == 'COBRADA')
                        checked = st.checkbox(label, key=f"chk_fac_{fidx}", value=False, disabled=ya_cobrada,
                                              help="Ya cobrada" if ya_cobrada else "")
                        facturas_checkeadas[fidx] = (checked, float(frow['Total']))

                    total_seleccionado = sum(v for c, v in facturas_checkeadas.values() if c)
                    facturas_sel_indices = [i for i, (c, v) in facturas_checkeadas.items() if c]

                    if total_seleccionado > 0:
                        st.markdown(
                            f"<div style='background:#eafaf1;border:2px solid #27ae60;border-radius:8px;"
                            f"padding:12px 20px;margin:10px 0;display:flex;justify-content:space-between;'>"
                            f"<b>Total seleccionado:</b>"
                            f"<b style='color:#27ae60;font-size:18px;'>$ {total_seleccionado:,.2f}</b></div>",
                            unsafe_allow_html=True
                        )

            st.markdown("---")

            # ── Cheques: FUERA del form para que el número de filas sea reactivo ──
            cheques_fac_lista = []
            chf_nro = chf_tipo = chf_banco = chf_librador = chf_obs = ""
            chf_fvenc = chf_femision = date.today()
            if es_cheque_fac:
                st.markdown("##### 🏦 Datos del/los Cheque/s Recibidos")
                cant_cheques_fac = st.number_input("¿Cuántos cheques recibiste?", min_value=1, max_value=30, value=1, step=1, key="cant_cheques_fac")
                chfg1, chfg2, chfg3 = st.columns(3)
                chf_librador = chfg1.text_input("Librador * (para todos)", key="chf_librador")
                chf_banco    = chfg2.text_input("Banco Librador * (para todos)", key="chf_banco")
                chf_tipo     = chfg3.selectbox("Tipo (para todos)", ["COMÚN", "DIFERIDO", "ELECTRÓNICO"], key="chf_tipo")
                chf_femision = st.date_input("Fecha Emisión (para todos)", date.today(), key="chf_femision")
                st.markdown("**Detalle por cheque:**")
                for _ci in range(int(cant_cheques_fac)):
                    _c1, _c2, _c3, _c4 = st.columns([2, 2, 2, 2])
                    _nro = _c1.text_input(f"Nro Cheque #{_ci+1}", key=f"chf_nro_{_ci}")
                    _imp = _c2.number_input(f"Importe #{_ci+1} $", min_value=0.0, step=0.01, key=f"chf_imp_{_ci}")
                    _fv  = _c3.date_input(f"Vencimiento #{_ci+1}", date.today() + timedelta(days=30), key=f"chf_fv_{_ci}")
                    _obs = _c4.text_input(f"Obs #{_ci+1}", key=f"chf_obs_{_ci}")
                    if _nro and _imp > 0:
                        cheques_fac_lista.append({"nro": _nro, "importe": _imp, "fvenc": _fv, "obs": _obs})
                chf_nro   = cheques_fac_lista[0]["nro"]   if cheques_fac_lista else ""
                chf_fvenc = cheques_fac_lista[0]["fvenc"] if cheques_fac_lista else date.today()
                st.markdown("---")

            with st.form("f_cob_fac", clear_on_submit=True):
                st.markdown("##### 💰 Datos del Cobro")
                fcf1, fcf2 = st.columns(2)
                if es_admin:
                    cj_fac = fcf1.selectbox("Caja Destino", opc_cajas, key="cob_fac_caja")
                else:
                    cj_fac = caja_propia
                    fcf1.markdown(f"**Caja:** {caja_propia}")

                monto_cobro = fcf2.number_input("Monto cobrado $", min_value=0.0,
                                                value=float(round(total_seleccionado, 2)),
                                                step=100.0, format="%.2f", key="cob_fac_monto")

                # ── Retenciones ──
                st.markdown("##### ✂️ Retenciones (opcional)")
                ret1, ret2, ret3 = st.columns(3)
                ret_iva      = ret1.number_input("Ret. IVA $",      min_value=0.0, step=10.0, format="%.2f", key="ret_iva")
                ret_ganancias = ret2.number_input("Ret. Ganancias $", min_value=0.0, step=10.0, format="%.2f", key="ret_gan")
                ret_suss     = ret3.number_input("Ret. SUSS $",     min_value=0.0, step=10.0, format="%.2f", key="ret_suss")
                total_retenciones = ret_iva + ret_ganancias + ret_suss
                neto_cobro = monto_cobro - total_retenciones

                # ── Transferencia ──
                if es_transf_fac:
                    st.markdown("---")
                    st.markdown("##### 🏦 Datos de la Transferencia")
                    tr1, tr2 = st.columns(2)
                    transf_banco   = tr1.text_input("Banco destino", key="transf_banco")
                    transf_ref     = tr2.text_input("Nro. de referencia / CBU", key="transf_ref")
                else:
                    transf_banco = transf_ref = ""

                nro_recibo = st.text_input("Nro. Recibo / Referencia (opcional)", key="cob_fac_ref")
                obs_cobro  = st.text_area("Observaciones", height=50, key="cob_fac_obs")

                if st.form_submit_button("✅ REGISTRAR COBRANZA"):
                    if not cli_fac_sel:
                        st.warning("Seleccioná un cliente.")
                    elif monto_cobro <= 0:
                        st.warning("Ingresá un monto mayor a cero.")
                    elif es_cheque_fac and (not chf_nro or not chf_banco or not chf_librador):
                        st.warning("Completá Nro. de Cheque, Banco y Librador.")
                    else:
                        # 1) Movimiento principal en tesorería
                        concepto_cob = f"Cobro Factura"
                        if facturas_sel_indices:
                            nros = [f"{st.session_state.facturas.loc[i,'Punto Venta']}-{st.session_state.facturas.loc[i,'Numero']}" for i in facturas_sel_indices if i in st.session_state.facturas.index]
                            concepto_cob = f"Cobro Facturas: {', '.join(nros)}"
                        if chf_nro:
                            concepto_cob += f" | Cheque #{chf_nro}"
                        if transf_ref:
                            concepto_cob += f" | Transf. {transf_banco} Ref:{transf_ref}"

                        forma_final = "CHEQUE TERCERO" if es_cheque_fac else ("TRANSFERENCIA" if es_transf_fac else forma_fac_sel)
                        nt_cob = pd.DataFrame([[
                            str(date.today()), "COBRANZA FACTURA", cj_fac, forma_final,
                            concepto_cob, cli_fac_sel, -monto_cobro, nro_recibo or "-"
                        ]], columns=COL_TESORERIA)
                        st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, nt_cob], ignore_index=True)

                        # 2) Retenciones como egresos separados
                        if ret_iva > 0:
                            ret_row = pd.DataFrame([[str(date.today()), "RETENCION", cj_fac, "RETENCION IVA",
                                f"Ret. IVA — {cli_fac_sel}", cli_fac_sel, -ret_iva, nro_recibo or "-"]], columns=COL_TESORERIA)
                            st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, ret_row], ignore_index=True)
                        if ret_ganancias > 0:
                            ret_row = pd.DataFrame([[str(date.today()), "RETENCION", cj_fac, "RETENCION GANANCIAS",
                                f"Ret. Ganancias — {cli_fac_sel}", cli_fac_sel, -ret_ganancias, nro_recibo or "-"]], columns=COL_TESORERIA)
                            st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, ret_row], ignore_index=True)
                        if ret_suss > 0:
                            ret_row = pd.DataFrame([[str(date.today()), "RETENCION", cj_fac, "RETENCION SUSS",
                                f"Ret. SUSS — {cli_fac_sel}", cli_fac_sel, -ret_suss, nro_recibo or "-"]], columns=COL_TESORERIA)
                            st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, ret_row], ignore_index=True)

                        guardar_datos("tesoreria", st.session_state.tesoreria)

                        # 3) Marcar facturas como COBRADAS
                        for fidx in facturas_sel_indices:
                            if fidx in st.session_state.facturas.index:
                                st.session_state.facturas.loc[fidx, 'Estado'] = 'COBRADA'
                        if facturas_sel_indices:
                            guardar_datos("facturas", st.session_state.facturas)

                        # 4) Registrar cheque/s en cartera si corresponde
                        if es_cheque_fac and chf_banco and chf_librador and cheques_fac_lista:
                            filas_cheq = []
                            for _ch in cheques_fac_lista:
                                filas_cheq.append([
                                    str(date.today()), _ch["nro"], chf_tipo, chf_banco, chf_librador,
                                    _ch["importe"], str(_ch["fvenc"]), "EN CARTERA", "-", "-",
                                    f"Emisión:{chf_femision} | Cob.Factura {cli_fac_sel} | {_ch['obs']}"
                                ])
                            df_cheqs_nuevos = pd.DataFrame(filas_cheq, columns=COL_CHEQ_CARTERA)
                            st.session_state.cheques_cartera = pd.concat([st.session_state.cheques_cartera, df_cheqs_nuevos], ignore_index=True)
                            guardar_datos("cheques_cartera", st.session_state.cheques_cartera)

                        # 5) Generar recibo
                        detalle_ret = ""
                        if total_retenciones > 0:
                            detalle_ret = f" | Ret. IVA: ${ret_iva:,.2f} | Ret. Gan.: ${ret_ganancias:,.2f} | Ret. SUSS: ${ret_suss:,.2f} | Neto: ${neto_cobro:,.2f}"
                        st.session_state.html_recibo_fac_ready = generar_html_recibo({
                            "Fecha": date.today(),
                            "Cliente/Proveedor": cli_fac_sel,
                            "Concepto": concepto_cob + detalle_ret,
                            "Caja/Banco": f"{cj_fac} — {forma_final}",
                            "Monto": monto_cobro,
                            "Ref AFIP": nro_recibo or "-"
                        })
                        st.session_state.cli_fac_ready = cli_fac_sel
                        st.rerun()

    # ── VER MOVIMIENTOS ──
    with tab_ver:
        # Admin puede elegir cualquier caja. Operador solo ve la suya.
        if es_admin:
            cj_v = st.selectbox("Seleccionar Caja", opc_cajas)
        else:
            cj_v = caja_propia
            st.markdown(f"#### 🏦 {caja_propia}")

        df_caja_full = st.session_state.tesoreria[st.session_state.tesoreria['Caja/Banco'].astype(str).str.startswith(cj_v)].copy()

        # ── Mostrar solo movimientos DESDE la última rendición/cierre ──
        # El corte se hace por forma: efectivo/transferencia/tarjeta usan el último cierre general,
        # dólares usan su propio último cierre de DÓLARES para no quedar en cero tras una rendición de efectivo.
        cierres_idx = df_caja_full[
            df_caja_full['Tipo'].isin(['CIERRE DE CAJA', 'RENDICION', 'RENDICIÓN'])
        ].index
        if len(cierres_idx) > 0:
            ultimo_cierre_idx = cierres_idx[-1]
            df_ver = df_caja_full[df_caja_full.index > ultimo_cierre_idx].copy()
            ultimo_cierre_row = df_caja_full.loc[ultimo_cierre_idx]
            st.caption(f"📌 Última rendición: {ultimo_cierre_row['Fecha']} — mostrando movimientos posteriores.")
        else:
            df_ver = df_caja_full.copy()

        # ── Resumen desglosado por Forma ──
        # Para DÓLARES: corte independiente basado en la última rendición de DÓLARES
        FORMAS_RESUMEN = ["EFECTIVO", "TRANSFERENCIA", "TARJETA DE CREDITO", "DÓLARES", "OTROS"]
        ICONOS_FORMA   = {"EFECTIVO": "💵", "TRANSFERENCIA": "🏦", "TARJETA DE CREDITO": "💳", "DÓLARES": "💲", "OTROS": "📋"}

        # Calcular saldo de dólares con su propio corte independiente.
        # Los dólares pueden estar en df_caja_full (Forma=DÓLARES) O en la caja DOLAR separada.
        caja_dolar_vis = f"DOLAR {cj_v}"
        df_dolar_full_vis = st.session_state.tesoreria[
            st.session_state.tesoreria['Caja/Banco'].astype(str).str.startswith(caja_dolar_vis)
        ].copy()
        # Calcular saldo de dólares: siempre usar corte independiente en AMBAS fuentes y sumar
        # Fuente 1: caja DOLAR separada (DOLAR CAJA JUNIN)
        cierres_dolar_sep = df_dolar_full_vis[
            df_dolar_full_vis['Tipo'].isin(['CIERRE DE CAJA', 'RENDICION', 'RENDICIÓN'])
        ].index if not df_dolar_full_vis.empty else []
        if len(cierres_dolar_sep) > 0:
            df_dolar_sep_activo = df_dolar_full_vis[df_dolar_full_vis.index > cierres_dolar_sep[-1]]
        else:
            df_dolar_sep_activo = df_dolar_full_vis
        saldo_dolar_sep = df_dolar_sep_activo['Monto'].sum()

        # Fuente 2: dólares mezclados en caja principal con Forma=DÓLARES, con su propio corte
        mask_rend_dolar_vis = (
            df_caja_full['Tipo'].isin(['CIERRE DE CAJA', 'RENDICION', 'RENDICIÓN']) &
            mask_forma(df_caja_full['Forma'], 'DOLARES')
        )
        cierres_dolar_mix = df_caja_full[mask_rend_dolar_vis].index
        if len(cierres_dolar_mix) > 0:
            df_dolar_mix_activo = df_caja_full[df_caja_full.index > cierres_dolar_mix[-1]]
        else:
            df_dolar_mix_activo = df_caja_full
        saldo_dolar_mix = df_dolar_mix_activo[mask_forma(df_dolar_mix_activo['Forma'], 'DOLARES')]['Monto'].sum()

        saldo_dolares_vis = saldo_dolar_sep + saldo_dolar_mix

        cols_formas = st.columns(len(FORMAS_RESUMEN))

        for idx, forma_r in enumerate(FORMAS_RESUMEN):
            if forma_r == "DÓLARES":
                saldo_forma = saldo_dolares_vis
            else:
                mask = mask_forma(df_ver['Forma'], forma_r.replace("TARJETA DE CREDITO","TARJETA"))
                saldo_forma = df_ver[mask]['Monto'].sum()
            icono = ICONOS_FORMA.get(forma_r, "💰")
            color = "#2ecc71" if saldo_forma >= 0 else "#e74c3c"
            cols_formas[idx].markdown(
                f"<div style='background:#f8f9fa;border-radius:10px;padding:12px;text-align:center;border-left:4px solid {color};'>"
                f"<div style='font-size:22px;'>{icono}</div>"
                f"<div style='font-size:11px;color:#666;font-weight:bold;'>{forma_r}</div>"
                f"<div style='font-size:16px;font-weight:bold;color:{color};'>$ {saldo_forma:,.2f}</div>"
                f"</div>",
                unsafe_allow_html=True
            )

        st.markdown("---")
        st.markdown("##### 📋 Detalle de Movimientos")
        st.dataframe(df_ver, use_container_width=True)

    # ── RENDICIÓN DE CAJA ──
    with tab_cierre:
        if "html_cierre_ready" not in st.session_state:
            st.session_state.html_cierre_ready = None

        if st.session_state.html_cierre_ready:
            st.success("✅ Rendición generada. Descargá el documento para imprimir.")
            st.download_button(
                "🖨️ DESCARGAR RENDICIÓN",
                st.session_state.html_cierre_ready,
                file_name=f"Rendicion_{date.today()}.html",
                mime="text/html"
            )
            if st.button("🔄 Nueva Rendición"):
                st.session_state.html_cierre_ready = None
                st.rerun()
        else:
            st.markdown("Registrá lo que rendís y el sistema calcula automáticamente cuánto queda en caja.")

            hoy = date.today()

            # Solo cajas con efectivo/dólares pueden rendir
            CAJAS_NO_RENDICION = ["BANCO GALICIA", "BANCO PROVINCIA", "BANCO SUPERVIELLE", "TARJETA DE CREDITO"]
            CAJAS_RENDICION = [c for c in TODAS_CAJAS if not any(c.upper().startswith(b) for b in CAJAS_NO_RENDICION)]

            c_cie1, c_cie2 = st.columns([2, 3])
            if es_admin:
                caja_cierre = c_cie1.selectbox("Caja", CAJAS_RENDICION, key="cierre_caja_sel")
            else:
                if caja_propia in CAJAS_RENDICION:
                    caja_cierre = caja_propia
                    c_cie1.markdown(f"**Caja:** {caja_propia}")
                else:
                    st.info("ℹ️ Tu cuenta no requiere rendición de efectivo.")
                    st.stop()
                    caja_cierre = caja_propia
            c_cie2.markdown(f"📅 **Fecha:** {hoy.strftime('%d/%m/%Y')}")

            obs_cierre = st.text_area("Observaciones (opcional)", placeholder="Ej: Se rindieron $400.000 al supervisor.", key="cierre_obs", height=70)

            st.markdown("---")

            # ── Base: TODOS los movimientos de la caja ──
            df_caja_base = st.session_state.tesoreria[
                st.session_state.tesoreria['Caja/Banco'].astype(str).str.startswith(caja_cierre)
            ].copy()

            # ── Base dólares: movimientos de la caja DOLAR correspondiente ──
            caja_dolar_nombre = f"DOLAR {caja_cierre}"
            df_dolar_base = st.session_state.tesoreria[
                st.session_state.tesoreria['Caja/Banco'].astype(str).str.startswith(caja_dolar_nombre)
            ].copy()

            # ── df_cierre = solo movimientos DESDE el último cierre/rendición ──
            # Esto evita que el saldo acumule movimientos de rendiciones anteriores ya procesadas
            cierres_cierre_idx = df_caja_base[
                df_caja_base['Tipo'].isin(['CIERRE DE CAJA', 'RENDICION', 'RENDICIÓN'])
            ].index
            if len(cierres_cierre_idx) > 0:
                ultimo_idx = cierres_cierre_idx[-1]
                df_cierre = df_caja_base[df_caja_base.index > ultimo_idx].copy()
            else:
                df_cierre = df_caja_base.copy()

            # ── Corte de dólares: usa el último cierre/rendición PROPIO de la caja dólar ──
            # (independiente del corte de pesos, para que una rendición de efectivo no borre los dólares)
            if not df_dolar_base.empty:
                cierres_dolar_idx = df_dolar_base[
                    df_dolar_base['Tipo'].isin(['CIERRE DE CAJA', 'RENDICION', 'RENDICIÓN'])
                ].index
                if len(cierres_dolar_idx) > 0:
                    ultimo_idx_dolar = cierres_dolar_idx[-1]
                    df_dolar_cierre = df_dolar_base[df_dolar_base.index > ultimo_idx_dolar].copy()
                else:
                    df_dolar_cierre = df_dolar_base.copy()
            else:
                df_dolar_cierre = df_dolar_base.copy()

            # ── Saldo disponible por forma ──
            mask_efec_base  = mask_forma(df_cierre['Forma'], "EFECTIVO")
            efectivo_disponible = df_cierre[mask_efec_base]['Monto'].sum()

            # Dólares: sumar AMBAS fuentes con corte independiente cada una.
            # Fuente 1: caja DOLAR separada
            saldo_dolar_sep = df_dolar_cierre['Monto'].sum() if not df_dolar_cierre.empty else 0.0
            # Fuente 2: dólares con Forma=DÓLARES en la caja principal, con su propio corte
            mask_rend_dolar = (
                df_caja_base['Tipo'].isin(['CIERRE DE CAJA', 'RENDICION', 'RENDICIÓN']) &
                mask_forma(df_caja_base['Forma'], "DOLARES")
            )
            cierres_dolar_en_base = df_caja_base[mask_rend_dolar].index
            if len(cierres_dolar_en_base) > 0:
                df_dolar_mix = df_caja_base[df_caja_base.index > cierres_dolar_en_base[-1]]
            else:
                df_dolar_mix = df_caja_base
            saldo_dolar_mix = df_dolar_mix[mask_forma(df_dolar_mix['Forma'], "DOLARES")]['Monto'].sum()
            dolares_disponibles = saldo_dolar_sep + saldo_dolar_mix

            # ── Panel: saldo disponible ──
            st.markdown("##### 💰 Disponible en caja")
            col_ef1, col_ef2 = st.columns(2)
            col_ef1_color = "#27ae60" if efectivo_disponible >= 0 else "#e74c3c"
            col_ef2_color = "#27ae60" if dolares_disponibles >= 0 else "#e74c3c"
            col_ef1.markdown(
                f"<div style='background:#f0fff4;border-radius:10px;padding:16px;text-align:center;border:2px solid {col_ef1_color};'>"
                f"<div style='font-size:24px;'>💵</div>"
                f"<div style='font-size:11px;color:#666;font-weight:bold;margin-top:4px;'>TOTAL EN CAJA (PESOS)</div>"
                f"<div style='font-size:28px;font-weight:bold;color:{col_ef1_color};margin-top:6px;'>$ {efectivo_disponible:,.2f}</div>"
                f"</div>", unsafe_allow_html=True
            )
            col_ef2.markdown(
                f"<div style='background:#fffaf0;border-radius:10px;padding:16px;text-align:center;border:2px solid {col_ef2_color};'>"
                f"<div style='font-size:24px;'>💲</div>"
                f"<div style='font-size:11px;color:#666;font-weight:bold;margin-top:4px;'>DÓLARES</div>"
                f"<div style='font-size:28px;font-weight:bold;color:{col_ef2_color};margin-top:6px;'>USD {dolares_disponibles:,.2f}</div>"
                f"</div>", unsafe_allow_html=True
            )

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Ingreso del monto a rendir ──
            st.markdown("##### 📤 ¿Cuánto vas a rendir?")
            c_rend1, c_rend2 = st.columns([2, 3])
            tipo_rendicion = c_rend1.selectbox(
                "Moneda", ["EFECTIVO", "DÓLARES"],
                key="rend_tipo",
                help="Elegí si rendís pesos o dólares"
            )
            simbolo = "$" if tipo_rendicion == "EFECTIVO" else "USD"
            disponible_sel = efectivo_disponible if tipo_rendicion == "EFECTIVO" else dolares_disponibles
            monto_rendicion = c_rend2.number_input(
                f"Monto a rendir ({simbolo})",
                min_value=0.0, step=100.0, value=0.0,
                key="rend_monto", format="%.2f"
            )

            saldo_restante = disponible_sel - monto_rendicion

            # ── Resultado en tiempo real ──
            if monto_rendicion > 0:
                if saldo_restante >= 0:
                    st.markdown(
                        f"<div style='background:#eafaf1;border:2px solid #27ae60;border-radius:10px;padding:16px 20px;margin-top:8px;"
                        f"display:flex;justify-content:space-between;align-items:center;'>"
                        f"<div><div style='font-size:11px;color:#1a7a40;font-weight:bold;'>📤 RENDÍS</div>"
                        f"<div style='font-size:24px;font-weight:bold;color:#c0392b;'>{simbolo} {monto_rendicion:,.2f}</div></div>"
                        f"<div style='font-size:26px;color:#aaa;'>→</div>"
                        f"<div style='text-align:right;'><div style='font-size:11px;color:#1a7a40;font-weight:bold;'>✅ QUEDA EN CAJA</div>"
                        f"<div style='font-size:24px;font-weight:bold;color:#27ae60;'>{simbolo} {saldo_restante:,.2f}</div></div>"
                        f"</div>", unsafe_allow_html=True
                    )
                else:
                    st.markdown(
                        f"<div style='background:#fef9f0;border:2px solid #e67e22;border-radius:10px;padding:16px 20px;margin-top:8px;"
                        f"display:flex;justify-content:space-between;align-items:center;'>"
                        f"<div><div style='font-size:11px;color:#935116;font-weight:bold;'>📤 RENDÍS</div>"
                        f"<div style='font-size:24px;font-weight:bold;color:#c0392b;'>{simbolo} {monto_rendicion:,.2f}</div></div>"
                        f"<div style='font-size:26px;color:#aaa;'>→</div>"
                        f"<div style='text-align:right;'><div style='font-size:11px;color:#935116;font-weight:bold;'>⚠️ EXCEDE EL DISPONIBLE</div>"
                        f"<div style='font-size:24px;font-weight:bold;color:#e74c3c;'>{simbolo} {saldo_restante:,.2f}</div></div>"
                        f"</div>", unsafe_allow_html=True
                    )

            st.markdown("---")

            # ── Vista previa: desglose por forma de pago ──
            FORMAS_PREV = ["EFECTIVO", "TRANSFERENCIA", "TARJETA DE CREDITO", "DÓLARES", "OTROS"]
            ICONOS_PREV = {"EFECTIVO":"💵","TRANSFERENCIA":"🏦","TARJETA DE CREDITO":"💳","DÓLARES":"💲","OTROS":"📋"}
            st.markdown(f"##### 📊 Saldo actual en caja — {caja_cierre}")
            cols_prev = st.columns(len(FORMAS_PREV))
            for idx_p, fr in enumerate(FORMAS_PREV):
                if fr == "DÓLARES":
                    sub_p = dolares_disponibles
                else:
                    mask_p = mask_forma(df_cierre['Forma'], fr.replace("TARJETA DE CREDITO","TARJETA"))
                    sub_p  = df_cierre[mask_p]['Monto'].sum()
                col_p  = "#2ecc71" if sub_p >= 0 else "#e74c3c"
                cols_prev[idx_p].markdown(
                    f"<div style='background:#f8f9fa;border-radius:8px;padding:10px;text-align:center;border-left:3px solid {col_p};'>"
                    f"<div>{ICONOS_PREV.get(fr,'💰')}</div>"
                    f"<div style='font-size:10px;color:#666;font-weight:bold;'>{fr}</div>"
                    f"<div style='font-size:14px;font-weight:bold;color:{col_p};'>$ {sub_p:,.2f}</div>"
                    f"</div>", unsafe_allow_html=True
                )
            st.caption(f"Saldo calculado sobre {len(df_cierre)} movimiento(s) desde la última rendición en {caja_cierre}")
            st.markdown("---")

            if st.button("📋 GENERAR RENDICIÓN", type="primary"):
                responsable = st.session_state.nombre_usuario

                # ── Registrar la rendición como egreso en tesorería ──
                if monto_rendicion > 0:
                    # Si se rinde DÓLARES, el egreso va a la caja DOLAR correspondiente
                    caja_rend = caja_dolar_nombre if tipo_rendicion == "DÓLARES" else caja_cierre
                    nuevos_movs = []
                    # Egreso: lo que se rinde (sale de caja)
                    nuevos_movs.append([
                        date.today(), "RENDICIÓN", caja_rend, tipo_rendicion,
                        f"Rendición — {responsable}", "INTERNO", -monto_rendicion, f"Rendición {hoy}"
                    ])
                    # Si queda remanente, grabarlo como ingreso en el nuevo período
                    if saldo_restante > 0:
                        nuevos_movs.append([
                            date.today(), "SALDO REMANENTE", caja_rend, tipo_rendicion,
                            f"Remanente tras rendición — {responsable}", "INTERNO", saldo_restante, f"Rendición {hoy}"
                        ])
                    df_nuevos = pd.DataFrame(nuevos_movs, columns=COL_TESORERIA)
                    st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, df_nuevos], ignore_index=True)
                    guardar_datos("tesoreria", st.session_state.tesoreria)

                html_cierre = generar_html_cierre_caja({
                    "caja":                caja_cierre,
                    "fecha_cierre":        str(hoy),
                    "responsable":         responsable,
                    "movimientos":         df_cierre.drop(columns=['Fecha_dt'], errors='ignore'),
                    "total":               df_cierre['Monto'].sum(),
                    "efectivo_disponible": efectivo_disponible,
                    "dolares_disponibles": dolares_disponibles,
                    "monto_rendicion":     monto_rendicion,
                    "tipo_rendicion":      tipo_rendicion,
                    "saldo_restante":      saldo_restante,
                    "observaciones":       obs_cierre.strip()
                })
                st.session_state.html_cierre_ready = html_cierre
                st.rerun()

    # ── PASE DE EFECTIVO: disponible para todos (operador pasa desde su caja, admin elige) ──
    with tab_pase:
        if st.session_state.get("msg_pase"):
            st.success(st.session_state.msg_pase)
            st.session_state.msg_pase = None
        st.markdown("Registrá el pase de efectivo físico de una caja a otra. Se genera un egreso en la caja origen y un ingreso en la caja destino.")
        with st.form("f_pase", clear_on_submit=True):
            c1, c2 = st.columns(2)
            if es_admin:
                origen_pase  = c1.selectbox("Caja Origen (sale el efectivo)", TODAS_CAJAS, key="pase_orig")
                destino_pase = c2.selectbox("Caja Destino (entra el efectivo)", TODAS_CAJAS, key="pase_dest")
            else:
                st.markdown(f"**Caja Origen:** {caja_propia}")
                origen_pase  = caja_propia
                # El operador puede mandar efectivo a cualquier otra caja
                otras_cajas  = [c for c in TODAS_CAJAS if c != caja_propia]
                destino_pase = c2.selectbox("Caja Destino (entra el efectivo)", otras_cajas, key="pase_dest")
            forma_pase = st.selectbox("Tipo de valor", ["EFECTIVO", "DÓLARES"], key="pase_forma")
            monto_pase = st.number_input("Monto a pasar $", min_value=0.0, key="pase_monto")
            concepto_pase = st.text_input("Concepto (opcional)", key="pase_concepto")
            if st.form_submit_button("💱 REGISTRAR PASE"):
                if monto_pase > 0 and origen_pase != destino_pase:
                    desc = concepto_pase if concepto_pase else f"Pase a {destino_pase}"
                    desc_dest = concepto_pase if concepto_pase else f"Pase desde {origen_pase}"
                    p1 = pd.DataFrame([[date.today(), "PASE EFECTIVO", origen_pase, forma_pase, desc,       "INTERNO", -monto_pase, "-"]], columns=COL_TESORERIA)
                    p2 = pd.DataFrame([[date.today(), "PASE EFECTIVO", destino_pase, forma_pase, desc_dest, "INTERNO",  monto_pase, "-"]], columns=COL_TESORERIA)
                    st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, p1, p2], ignore_index=True)
                    guardar_datos("tesoreria", st.session_state.tesoreria)
                    st.session_state.msg_pase = f"✅ Pase de {forma_pase} por $ {monto_pase:,.2f} de {origen_pase} → {destino_pase} registrado."
                    st.rerun()
                elif origen_pase == destino_pase:
                    st.warning("La caja origen y destino no pueden ser la misma.")
                else:
                    st.warning("Ingresá un monto mayor a cero.")

    # Traspasos y Orden de Pago: SOLO ADMIN
    if tab_tras is not None:
        with tab_tras:
            if st.session_state.get("msg_traspaso"):
                st.success(st.session_state.msg_traspaso)
                st.session_state.msg_traspaso = None
            with st.form("f_tras", clear_on_submit=True):
                o = st.selectbox("Desde", opc_cajas)
                d = st.selectbox("Hacia", opc_cajas)
                m = st.number_input("Monto a Traspasar", min_value=0.0)
                if st.form_submit_button("EJECUTAR"):
                    if m > 0:
                        tr1 = pd.DataFrame([[date.today(), "TRASPASO", o, "INTERNO", f"Hacia {d}", "INTERNO", -m, "-"]], columns=COL_TESORERIA)
                        tr2 = pd.DataFrame([[date.today(), "TRASPASO", d, "INTERNO", f"Desde {o}", "INTERNO",  m, "-"]], columns=COL_TESORERIA)
                        st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, tr1, tr2], ignore_index=True)
                        guardar_datos("tesoreria", st.session_state.tesoreria)
                        st.session_state.msg_traspaso = f"✅ Traspaso de $ {m:,.2f} de {o} hacia {d} ejecutado."
                        st.rerun()
                    else:
                        st.warning("Ingresá un monto mayor a cero.")

    if tab_op is not None:
        with tab_op:
            st.subheader("💸 Generar Orden de Pago a Proveedor")
            if "html_op_ready" not in st.session_state: st.session_state.html_op_ready = None

            if st.session_state.html_op_ready:
                st.success(f"✅ Orden de Pago para '{st.session_state.prov_ready}' registrada con éxito.")
                st.download_button("🖨️ IMPRIMIR ORDEN DE PAGO", st.session_state.html_op_ready, file_name=f"OrdenPago_{st.session_state.prov_ready}.html", mime="text/html")
                if st.button("Limpiar OP"): st.session_state.html_op_ready = None; st.rerun()
            else:
                # ── Proveedor y forma de pago FUERA del form para reaccionar en tiempo real ──
                p_sel   = st.selectbox("Seleccionar Proveedor", st.session_state.proveedores['Razón Social'].unique() if not st.session_state.proveedores.empty else [""], key="op_prov")
                forma_op = st.selectbox("Forma de Pago", ["TRANSFERENCIA", "EFECTIVO", "CHEQUE PROPIO", "CHEQUE DE TERCERO"], key="op_forma")
                afip_p  = st.text_input("Referencia AFIP / Concepto", key="op_afip")

                # ── FACTURAS/COMPRAS PENDIENTES DEL PROVEEDOR ──
                # Lógica FIFO: distribuir pagos negativos contra facturas positivas
                # ordenadas por fecha, mostrar solo las que tienen saldo > 0.
                monto_sugerido_op = 0.0
                facturas_op_desc = []
                if p_sel and not st.session_state.compras.empty:
                    df_comp_prov = st.session_state.compras[
                        st.session_state.compras['Proveedor'] == p_sel
                    ].copy()
                    df_facturas_prov = df_comp_prov[df_comp_prov['Total'] > 0].copy()
                    pagos_op_total   = df_comp_prov[df_comp_prov['Total'] < 0]['Total'].sum()
                    total_facturado_op = df_facturas_prov['Total'].sum() if not df_facturas_prov.empty else 0.0
                    saldo_prov = round(total_facturado_op + pagos_op_total, 2)

                    # Calcular saldo pendiente por factura (FIFO)
                    saldo_restante_pago_op = abs(pagos_op_total)
                    facturas_con_saldo = []
                    for idx, crow in df_facturas_prov.sort_values('Fecha').iterrows():
                        total_c = float(crow['Total'])
                        if saldo_restante_pago_op >= total_c:
                            saldo_restante_pago_op -= total_c
                            # Factura totalmente pagada → no mostrar
                        else:
                            saldo_pend_c = total_c - saldo_restante_pago_op
                            saldo_restante_pago_op = 0.0
                            facturas_con_saldo.append((idx, crow, round(saldo_pend_c, 2)))

                    if facturas_con_saldo:
                        st.markdown("---")
                        st.markdown(f"##### 📋 Comprobantes pendientes de {p_sel}")
                        op1m, op2m, op3m = st.columns(3)
                        op1m.metric("Total Facturado", f"$ {total_facturado_op:,.2f}")
                        op2m.metric("Ya Pagado",       f"$ {abs(pagos_op_total):,.2f}")
                        op3m.metric("Saldo a Pagar",   f"$ {saldo_prov:,.2f}")
                        st.markdown("---")
                        st.markdown("**Seleccioná las facturas a pagar:**")
                        comp_chequeados = {}
                        for cidx, crow, saldo_c in facturas_con_saldo:
                            fecha_c = crow.get('Fecha', '-')
                            tipo_c  = crow.get('Tipo Factura', '-')
                            pv_c    = crow.get('Punto Venta', '-')
                            total_c = float(crow['Total'])
                            if saldo_c < total_c:
                                label_c = f"📄 {fecha_c} | {tipo_c} {pv_c} | $ {saldo_c:,.2f} (saldo de $ {total_c:,.2f})"
                            else:
                                label_c = f"📄 {fecha_c} | {tipo_c} {pv_c} | $ {saldo_c:,.2f}"
                            chk_c = st.checkbox(label_c, key=f"chk_comp_{cidx}", value=False)
                            comp_chequeados[cidx] = (chk_c, saldo_c, label_c)

                        monto_sugerido_op = sum(v for c, v, l in comp_chequeados.values() if c)
                        facturas_op_desc  = [l for c, v, l in comp_chequeados.values() if c]

                        if monto_sugerido_op > 0:
                            st.markdown(
                                f"<div style='background:#fef9f0;border:2px solid #e67e22;border-radius:8px;"
                                f"padding:12px 20px;margin:10px 0;display:flex;justify-content:space-between;'>"
                                f"<b>Total seleccionado a pagar:</b>"
                                f"<b style='color:#e67e22;font-size:18px;'>$ {monto_sugerido_op:,.2f}</b></div>",
                                unsafe_allow_html=True
                            )
                        st.markdown("---")
                    elif not df_facturas_prov.empty:
                        st.success(f"✅ {p_sel} no tiene comprobantes pendientes de pago.")
                    else:
                        st.info(f"No hay comprobantes de compra registrados para {p_sel}.")

                # ── TRANSFERENCIA / EFECTIVO ─────────────────────────────────────────
                if forma_op in ["TRANSFERENCIA", "EFECTIVO"]:
                    with st.form("f_op_std", clear_on_submit=True):
                        cj_p  = st.selectbox("Caja de Salida", opc_cajas)
                        mon_p = st.number_input("Monto $", min_value=0.0, step=0.01, value=float(round(monto_sugerido_op, 2)))
                        st.caption(f"Proveedor: **{p_sel}** | Forma: **{forma_op}**")
                        if facturas_op_desc:
                            st.markdown(f"*Comprobantes seleccionados: {len(facturas_op_desc)}*")
                        if st.form_submit_button("✅ GENERAR ORDEN DE PAGO"):
                            _prov  = st.session_state.get("op_prov", p_sel)
                            _forma = st.session_state.get("op_forma", forma_op)
                            _afip  = st.session_state.get("op_afip", afip_p)
                            if _prov and mon_p > 0:
                                nt = pd.DataFrame([[date.today(), "PAGO PROV", cj_p, _forma, "Orden de Pago", _prov, -mon_p, _afip]], columns=COL_TESORERIA)
                                nc = pd.DataFrame([[date.today(), _prov, "-", "ORDEN PAGO", 0, 0, 0, 0, 0, 0, -mon_p]], columns=COL_COMPRAS)
                                st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, nt], ignore_index=True)
                                st.session_state.compras   = pd.concat([st.session_state.compras, nc], ignore_index=True)
                                ok = guardar_tesoreria_y_compras()
                                if ok:
                                    st.session_state.html_op_ready = generar_html_orden_pago({"Fecha": date.today(), "Proveedor": _prov, "Concepto": f"Pago {_forma}", "Caja/Banco": cj_p, "Monto": mon_p, "Ref AFIP": _afip})
                                    st.session_state.prov_ready = _prov
                                    st.rerun()
                            else:
                                st.warning("Seleccioná proveedor y completá el monto.")

                # ── CHEQUE PROPIO ────────────────────────────────────────────────────
                elif forma_op == "CHEQUE PROPIO":
                    st.markdown("##### 📝 Datos del Cheque Propio a Emitir")
                    with st.form("f_op_cheq_propio", clear_on_submit=True):
                        op1, op2, op3 = st.columns(3)
                        nro_op   = op1.text_input("Nro de Cheque")
                        tipo_op  = op2.selectbox("Tipo", ["FÍSICO", "ECHEQ"])
                        banco_op = op3.text_input("Banco Emisor")
                        op4, op5 = st.columns(2)
                        mon_op   = op4.number_input("Importe $", min_value=0.0, step=0.01, value=float(round(monto_sugerido_op, 2)))
                        f_venc_op = op5.date_input("Fecha de Vencimiento del Cheque", value=date.today() + timedelta(days=30))
                        obs_op   = st.text_input("Observaciones")
                        if st.form_submit_button("✅ EMITIR CHEQUE Y GENERAR ORDEN DE PAGO"):
                            if p_sel and nro_op and mon_op > 0:
                                # Registrar en cheques_emitidos
                                nuevo_cheq = pd.DataFrame([[
                                    str(date.today()), nro_op, tipo_op, banco_op, p_sel,
                                    mon_op, str(f_venc_op), "PENDIENTE", "-", obs_op
                                ]], columns=COL_CHEQ_EMITIDOS)
                                st.session_state.cheques_emitidos = pd.concat([st.session_state.cheques_emitidos, nuevo_cheq], ignore_index=True)
                                # Registrar egreso en tesorería y cuenta corriente
                                nt = pd.DataFrame([[date.today(), "PAGO PROV", f"CHEQUE {tipo_op}", f"CHEQUE PROPIO #{nro_op}", "Orden de Pago", p_sel, -mon_op, afip_p]], columns=COL_TESORERIA)
                                nc = pd.DataFrame([[date.today(), p_sel, "-", "ORDEN PAGO CHEQUE", 0, 0, 0, 0, 0, 0, -mon_op]], columns=COL_COMPRAS)
                                st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, nt], ignore_index=True)
                                st.session_state.compras   = pd.concat([st.session_state.compras, nc], ignore_index=True)
                                ok = guardar_tesoreria_y_compras()
                                if ok:
                                    guardar_datos("cheques_emitidos", st.session_state.cheques_emitidos)
                                    st.session_state.html_op_ready = generar_html_orden_pago({"Fecha": date.today(), "Proveedor": p_sel, "Concepto": f"Cheque {tipo_op} #{nro_op} — Vto: {f_venc_op}", "Caja/Banco": f"Cheque {banco_op}", "Monto": mon_op, "Ref AFIP": afip_p})
                                    st.session_state.prov_ready = p_sel
                                    st.rerun()
                            else:
                                st.warning("Completá proveedor, número de cheque e importe.")

                # ── CHEQUE DE TERCERO (de cartera) ───────────────────────────────────
                elif forma_op == "CHEQUE DE TERCERO":
                    # Filtrar cheques en cartera disponibles
                    df_cartera_disp = st.session_state.cheques_cartera[
                        st.session_state.cheques_cartera['Estado'] == 'EN CARTERA'
                    ].copy() if not st.session_state.cheques_cartera.empty else pd.DataFrame(columns=COL_CHEQ_CARTERA)

                    if df_cartera_disp.empty:
                        st.warning("⚠️ No hay cheques de terceros en cartera disponibles. Ingresá uno desde el módulo **CHEQUES → Cheques en Cartera**.")
                    else:
                        st.markdown("##### 📂 Seleccionar Cheque de Cartera")
                        # Mostrar cartera disponible con info útil
                        opciones_cartera = []
                        for idx, r in df_cartera_disp.iterrows():
                            d_v = (pd.to_datetime(r['Fecha Vencimiento']).date() - date.today()).days if r['Fecha Vencimiento'] not in ['-',''] else '?'
                            opciones_cartera.append(f"#{r['Nro Cheque']} — {r['Banco Librador']} — {r['Librador']} — $ {float(r['Importe']):,.2f} — Vto: {r['Fecha Vencimiento']} ({d_v}d)")

                        idx_map = {opc: idx for opc, idx in zip(opciones_cartera, df_cartera_disp.index)}

                        with st.form("f_op_cheq_tercero", clear_on_submit=True):
                            cheq_sel_str = st.selectbox("Cheque a utilizar", opciones_cartera)
                            if st.form_submit_button("✅ APLICAR CHEQUE Y GENERAR ORDEN DE PAGO"):
                                if p_sel and cheq_sel_str:
                                    cheq_idx = idx_map[cheq_sel_str]
                                    cheq_row = st.session_state.cheques_cartera.loc[cheq_idx]
                                    mon_ct   = float(cheq_row['Importe'])
                                    # Marcar cheque como APLICADO PAGO
                                    st.session_state.cheques_cartera.loc[cheq_idx, 'Estado']          = 'APLICADO PAGO'
                                    st.session_state.cheques_cartera.loc[cheq_idx, 'Destino']         = p_sel
                                    st.session_state.cheques_cartera.loc[cheq_idx, 'Fecha Aplicación']= str(date.today())
                                    guardar_datos("cheques_cartera", st.session_state.cheques_cartera)
                                    # Registrar en tesorería
                                    nt = pd.DataFrame([[date.today(), "PAGO PROV", "CARTERA", f"CHEQUE TERCERO #{cheq_row['Nro Cheque']}", "Orden de Pago", p_sel, -mon_ct, afip_p]], columns=COL_TESORERIA)
                                    nc = pd.DataFrame([[date.today(), p_sel, "-", "ORDEN PAGO CHEQUE TERCERO", 0, 0, 0, 0, 0, 0, -mon_ct]], columns=COL_COMPRAS)
                                    st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, nt], ignore_index=True)
                                    st.session_state.compras   = pd.concat([st.session_state.compras, nc], ignore_index=True)
                                    ok = guardar_tesoreria_y_compras()
                                    if ok:
                                        st.session_state.html_op_ready = generar_html_orden_pago({"Fecha": date.today(), "Proveedor": p_sel, "Concepto": f"Cheque de {cheq_row['Librador']} #{cheq_row['Nro Cheque']} — Vto: {cheq_row['Fecha Vencimiento']}", "Caja/Banco": f"Cheque {cheq_row['Banco Librador']}", "Monto": mon_ct, "Ref AFIP": afip_p})
                                        st.session_state.prov_ready = p_sel
                                        st.rerun()
                                else:
                                    st.warning("Seleccioná proveedor y cheque.")

elif sel == "CTA CTE INDIVIDUAL":
    st.header("📑 Cuenta Corriente por Cliente")
    if not st.session_state.clientes.empty:
        cl = st.selectbox("Seleccionar Cliente", sorted(st.session_state.clientes['Razón Social'].unique()))

        # ── SALDO INICIAL (MIGRACIÓN) ──
        with st.expander("⚙️ Cargar Saldo Inicial (migración desde otro sistema)", expanded=False):
            st.markdown(
                "<div style='background:#fff8e1;border-left:4px solid #f39c12;border-radius:6px;"
                "padding:10px 16px;font-size:13px;margin-bottom:12px;'>"
                "⚠️ <b>Usá esta opción solo para cargar el saldo que traía el cliente de un sistema anterior.</b><br>"
                "Se registrará como <i>SALDO INICIAL</i> con la fecha que indiques. No afecta facturas ni viajes."
                "</div>", unsafe_allow_html=True
            )
            with st.form("f_saldo_ini_cli", clear_on_submit=True):
                si1, si2, si3 = st.columns(3)
                fecha_si   = si1.date_input("Fecha del saldo", value=date.today(), key="si_cli_fecha")
                monto_si   = si2.number_input("Saldo a favor de la empresa (deuda del cliente) $",
                                              min_value=0.0, step=100.0, format="%.2f", key="si_cli_monto")
                favor_cli  = si3.checkbox("¿Saldo a favor del CLIENTE?", value=False, key="si_cli_favor")
                obs_si     = st.text_input("Observación (opcional)", placeholder="Ej: Saldo migrado de sistema anterior", key="si_cli_obs")
                if st.form_submit_button("✅ CARGAR SALDO INICIAL"):
                    if monto_si > 0:
                        importe_final = -monto_si if favor_cli else monto_si
                        concepto_si   = obs_si if obs_si else "SALDO INICIAL — migración"
                        nv_si = pd.DataFrame([[
                            str(fecha_si), cl, str(fecha_si), "AJUSTE", "SALDO INICIAL",
                            "-", importe_final, "AJUSTE", concepto_si
                        ]], columns=COL_VIAJES)
                        st.session_state.viajes = pd.concat([st.session_state.viajes, nv_si], ignore_index=True)
                        guardar_datos("viajes", st.session_state.viajes)
                        signo_txt = "a favor del cliente" if favor_cli else "a cobrar"
                        st.success(f"✅ Saldo inicial de $ {monto_si:,.2f} ({signo_txt}) cargado para {cl}.")
                        st.rerun()
                    else:
                        st.warning("Ingresá un monto mayor a cero.")

        st.markdown("---")

        # ══════════════════════════════════════════════════════════════
        # CUENTA CORRIENTE UNIFICADA: viajes + facturas + cobros
        # ══════════════════════════════════════════════════════════════
        filas_unif = []

        # 1) VIAJES (débitos por servicios prestados y pagos de viajes)
        df_v_cli = st.session_state.viajes[
            st.session_state.viajes['Cliente'] == cl
        ].copy()
        for _, r in df_v_cli.iterrows():
            imp = float(r['Importe'])
            origen = str(r.get('Origen', '-'))
            destino = str(r.get('Destino', '-'))
            comp_nro = str(r.get('Nro Comp Asoc', '-'))

            if origen in ("AJUSTE", "SALDO INICIAL"):
                tipo_mov = "SALDO INICIAL"
                concepto = str(r.get('Nro Comp Asoc', 'Saldo inicial migrado'))
            elif imp < 0:
                tipo_mov = "COBRANZA VIAJE"
                concepto = f"Cobro viaje | Ref: {comp_nro}"
            else:
                tipo_mov = "VIAJE"
                concepto = f"{origen} → {destino}" + (f" | {comp_nro}" if comp_nro not in ['-','','nan'] else "")

            filas_unif.append({
                "Fecha":       str(r.get('Fecha Viaje', r.get('Fecha Carga', '-'))),
                "Tipo":        tipo_mov,
                "Comprobante": concepto,
                "Debe":        imp if imp > 0 else 0.0,
                "Haber":       abs(imp) if imp < 0 else 0.0,
            })

        # 2) FACTURAS emitidas (desde tesorería: FACTURA, NC, ND)
        TIPOS_FAC = ['FACTURA', 'NOTA DE CREDITO', 'NOTA DE DEBITO']
        df_t_cli = st.session_state.tesoreria[
            (st.session_state.tesoreria['Cliente/Proveedor'] == cl) &
            (st.session_state.tesoreria['Tipo'].isin(TIPOS_FAC))
        ].copy()
        for _, r in df_t_cli.iterrows():
            monto = float(r['Monto'])
            filas_unif.append({
                "Fecha":       str(r['Fecha']),
                "Tipo":        r['Tipo'],
                "Comprobante": str(r['Concepto']),
                "Debe":        monto if monto > 0 else 0.0,
                "Haber":       abs(monto) if monto < 0 else 0.0,
            })

        # 3) COBRANZAS de facturas (desde tesorería: COBRO, COBRANZA, COBRANZA FACTURA)
        TIPOS_COB = ['COBRO', 'COBRANZA', 'COBRANZA FACTURA']
        df_cob_cli = st.session_state.tesoreria[
            (st.session_state.tesoreria['Cliente/Proveedor'] == cl) &
            (st.session_state.tesoreria['Tipo'].isin(TIPOS_COB))
        ].copy()
        for _, r in df_cob_cli.iterrows():
            monto = float(r['Monto'])
            filas_unif.append({
                "Fecha":       str(r['Fecha']),
                "Tipo":        "COBRO",
                "Comprobante": str(r['Concepto']),
                "Debe":        monto if monto > 0 else 0.0,
                "Haber":       abs(monto) if monto < 0 else 0.0,
            })

        if not filas_unif:
            st.info(f"No hay movimientos registrados para {cl}.")
        else:
            # Ordenar por fecha y calcular saldo acumulado
            df_unif = pd.DataFrame(filas_unif)
            df_unif['Fecha_dt'] = pd.to_datetime(df_unif['Fecha'], errors='coerce')
            df_unif = df_unif.sort_values('Fecha_dt', na_position='last').reset_index(drop=True)

            saldo_acum = 0.0
            filas_cc = []
            for _, r in df_unif.iterrows():
                saldo_acum += r['Debe'] - r['Haber']
                filas_cc.append({
                    "Fecha":       r['Fecha'],
                    "Tipo":        r['Tipo'],
                    "Comprobante": r['Comprobante'],
                    "Debe":        f"$ {r['Debe']:,.2f}" if r['Debe'] > 0 else "",
                    "Haber":       f"$ {r['Haber']:,.2f}" if r['Haber'] > 0 else "",
                    "Saldo":       f"$ {saldo_acum:,.2f}",
                })

            df_cc_show = pd.DataFrame(filas_cc)
            total_debe  = df_unif['Debe'].sum()
            total_haber = df_unif['Haber'].sum()
            saldo_final = total_debe - total_haber

            # Métricas
            m1, m2, m3 = st.columns(3)
            m1.metric("Total Facturado / Viajes", f"$ {total_debe:,.2f}")
            m2.metric("Total Cobrado",             f"$ {total_haber:,.2f}")
            color_s = "inverse" if saldo_final > 0 else "normal"
            m3.metric("Saldo Pendiente",            f"$ {saldo_final:,.2f}", delta_color=color_s)
            st.markdown("---")

            # Tabla con colores por tipo
            st.dataframe(df_cc_show, use_container_width=True, hide_index=True)

            # Descarga
            html_reporte = generar_html_resumen(cl, df_cc_show, saldo_final)
            st.download_button(
                label="📄 DESCARGAR RESUMEN",
                data=html_reporte,
                file_name=f"CuentaCorriente_{cl.replace(' ','_')}.html",
                mime="text/html"
            )
    else:
        st.info("No hay clientes registrados.")

elif sel == "CTA CTE GENERAL":
    st.header("🌎 Estado Global de Deudores")

    # Unificar saldos: viajes + facturas/cobros de tesorería
    saldos_dict = {}

    # 1) Viajes
    if not st.session_state.viajes.empty:
        for cli, grp in st.session_state.viajes.groupby('Cliente'):
            saldos_dict[cli] = saldos_dict.get(cli, 0.0) + float(grp['Importe'].sum())

    # 2) Tesorería (facturas, NC, ND, cobros de factura)
    TIPOS_CC_GRAL = ['FACTURA', 'NOTA DE CREDITO', 'NOTA DE DEBITO', 'COBRO', 'COBRANZA', 'COBRANZA FACTURA']
    if not st.session_state.tesoreria.empty:
        df_t_gral = st.session_state.tesoreria[
            st.session_state.tesoreria['Tipo'].isin(TIPOS_CC_GRAL)
        ]
        for cli, grp in df_t_gral.groupby('Cliente/Proveedor'):
            saldos_dict[cli] = saldos_dict.get(cli, 0.0) + float(grp['Monto'].sum())

    if saldos_dict:
        res = pd.DataFrame(list(saldos_dict.items()), columns=['Cliente', 'Saldo'])
        # Redondear y excluir saldos cero o residuales
        res['Saldo'] = res['Saldo'].apply(lambda x: round(float(x), 2))
        res = res[res['Saldo'].apply(lambda x: abs(x) > 0.01)].sort_values('Saldo', ascending=False).reset_index(drop=True)

        if res.empty:
            st.success("✅ Todos los clientes tienen saldo en cero.")
        else:
            m1, m2, m3 = st.columns(3)
            m1.metric("Clientes con saldo", len(res))
            m2.metric("Total a cobrar",     f"$ {res[res['Saldo']>0]['Saldo'].sum():,.2f}")
            m3.metric("Total a favor clientes", f"$ {res[res['Saldo']<0]['Saldo'].sum():,.2f}")
            # Renombrar para compatibilidad con generar_html_cta_cte_general
            res_show = res.rename(columns={'Saldo': 'Importe'})
            st.table(res_show.style.format({"Importe": "$ {:,.2f}"}))
            html_cta_cte = generar_html_cta_cte_general("Clientes", res_show, date.today())
            st.download_button(
                label="🖨️ DESCARGAR PDF / IMPRIMIR",
                data=html_cta_cte,
                file_name=f"CTA_CTE_General_Clientes_{date.today()}.html",
                mime="text/html",
                help="Abrí el archivo descargado en el navegador y usá Ctrl+P para imprimir o guardar como PDF"
            )
    else:
        st.info("No hay movimientos registrados.")

elif sel == "CARGA PROVEEDOR":
    st.header("👤 Gestión de Proveedores")
    if st.session_state.get("msg_proveedor"):
        st.success(st.session_state.msg_proveedor)
        st.session_state.msg_proveedor = None
    with st.expander("➕ ALTA DE NUEVO PROVEEDOR", expanded=False):
        with st.form("f_prov", clear_on_submit=True):
            c1, c2  = st.columns(2)
            rs      = c1.text_input("Razón Social")
            doc     = c2.text_input("CUIT o DNI")
            cuenta  = c1.selectbox("Cuenta de Gastos", sorted(st.session_state.cuentas_gastos))
            cat_iva = c2.selectbox("Categoría IVA", ["Responsable Inscripto", "Exento en IVA", "Consumidor Final", "Monotributista", "No Inscripto"])
            c3, c4  = st.columns(2)
            cbu     = c3.text_input("CBU")
            alias   = c4.text_input("Alias")
            if st.form_submit_button("REGISTRAR PROVEEDOR"):
                if rs and doc:
                    np_row = pd.DataFrame([[rs, doc, cuenta, cat_iva, cbu, alias]], columns=COL_PROVEEDORES)
                    st.session_state.proveedores = pd.concat([st.session_state.proveedores, np_row], ignore_index=True)
                    guardar_datos("proveedores", st.session_state.proveedores)
                    st.session_state.msg_proveedor = f"✅ Proveedor '{rs}' registrado correctamente."
                    st.rerun()
                else:
                    st.warning("Completá Razón Social y CUIT/DNI para continuar.")
    st.subheader("📋 Base de Proveedores")
    if not st.session_state.proveedores.empty:
        for i, row in st.session_state.proveedores.iterrows():
            with st.container():
                c_inf, c_ed, c_el = st.columns([0.7, 0.15, 0.15])
                c_inf.markdown(f"**{row['Razón Social']}** | CUIT: {row['CUIT/DNI']}")
                c_inf.caption(f"📂 Cuenta: {row['Cuenta de Gastos']} | {row['Categoría IVA']} | 🏦 CBU: {row['CBU']} | Alias: {row['Alias']}")
                if c_ed.button("📝 Editar", key=f"edit_p_{i}"): st.session_state[f"edit_p_mode_{i}"] = True
                if c_el.button("🗑️", key=f"del_p_{i}"):
                    tiene_compras = not st.session_state.compras[st.session_state.compras['Proveedor'] == row['Razón Social']].empty
                    if tiene_compras: st.error("No se puede eliminar: tiene comprobantes asociados.")
                    else:
                        st.session_state.proveedores = st.session_state.proveedores.drop(i).reset_index(drop=True)
                        guardar_datos("proveedores", st.session_state.proveedores)
                        st.rerun()
                if st.session_state.get(f"edit_p_mode_{i}", False):
                    with st.form(f"f_edit_p_{i}"):
                        ce1, ce2 = st.columns(2)
                        n_rs    = ce1.text_input("Razón Social", value=row['Razón Social'])
                        n_doc   = ce2.text_input("CUIT/DNI", value=row['CUIT/DNI'])
                        ce3, ce4 = st.columns(2)
                        n_cbu   = ce3.text_input("CBU", value=row['CBU'])
                        n_alias = ce4.text_input("Alias", value=row['Alias'])
                        if st.form_submit_button("✅ Guardar"):
                            st.session_state.proveedores.loc[i] = [n_rs, n_doc, row['Cuenta de Gastos'], row['Categoría IVA'], n_cbu, n_alias]
                            guardar_datos("proveedores", st.session_state.proveedores)
                            st.session_state[f"edit_p_mode_{i}"] = False; st.rerun()
            st.divider()

elif sel == "CARGA GASTOS":
    st.header("💸 Carga de Gastos")
    if st.session_state.get("msg_gasto"):
        st.success(st.session_state.msg_gasto)
        st.session_state.msg_gasto = None

    # ── Inputs fuera del form para que el total se actualice en tiempo real ──
    prov_sel = st.selectbox("Proveedor", st.session_state.proveedores['Razón Social'].unique() if not st.session_state.proveedores.empty else [""])
    c1, c2, c3_fecha = st.columns(3)
    fecha_comp = c1.date_input("Fecha del Comprobante", value=date.today())
    pv         = c2.text_input("Punto de Venta")
    tipo_f     = c3_fecha.selectbox("Tipo de Factura", ["A", "B", "C", "B - OP EXENTA", "C - OP EXENTA", "REMITO", "NOTA DE CREDITO", "NOTA DE DEBITO"])
    c3, c4   = st.columns(2)
    n21      = c3.number_input("Importe Neto (21%)", min_value=0.0, step=0.01, key="g_n21")
    n10      = c4.number_input("Importe Neto (10.5%)", min_value=0.0, step=0.01, key="g_n10")
    c5, c6, c7 = st.columns(3)
    r_iva    = c5.number_input("Retención IVA", min_value=0.0, step=0.01, key="g_riva")
    r_gan    = c6.number_input("Retención Ganancia", min_value=0.0, step=0.01, key="g_rgan")
    r_iibb   = c7.number_input("Retención IIBB", min_value=0.0, step=0.01, key="g_riibb")
    nograv   = st.number_input("Conceptos No Gravados", min_value=0.0, step=0.01, key="g_nograv")

    # ── Total en tiempo real ──
    # Para OP Exentas (B o C): no se suma IVA, el neto es el total
    es_exenta = tipo_f in ("B - OP EXENTA", "C - OP EXENTA")
    if es_exenta:
        total = n21 + n10 + r_iva + r_gan + r_iibb + nograv
        st.info("ℹ️ **Operación Exenta**: no se aplica IVA sobre los importes netos.")
    else:
        total = (n21 * 1.21) + (n10 * 1.105) + r_iva + r_gan + r_iibb + nograv
    if tipo_f == "NOTA DE CREDITO": total = -total

    color_total = "#2ecc71" if total >= 0 else "#e74c3c"
    signo = "-" if total < 0 else ""
    st.markdown(
        f"<div style='background:#f0f2f6;border-radius:10px;padding:16px 24px;margin:12px 0;"
        f"border-left:5px solid {color_total};display:flex;align-items:center;gap:16px;'>"
        f"<span style='font-size:14px;color:#555;font-weight:bold;'>TOTAL DEL COMPROBANTE</span>"
        f"<span style='font-size:28px;font-weight:bold;color:{color_total};'>{signo}$ {abs(total):,.2f}</span>"
        f"</div>",
        unsafe_allow_html=True
    )

    if st.button("✅ REGISTRAR COMPROBANTE", type="primary"):
        if total != 0:
            ng = pd.DataFrame([[fecha_comp, prov_sel, pv, tipo_f, n21, n10, r_iva, r_gan, r_iibb, nograv, total]], columns=COL_COMPRAS)
            st.session_state.compras = pd.concat([st.session_state.compras, ng], ignore_index=True)
            guardar_datos("compras", st.session_state.compras)
            st.session_state.msg_gasto = f"✅ Comprobante de '{prov_sel}' guardado por $ {total:,.2f}."
            st.rerun()
        else:
            st.warning("Ingresá al menos un importe para registrar el comprobante.")

elif sel == "COMPROBANTES":
    st.header("📜 Historial de Viajes")

    tab_ver_comp, tab_editar = st.tabs(["📋 VER Y ELIMINAR", "✏️ EDITAR VIAJE"])

    with tab_ver_comp:
        if not st.session_state.viajes.empty:
            for i in reversed(st.session_state.viajes.index):
                row = st.session_state.viajes.loc[i]
                c1, c2, c3 = st.columns([0.2, 0.6, 0.1])
                c1.write(f"📅 {row['Fecha Viaje']}")
                c2.write(f"👤 **{row['Cliente']}** | {row['Origen']} a {row['Destino']} | **${row['Importe']}**")
                if c3.button("🗑️", key=f"del_{i}"):
                    st.session_state.viajes = st.session_state.viajes.drop(i)
                    guardar_datos("viajes", st.session_state.viajes); st.rerun()
                st.divider()
        else:
            st.info("No hay viajes registrados.")

    with tab_editar:
        if st.session_state.viajes.empty:
            st.info("No hay viajes para editar.")
        else:
            # Filtros para encontrar el viaje rápido
            col_f1, col_f2 = st.columns(2)
            clientes_unicos = ["(Todos)"] + sorted(st.session_state.viajes['Cliente'].unique().tolist())
            filtro_cli = col_f1.selectbox("Filtrar por cliente", clientes_unicos, key="edit_filtro_cli")
            filtro_txt = col_f2.text_input("Buscar por origen / destino", key="edit_filtro_txt").strip().lower()

            df_edit = st.session_state.viajes.copy()
            if filtro_cli != "(Todos)":
                df_edit = df_edit[df_edit['Cliente'] == filtro_cli]
            if filtro_txt:
                df_edit = df_edit[
                    df_edit['Origen'].str.lower().str.contains(filtro_txt, na=False) |
                    df_edit['Destino'].str.lower().str.contains(filtro_txt, na=False)
                ]

            if df_edit.empty:
                st.info("No se encontraron viajes con ese filtro.")
            else:
                st.markdown(f"**{len(df_edit)} viaje(s) encontrado(s)**")
                for i in reversed(df_edit.index):
                    row = st.session_state.viajes.loc[i]
                    with st.container():
                        # Encabezado del viaje
                        col_info, col_btn = st.columns([0.85, 0.15])
                        col_info.markdown(
                            f"📅 `{row['Fecha Viaje']}` — **{row['Cliente']}** | "
                            f"{row['Origen']} ➔ {row['Destino']} | **$ {row['Importe']:,.2f}**"
                        )
                        if col_btn.button("✏️ Editar", key=f"abrir_edit_{i}"):
                            st.session_state[f"modo_edit_viaje_{i}"] = not st.session_state.get(f"modo_edit_viaje_{i}", False)
                            st.rerun()

                        # Formulario de edición inline
                        if st.session_state.get(f"modo_edit_viaje_{i}", False):
                            with st.form(f"form_edit_viaje_{i}"):
                                st.markdown(f"##### ✏️ Editando viaje #{i}")
                                ec1, ec2 = st.columns(2)
                                # Parsear fecha correctamente
                                try:
                                    fecha_actual = pd.to_datetime(row['Fecha Viaje']).date()
                                except:
                                    fecha_actual = date.today()
                                n_fecha  = ec1.date_input("Fecha Viaje", value=fecha_actual)
                                n_cli    = ec2.selectbox("Cliente", st.session_state.clientes['Razón Social'].unique() if not st.session_state.clientes.empty else [""], index=list(st.session_state.clientes['Razón Social'].unique()).index(row['Cliente']) if row['Cliente'] in st.session_state.clientes['Razón Social'].unique() else 0)
                                ec3, ec4 = st.columns(2)
                                n_orig   = ec3.text_input("Origen", value=str(row['Origen']))
                                n_dest   = ec4.text_input("Destino", value=str(row['Destino']))
                                ec5, ec6 = st.columns(2)
                                n_pat    = ec5.text_input("Patente / Móvil", value=str(row['Patente / Móvil']))
                                n_imp    = ec6.number_input("Importe $", value=float(row['Importe']), min_value=0.0, step=0.01)
                                n_tipo   = st.selectbox("Tipo Comprobante", ["Factura (Cuenta Corriente)", "Factura (Contado)", "RECIBO", "REMITO"], index=0)
                                sb1, sb2 = st.columns(2)
                                if sb1.form_submit_button("💾 GUARDAR CAMBIOS", type="primary"):
                                    st.session_state.viajes.loc[i, 'Fecha Viaje']    = str(n_fecha)
                                    st.session_state.viajes.loc[i, 'Cliente']        = n_cli
                                    st.session_state.viajes.loc[i, 'Origen']         = n_orig
                                    st.session_state.viajes.loc[i, 'Destino']        = n_dest
                                    st.session_state.viajes.loc[i, 'Patente / Móvil']= n_pat
                                    st.session_state.viajes.loc[i, 'Importe']        = n_imp
                                    st.session_state.viajes.loc[i, 'Tipo Comp']      = n_tipo
                                    guardar_datos("viajes", st.session_state.viajes)
                                    st.session_state[f"modo_edit_viaje_{i}"] = False
                                    st.session_state.msg_viaje = f"✅ Viaje #{i} actualizado correctamente."
                                    st.rerun()
                                if sb2.form_submit_button("❌ Cancelar"):
                                    st.session_state[f"modo_edit_viaje_{i}"] = False
                                    st.rerun()
                    st.divider()
elif sel == "CTA CTE PROVEEDOR":
    st.header("📊 Cuenta Corriente por Proveedor")
    if not st.session_state.proveedores.empty:
        p_sel = st.selectbox("Seleccionar Proveedor", st.session_state.proveedores['Razón Social'].unique())

        # ── SALDO INICIAL (MIGRACIÓN) ──
        with st.expander("⚙️ Cargar Saldo Inicial (migración desde otro sistema)", expanded=False):
            st.markdown(
                "<div style='background:#fff8e1;border-left:4px solid #f39c12;border-radius:6px;"
                "padding:10px 16px;font-size:13px;margin-bottom:12px;'>"
                "⚠️ <b>Usá esta opción solo para cargar el saldo que le debías al proveedor en el sistema anterior.</b><br>"
                "Se registrará como <i>SALDO INICIAL</i> con la fecha que indiques. No genera comprobante fiscal."
                "</div>", unsafe_allow_html=True
            )
            with st.form("f_saldo_ini_prov", clear_on_submit=True):
                sp1, sp2, sp3 = st.columns(3)
                fecha_sp  = sp1.date_input("Fecha del saldo", value=date.today(), key="si_prov_fecha")
                monto_sp  = sp2.number_input("Saldo que le debés al proveedor $",
                                             min_value=0.0, step=100.0, format="%.2f", key="si_prov_monto",
                                             help="Ingresá lo que le debés. Si el proveedor tiene crédito a tu favor, marcá 'A favor de la empresa'.")
                favor_emp = sp3.checkbox("¿Saldo a favor de la EMPRESA?", value=False, key="si_prov_favor",
                                         help="Marcá si el proveedor te debe algo (anticipo, devolución, etc.).")
                obs_sp    = st.text_input("Observación (opcional)", placeholder="Ej: Saldo migrado de sistema anterior", key="si_prov_obs")
                if st.form_submit_button("✅ CARGAR SALDO INICIAL"):
                    if monto_sp > 0:
                        # Si debemos al proveedor → total positivo (deuda nuestra)
                        # Si el proveedor nos debe → total negativo (crédito nuestro)
                        total_sp  = -monto_sp if favor_emp else monto_sp
                        concepto_sp = obs_sp if obs_sp else "SALDO INICIAL — migración"
                        nc_si = pd.DataFrame([[
                            str(fecha_sp), p_sel, "-", "SALDO INICIAL",
                            0, 0, 0, 0, 0, 0, total_sp
                        ]], columns=COL_COMPRAS)
                        st.session_state.compras = pd.concat([st.session_state.compras, nc_si], ignore_index=True)
                        guardar_datos("compras", st.session_state.compras)
                        signo_txt = "a favor de la empresa" if favor_emp else "a pagar al proveedor"
                        st.success(f"✅ Saldo inicial de $ {monto_sp:,.2f} ({signo_txt}) cargado para {p_sel}.")
                        st.rerun()
                    else:
                        st.warning("Ingresá un monto mayor a cero.")

        st.markdown("---")
        df_p  = st.session_state.compras[st.session_state.compras['Proveedor'] == p_sel].copy()

        # Mostrar detalle cronológico con Debe/Haber/Saldo
        if df_p.empty:
            st.info(f"No hay movimientos registrados para {p_sel}.")
        else:
            df_p['Total'] = pd.to_numeric(df_p['Total'], errors='coerce').fillna(0)
            df_p_sorted = df_p.sort_values('Fecha', na_position='last').reset_index(drop=True)

            saldo_ini_prov   = df_p[df_p['Tipo Factura'] == 'SALDO INICIAL']['Total'].sum()
            saldo_total_prov = df_p['Total'].sum()
            total_facturas   = df_p[df_p['Total'] > 0]['Total'].sum()
            total_pagos      = df_p[df_p['Total'] < 0]['Total'].sum()

            sp_m1, sp_m2, sp_m3 = st.columns(3)
            sp_m1.metric("Total Facturado",    f"$ {total_facturas:,.2f}")
            sp_m2.metric("Total Pagado",        f"$ {abs(total_pagos):,.2f}")
            sp_m3.metric("SALDO PENDIENTE",     f"$ {saldo_total_prov:,.2f}")
            if saldo_ini_prov != 0:
                st.caption(f"Incluye saldo inicial migrado: $ {saldo_ini_prov:,.2f}")

            st.markdown("---")

            # Tabla con Debe / Haber / Saldo acumulado
            saldo_acum = 0.0
            filas_det = []
            for _, r in df_p_sorted.iterrows():
                monto = float(r['Total'])
                saldo_acum += monto
                filas_det.append({
                    "Fecha":        str(r.get('Fecha', '-')),
                    "Tipo":         str(r.get('Tipo Factura', '-')),
                    "Punto Venta":  str(r.get('Punto Venta', '-')),
                    "Debe":         f"$ {monto:,.2f}"   if monto > 0 else "",
                    "Haber":        f"$ {abs(monto):,.2f}" if monto < 0 else "",
                    "Saldo":        f"$ {saldo_acum:,.2f}",
                })
            st.dataframe(pd.DataFrame(filas_det), use_container_width=True, hide_index=True)
    else:
        st.info("No hay proveedores registrados.")

elif sel == "CTA CTE GENERAL PROV":
    st.header("🌎 Estado General de Proveedores")
    if not st.session_state.compras.empty:
        res_p = st.session_state.compras.groupby('Proveedor')['Total'].sum().reset_index()
        res_p = res_p.merge(
            st.session_state.proveedores[['Razón Social', 'CBU', 'Alias']],
            left_on='Proveedor', right_on='Razón Social', how='left'
        ).drop(columns='Razón Social')
        res_p['CBU']   = res_p['CBU'].fillna('-')
        res_p['Alias'] = res_p['Alias'].fillna('-')
        res_p = res_p[['Proveedor', 'CBU', 'Alias', 'Total']]
        # Excluir saldos cero o residuales
        res_p['Total'] = res_p['Total'].apply(lambda x: round(float(x), 2))
        res_p = res_p[res_p['Total'].apply(lambda x: abs(x) > 0.01)].reset_index(drop=True)
        # Métricas
        mp1, mp2 = st.columns(2)
        mp1.metric("Total proveedores con saldo", len(res_p))
        mp2.metric("Total a pagar", f"$ {res_p['Total'].sum():,.2f}")
        st.dataframe(res_p.style.format({"Total": "$ {:,.2f}"}), use_container_width=True)
        # Descarga PDF — excluir saldos en cero o residuales
        res_p_pdf = res_p[['Proveedor','Total']].copy()
        res_p_pdf['Total'] = res_p_pdf['Total'].apply(lambda x: round(float(x), 2))
        res_p_pdf = res_p_pdf[res_p_pdf['Total'].apply(lambda x: abs(x) > 0.01)]
        res_p_pdf = res_p_pdf.rename(columns={'Proveedor':'Cliente','Total':'Importe'})
        html_prov_pdf = generar_html_cta_cte_general("Proveedores", res_p_pdf, date.today())
        st.download_button(
            label="🖨️ DESCARGAR PDF / IMPRIMIR",
            data=html_prov_pdf,
            file_name=f"CTA_CTE_General_Proveedores_{date.today()}.html",
            mime="text/html",
            help="Abrí el archivo descargado en el navegador y usá Ctrl+P para imprimir o guardar como PDF"
        )
    else:
        st.info("No hay comprobantes registrados.")

elif sel == "HISTORICO COMPRAS":
    st.header("📜 Comprobantes Cargados")
    if not st.session_state.compras.empty:
        for i in reversed(st.session_state.compras.index):
            row = st.session_state.compras.loc[i]
            c1, c2, c3 = st.columns([0.2, 0.6, 0.1])
            c1.write(f"📅 {row['Fecha']}")
            c2.write(f"👤 **{row['Proveedor']}** | {row['Tipo Factura']} {row['Punto Venta']} | **${row['Total']:,.2f}**")
            if c3.button("🗑️", key=f"del_comp_{i}"):
                st.session_state.compras = st.session_state.compras.drop(i)
                guardar_datos("compras", st.session_state.compras); st.rerun()
            st.divider()

elif sel == "MAYOR DE CUENTAS":
    st.header("📒 Mayor de Cuentas")
    st.caption("Resumen contable por cuenta: Ingresos, IVA Ventas, Gastos, IVA Compras y retenciones.")

    # ── Filtros de período ──
    mc1, mc2 = st.columns(2)
    mc_desde = mc1.date_input("Desde", value=date(date.today().year, 1, 1), key="mc_desde")
    mc_hasta = mc2.date_input("Hasta", value=date.today(), key="mc_hasta")

    # ════════════════════════════════════════════
    # SECCIÓN 1: INGRESOS (viajes)
    # ════════════════════════════════════════════
    df_viajes_mc = st.session_state.viajes.copy()
    df_viajes_mc['Importe'] = pd.to_numeric(df_viajes_mc['Importe'], errors='coerce').fillna(0)

    # Buscar la columna de fecha correcta (puede ser 'Fecha Viaje' o 'Fecha Carga')
    col_fecha_viaje = None
    for _cf in ['Fecha Viaje', 'Fecha Carga', 'Fecha']:
        if _cf in df_viajes_mc.columns:
            col_fecha_viaje = _cf
            break

    if col_fecha_viaje and not df_viajes_mc.empty:
        df_viajes_mc['Fecha_dt'] = pd.to_datetime(df_viajes_mc[col_fecha_viaje], errors='coerce')
        mask_v = (df_viajes_mc['Fecha_dt'].dt.date >= mc_desde) & (df_viajes_mc['Fecha_dt'].dt.date <= mc_hasta)
        df_viajes_mc = df_viajes_mc[mask_v]

    total_ingresos = df_viajes_mc['Importe'].sum()
    iva_ventas = 0.0
    neto_ventas = total_ingresos

    # ════════════════════════════════════════════
    # SECCIÓN 2: COMPRAS (gastos + IVA compras)
    # ════════════════════════════════════════════
    df_compras_mc = st.session_state.compras.copy()
    try:
        df_compras_mc['Fecha_dt'] = pd.to_datetime(df_compras_mc['Fecha'], errors='coerce')
        df_compras_mc = df_compras_mc[
            (df_compras_mc['Fecha_dt'].dt.date >= mc_desde) &
            (df_compras_mc['Fecha_dt'].dt.date <= mc_hasta)
        ]
    except: pass

    # Calcular IVA compras: neto21*0.21 + neto10.5*0.105
    df_compras_mc['IVA_21']  = df_compras_mc['Neto 21']  * 0.21
    df_compras_mc['IVA_105'] = df_compras_mc['Neto 10.5'] * 0.105
    total_iva_compras = df_compras_mc['IVA_21'].sum() + df_compras_mc['IVA_105'].sum()
    total_neto_compras = df_compras_mc['Neto 21'].sum() + df_compras_mc['Neto 10.5'].sum()
    total_no_grav = df_compras_mc['No Gravados'].sum()
    total_ret_iva = df_compras_mc['Ret IVA'].sum()
    total_ret_gan = df_compras_mc['Ret Ganancia'].sum()
    total_ret_iibb = df_compras_mc['Ret IIBB'].sum()
    total_compras  = df_compras_mc['Total'].sum()

    # Gastos por cuenta (join con proveedores para la cuenta de gastos)
    gastos_por_cuenta = pd.DataFrame()
    if not df_compras_mc.empty and not st.session_state.proveedores.empty:
        df_gc = df_compras_mc.merge(
            st.session_state.proveedores[['Razón Social', 'Cuenta de Gastos']],
            left_on='Proveedor', right_on='Razón Social', how='left'
        )
        df_gc['Cuenta de Gastos'] = df_gc['Cuenta de Gastos'].fillna('SIN CLASIFICAR')
        gastos_por_cuenta = df_gc.groupby('Cuenta de Gastos')['Total'].sum().reset_index()
        gastos_por_cuenta = gastos_por_cuenta.sort_values('Total', ascending=False)

    # ── IVA Neto a pagar/favor ──
    iva_neto = iva_ventas - total_iva_compras

    # ════════════════════════════════════════════
    # MÉTRICAS RESUMEN
    # ════════════════════════════════════════════
    st.markdown("### 📊 Resumen del Período")
    km1, km2, km3, km4 = st.columns(4)
    km1.metric("💰 Ingresos Totales", f"$ {total_ingresos:,.2f}")
    km2.metric("🚌 IVA Ventas", "No aplica (transporte)")
    km3.metric("🛒 Gastos Totales", f"$ {total_compras:,.2f}")
    km4.metric("📋 IVA Compras", f"$ {total_iva_compras:,.2f}")
    # IVA neto = solo crédito fiscal de compras (no hay débito fiscal en ventas de transporte)
    iva_neto = -total_iva_compras  # es crédito a favor
    color_iva_neto = "#2ecc71"
    st.markdown(
        f"<div style='background:#f0f2f6;border-radius:10px;padding:14px 24px;margin:10px 0;"
        f"border-left:5px solid {color_iva_neto};'>"
        f"<b>IVA Crédito Fiscal (a favor — solo compras):</b> "
        f"<span style='font-size:22px;font-weight:bold;color:{color_iva_neto};'>$ {abs(iva_neto):,.2f}</span>"
        f"<span style='font-size:12px;color:#888;'> — Los ingresos por transporte de pasajeros no generan débito fiscal</span>"
        f"</div>", unsafe_allow_html=True
    )

    st.markdown("---")

    # ════════════════════════════════════════════
    # TABS DEL MAYOR
    # ════════════════════════════════════════════
    tab_ing, tab_gtos, tab_iva, tab_ret = st.tabs(["💰 Ingresos", "🛒 Gastos por Cuenta", "📋 IVA", "📌 Retenciones"])

    with tab_ing:
        st.markdown("#### Cuenta: INGRESOS POR SERVICIOS DE TRANSPORTE")
        if df_viajes_mc.empty:
            st.info("No hay viajes registrados en el período seleccionado.")
        else:
            df_mayor_ing = df_viajes_mc.copy()

            # Usar la columna de fecha que encontramos
            if col_fecha_viaje and col_fecha_viaje in df_mayor_ing.columns:
                df_mayor_ing['Fecha'] = pd.to_datetime(df_mayor_ing[col_fecha_viaje], errors='coerce').dt.strftime('%d/%m/%Y')
            else:
                df_mayor_ing['Fecha'] = '-'

            df_mayor_ing = df_mayor_ing.sort_values('Fecha_dt', na_position='last') if 'Fecha_dt' in df_mayor_ing.columns else df_mayor_ing
            df_mayor_ing['Debe'] = 0.0
            df_mayor_ing = df_mayor_ing.rename(columns={'Importe': 'Haber'})
            df_mayor_ing['Saldo Acum.'] = df_mayor_ing['Haber'].cumsum()

            cols_mostrar = ['Fecha', 'Cliente', 'Debe', 'Haber', 'Saldo Acum.']
            cols_ok = [c for c in cols_mostrar if c in df_mayor_ing.columns]
            st.dataframe(
                df_mayor_ing[cols_ok].style.format({"Debe": "$ {:,.2f}", "Haber": "$ {:,.2f}", "Saldo Acum.": "$ {:,.2f}"}),
                use_container_width=True
            )
            st.markdown(
                f"<div style='background:#5e2d61;color:white;border-radius:8px;padding:12px 20px;"
                f"display:flex;justify-content:space-between;align-items:center;margin-top:8px;'>"
                f"<span style='font-size:14px;font-weight:bold;'>TOTAL INGRESOS DEL PERÍODO</span>"
                f"<span style='font-size:22px;font-weight:bold;'>$ {total_ingresos:,.2f}</span>"
                f"</div>", unsafe_allow_html=True
            )

    with tab_gtos:
        st.markdown("#### Gastos por Cuenta Contable")
        if not gastos_por_cuenta.empty:
            fig_gc = px.bar(gastos_por_cuenta, x='Cuenta de Gastos', y='Total',
                           color_discrete_sequence=['#5e2d61'],
                           labels={'Total': 'Total $', 'Cuenta de Gastos': 'Cuenta'})
            fig_gc.update_layout(showlegend=False, plot_bgcolor='white', height=320)
            st.plotly_chart(fig_gc, use_container_width=True)
            st.dataframe(gastos_por_cuenta.style.format({"Total": "$ {:,.2f}"}), use_container_width=True)
        else:
            st.info("No hay compras en el período seleccionado.")

    with tab_iva:
        st.markdown("#### Posición de IVA")
        st.info("ℹ️ Los servicios de transporte de pasajeros **no están gravados con IVA** (exentos). Solo se registra el crédito fiscal de compras.")
        iva_data = pd.DataFrame({
            "Concepto": ["IVA Ventas (Débito Fiscal)",
                         "IVA Compras 21% (Crédito Fiscal)",
                         "IVA Compras 10.5% (Crédito Fiscal)",
                         "IVA CRÉDITO FISCAL NETO (a favor)"],
            "Importe": [0.0,
                        df_compras_mc['IVA_21'].sum(),
                        df_compras_mc['IVA_105'].sum(),
                        total_iva_compras]
        })
        st.dataframe(iva_data.style.format({"Importe": "$ {:,.2f}"}), use_container_width=True)
        # Detalle facturas con IVA
        if not df_compras_mc.empty:
            with st.expander("Ver detalle comprobantes"):
                cols_iva = ['Fecha', 'Proveedor', 'Tipo Factura', 'Neto 21', 'Neto 10.5', 'IVA_21', 'IVA_105', 'Total']
                st.dataframe(df_compras_mc[cols_iva].rename(columns={'IVA_21':'IVA 21%','IVA_105':'IVA 10.5%'})\
                    .style.format({c: "$ {:,.2f}" for c in ['Neto 21','Neto 10.5','IVA 21%','IVA 10.5%','Total']}),
                    use_container_width=True)

    with tab_ret:
        st.markdown("#### Retenciones del Período")
        ret_data = pd.DataFrame({
            "Retención": ["Retención IVA", "Retención Ganancias", "Retención IIBB", "TOTAL"],
            "Importe": [total_ret_iva, total_ret_gan, total_ret_iibb,
                        total_ret_iva + total_ret_gan + total_ret_iibb]
        })
        st.dataframe(ret_data.style.format({"Importe": "$ {:,.2f}"}), use_container_width=True)

    st.markdown("---")

    # ════════════════════════════════════════════
    # DESCARGA TXT
    # ════════════════════════════════════════════
    st.markdown("### 📥 Exportar Mayor de Cuentas")

    lineas_txt = []
    lineas_txt.append("=" * 60)
    lineas_txt.append("    CHACABUCO NOROESTE TOUR S.R.L. — MAYOR DE CUENTAS")
    lineas_txt.append(f"    Período: {mc_desde} al {mc_hasta}")
    lineas_txt.append(f"    Emitido: {date.today()}")
    lineas_txt.append("=" * 60)
    lineas_txt.append("")
    lineas_txt.append("── INGRESOS ──────────────────────────────────────────")
    lineas_txt.append(f"  Ingresos Totales (servicios transp.): $ {total_ingresos:>12,.2f}")
    lineas_txt.append(f"  IVA Ventas: NO APLICA (exento - transporte de pasajeros)")
    lineas_txt.append("")
    if not df_viajes_mc.empty:
        lineas_txt.append("  Detalle por cliente:")
        for _, row_i in df_viajes_mc.groupby('Cliente')['Importe'].sum().reset_index().sort_values('Importe', ascending=False).iterrows():
            lineas_txt.append(f"    {row_i['Cliente']:<40} $ {row_i['Importe']:>12,.2f}")
        lineas_txt.append("")
        lineas_txt.append("  Detalle de viajes:")
        cols_txt = ['Fecha Viaje', 'Cliente', 'Origen', 'Destino', 'Importe']
        for _, row_v in df_viajes_mc.sort_values('Fecha Viaje', ascending=False).iterrows():
            fecha_v  = str(row_v.get('Fecha Viaje', '-'))
            cliente_v= str(row_v.get('Cliente', '-'))[:25]
            origen_v = str(row_v.get('Origen', '-'))[:15]
            destino_v= str(row_v.get('Destino', '-'))[:15]
            imp_v    = float(row_v.get('Importe', 0))
            lineas_txt.append(f"    {fecha_v:<12} {cliente_v:<26} {origen_v:<16} {destino_v:<16} $ {imp_v:>10,.2f}")
    lineas_txt.append("")
    if not df_viajes_mc.empty:
        for _, row_i in df_viajes_mc.groupby('Cliente')['Importe'].sum().reset_index().iterrows():
            lineas_txt.append(f"    {row_i['Cliente']:<40} $ {row_i['Importe']:>12,.2f}")
    lineas_txt.append("")
    lineas_txt.append("── GASTOS POR CUENTA ─────────────────────────────────")
    if not gastos_por_cuenta.empty:
        for _, row_g in gastos_por_cuenta.iterrows():
            lineas_txt.append(f"  {row_g['Cuenta de Gastos']:<40} $ {row_g['Total']:>12,.2f}")
    lineas_txt.append(f"  {'TOTAL GASTOS':<40} $ {total_compras:>12,.2f}")
    lineas_txt.append("")
    lineas_txt.append("── IVA ───────────────────────────────────────────────")
    lineas_txt.append(f"  IVA Ventas (Débito Fiscal):        NO APLICA (transporte exento)")
    lineas_txt.append(f"  IVA Compras 21% (Crédito Fiscal): $ {df_compras_mc['IVA_21'].sum():>15,.2f}")
    lineas_txt.append(f"  IVA Compras 10.5% (Crédito):      $ {df_compras_mc['IVA_105'].sum():>15,.2f}")
    lineas_txt.append(f"  {'CRÉDITO FISCAL NETO (a favor)':<38}  $ {total_iva_compras:>15,.2f}")
    lineas_txt.append("")
    lineas_txt.append("── RETENCIONES ───────────────────────────────────────")
    lineas_txt.append(f"  Retención IVA:                    $ {total_ret_iva:>15,.2f}")
    lineas_txt.append(f"  Retención Ganancias:              $ {total_ret_gan:>15,.2f}")
    lineas_txt.append(f"  Retención IIBB:                   $ {total_ret_iibb:>15,.2f}")
    lineas_txt.append(f"  {'TOTAL RETENCIONES':<38}  $ {(total_ret_iva+total_ret_gan+total_ret_iibb):>15,.2f}")
    lineas_txt.append("")
    lineas_txt.append("=" * 60)
    lineas_txt.append("  CHACAGEST Software System — www.chacagest.com.ar")
    lineas_txt.append("=" * 60)

    txt_content = "\n".join(lineas_txt)

    col_dl1, col_dl2 = st.columns(2)
    col_dl1.download_button(
        label="📄 DESCARGAR TXT",
        data=txt_content,
        file_name=f"Mayor_Cuentas_{mc_desde}_{mc_hasta}.txt",
        mime="text/plain"
    )

    # HTML para impresión
    filas_gtos_html = "".join(
        f"<tr><td>{r['Cuenta de Gastos']}</td><td style='text-align:right'>$ {r['Total']:,.2f}</td></tr>"
        for _, r in gastos_por_cuenta.iterrows()
    ) if not gastos_por_cuenta.empty else "<tr><td colspan='2'>Sin datos</td></tr>"

    html_mayor = f"""<html><head><style>
        body{{font-family:'Segoe UI',Arial,sans-serif;color:#333;padding:24px;font-size:13px;}}
        .tit{{color:#5e2d61;font-size:20px;font-weight:bold;}}
        h3{{color:#5e2d61;border-bottom:2px solid #5e2d61;padding-bottom:4px;}}
        table{{width:100%;border-collapse:collapse;margin-bottom:20px;}}
        th{{background:#5e2d61;color:white;padding:9px;text-align:left;font-size:12px;}}
        td{{border-bottom:1px solid #eee;padding:8px 10px;}}
        .box{{background:#f0f2f6;border-radius:8px;padding:12px 20px;margin:10px 0;
              border-left:4px solid #5e2d61;display:flex;justify-content:space-between;}}
        .num{{font-size:18px;font-weight:bold;color:#5e2d61;}}
    </style></head><body>
    <div class="tit">CHACABUCO NOROESTE TOUR S.R.L. — Mayor de Cuentas</div>
    <p style="color:#888;">Período: <b>{mc_desde}</b> al <b>{mc_hasta}</b> | Emitido: {date.today()}</p>
    <h3>💰 Ingresos</h3>
    <div class="box"><span>Ingresos Brutos</span><span class="num">$ {total_ingresos:,.2f}</span></div>
    <div class="box"><span>Ingresos Totales (servicios transporte)</span><span class="num">$ {total_ingresos:,.2f}</span></div>
    <div class="box"><span>IVA Ventas</span><span style="color:#888;font-style:italic;">No aplica — transporte de pasajeros exento</span></div>
    <h3>🛒 Gastos por Cuenta</h3>
    <table><tr><th>Cuenta de Gastos</th><th style="text-align:right">Total</th></tr>
    {filas_gtos_html}
    <tr style="font-weight:bold;background:#f8f9fa;"><td>TOTAL GASTOS</td><td style="text-align:right">$ {total_compras:,.2f}</td></tr>
    </table>
    <h3>📋 Posición IVA</h3>
    <table><tr><th>Concepto</th><th style="text-align:right">Importe</th></tr>
    <tr><td>IVA Ventas (Débito Fiscal)</td><td style="text-align:right;color:#888;font-style:italic;">No aplica (transporte exento)</td></tr>
    <tr><td>IVA Compras 21% (Crédito)</td><td style="text-align:right">$ {df_compras_mc['IVA_21'].sum():,.2f}</td></tr>
    <tr><td>IVA Compras 10.5% (Crédito)</td><td style="text-align:right">$ {df_compras_mc['IVA_105'].sum():,.2f}</td></tr>
    <tr style="font-weight:bold;background:#f8f9fa;"><td>Crédito Fiscal Neto (a favor)</td><td style="text-align:right;color:#2ecc71">$ {total_iva_compras:,.2f}</td></tr>
    </table>
    <h3>📌 Retenciones</h3>
    <table><tr><th>Tipo</th><th style="text-align:right">Importe</th></tr>
    <tr><td>IVA</td><td style="text-align:right">$ {total_ret_iva:,.2f}</td></tr>
    <tr><td>Ganancias</td><td style="text-align:right">$ {total_ret_gan:,.2f}</td></tr>
    <tr><td>IIBB</td><td style="text-align:right">$ {total_ret_iibb:,.2f}</td></tr>
    <tr style="font-weight:bold;background:#f8f9fa;"><td>TOTAL</td><td style="text-align:right">$ {(total_ret_iva+total_ret_gan+total_ret_iibb):,.2f}</td></tr>
    </table>
    </body></html>"""

    col_dl2.download_button(
        label="🖨️ DESCARGAR PDF / IMPRIMIR",
        data=html_mayor,
        file_name=f"Mayor_Cuentas_{mc_desde}_{mc_hasta}.html",
        mime="text/html",
        help="Abrí el HTML en el navegador y usá Ctrl+P para imprimir o guardar como PDF"
    )


# =============================================================
# FACTURACIÓN
# =============================================================
elif sel == "FACTURACION":
    st.header("🧾 Facturación")
    if "facturas" not in st.session_state:
        st.session_state.facturas = pd.DataFrame(columns=COL_FACTURAS)

    tab_nueva, tab_historial, tab_ctacte = st.tabs(["📝 Emitir Comprobante", "📂 Historial", "📒 Cta. Cte. Clientes"])

    # ─────────────────────────────────────────────
    # TAB 1 — EMITIR COMPROBANTE
    # ─────────────────────────────────────────────
    with tab_nueva:
        if "html_factura_ready" not in st.session_state:
            st.session_state.html_factura_ready = None
        if "items_factura" not in st.session_state:
            st.session_state.items_factura = []

        if st.session_state.html_factura_ready:
            st.success("✅ Comprobante generado. Descargá el documento para imprimir.")
            tipo_dl = st.session_state.get("ultimo_tipo_fac", "Factura")
            nro_dl  = st.session_state.get("ultimo_nro_fac", "")
            st.download_button(
                "🖨️ DESCARGAR COMPROBANTE",
                st.session_state.html_factura_ready,
                file_name=f"{tipo_dl}_{nro_dl}_{date.today()}.html",
                mime="text/html"
            )
            if st.button("🔄 Nuevo Comprobante"):
                st.session_state.html_factura_ready = None
                st.session_state.items_factura = []
                st.rerun()
        else:
            hoy = date.today()
            st.markdown("##### 📋 Datos del Comprobante")
            c1, c2, c3, c4 = st.columns(4)
            tipo_comp = c1.selectbox("Tipo", ["FACTURA", "NOTA DE CREDITO", "NOTA DE DEBITO"], key="fac_tipo")
            letra_comp = c2.selectbox("Letra", ["A", "B", "C", "M", "X"], key="fac_letra")
            pv_comp = c3.text_input("Punto de Venta", value="0001", max_chars=4, key="fac_pv")

            # Autonumerar según tipo+letra+pv
            mask_auto = (
                (st.session_state.facturas['Tipo'] == tipo_comp) &
                (st.session_state.facturas['Punto Venta'] == pv_comp.zfill(4))
            )
            ultimo_nro = 0
            if not st.session_state.facturas.empty and mask_auto.any():
                try:
                    ultimo_nro = int(st.session_state.facturas[mask_auto]['Numero'].apply(lambda x: int(str(x).replace("-",""))).max())
                except:
                    ultimo_nro = 0
            sugerido = str(ultimo_nro + 1).zfill(8)
            nro_comp = c4.text_input("Número", value=sugerido, max_chars=8, key="fac_nro")

            c5, c6 = st.columns(2)
            fecha_comp = c5.date_input("Fecha", value=hoy, key="fac_fecha")

            comp_asoc = ""
            if tipo_comp in ["NOTA DE CREDITO", "NOTA DE DEBITO"]:
                comp_asoc = c6.text_input("Comprobante Asociado (ej: FAC B 0001-00000012)", key="fac_comp_asoc")
            else:
                c6.markdown("")

            st.markdown("---")
            st.markdown("##### 👤 Datos del Cliente")
            clientes_lista = [""] + sorted(st.session_state.clientes['Razón Social'].tolist()) if not st.session_state.clientes.empty else [""]
            c7, c8 = st.columns(2)
            cliente_sel = c7.selectbox("Cliente", clientes_lista, key="fac_cliente")

            # Autocompletar datos del cliente
            cuit_auto, cond_iva_auto, dir_auto = "", "", ""
            if cliente_sel and not st.session_state.clientes.empty:
                row_cli = st.session_state.clientes[st.session_state.clientes['Razón Social'] == cliente_sel]
                if not row_cli.empty:
                    cuit_auto    = str(row_cli.iloc[0].get('CUIT / CUIL / DNI *', ''))
                    cond_iva_auto = str(row_cli.iloc[0].get('Condición IVA', ''))
                    dir_auto     = str(row_cli.iloc[0].get('Dirección Fiscal', ''))
            c8_cuit, c8_cond = st.columns(2)
            cuit_comp   = c8_cuit.text_input("CUIT / DNI", value=cuit_auto, key="fac_cuit")
            cond_iva_comp = c8_cond.text_input("Condición IVA", value=cond_iva_auto, key="fac_cond_iva")
            dir_comp    = st.text_input("Dirección", value=dir_auto, key="fac_dir")

            st.markdown("---")
            st.markdown("##### 🛒 Ítems del Comprobante")

            # Tabla de items
            if st.session_state.items_factura:
                hdr1, hdr2, hdr3, hdr4, hdr5, hdr6 = st.columns([4, 1.2, 2, 1.5, 2, 1])
                hdr1.markdown("**Descripción**"); hdr2.markdown("**Cant.**")
                hdr3.markdown("**P. Unitario**"); hdr4.markdown("**IVA %**")
                hdr5.markdown("**Subtotal**");    hdr6.markdown("")
                for idx_i, it in enumerate(st.session_state.items_factura):
                    ci1, ci2, ci3, ci4, ci5, ci6 = st.columns([4, 1.2, 2, 1.5, 2, 1])
                    ci1.markdown(f"{it['descripcion']}")
                    ci2.markdown(f"{it['cantidad']}")
                    ci3.markdown(f"$ {float(it['precio_unitario']):,.2f}")
                    ci4.markdown(f"{it['alicuota']}")
                    ci5.markdown(f"$ {float(it['subtotal']):,.2f} *(c/IVA)*")
                    if ci6.button("🗑️", key=f"del_item_{idx_i}"):
                        st.session_state.items_factura.pop(idx_i)
                        st.rerun()
                st.markdown("---")

            # Formulario para agregar item
            with st.expander("➕ Agregar ítem", expanded=len(st.session_state.items_factura) == 0):
                ni1, ni2, ni3, ni4 = st.columns([4, 1.2, 2, 1.5])
                desc_it  = ni1.text_input("Descripción del servicio/producto", key="it_desc")
                cant_it  = ni2.number_input("Cant.", min_value=1, value=1, step=1, key="it_cant")
                precio_it = ni3.number_input("Precio Unitario $ (sin IVA)", min_value=0.0, step=10.0, format="%.2f", key="it_precio")
                alicuota_it = ni4.selectbox("IVA %", ["21%", "10.5%", "27%", "0%", "Exento", "No Gravado"], key="it_alicuota")
                if st.button("➕ Agregar ítem", key="btn_add_item"):
                    if desc_it and precio_it > 0:
                        # precio_unitario es SIN IVA; subtotal = neto * cantidad
                        subtotal_neto = float(cant_it) * float(precio_it)
                        tasa = {"21%": 0.21, "10.5%": 0.105, "27%": 0.27}.get(alicuota_it, 0.0)
                        subtotal_con_iva = subtotal_neto * (1 + tasa)
                        st.session_state.items_factura.append({
                            "descripcion": desc_it,
                            "cantidad": cant_it,
                            "precio_unitario": precio_it,
                            "alicuota": alicuota_it,
                            "subtotal": round(subtotal_con_iva, 2),   # total con IVA para mostrar
                            "subtotal_neto": round(subtotal_neto, 2)  # neto para los cálculos
                        })
                        st.rerun()
                    else:
                        st.warning("Completá descripción y precio.")

            # Cálculo de totales — precio unitario ingresado SIN IVA
            neto_calc = 0.0
            iva_calc  = 0.0
            no_grav_calc = 0.0
            for it in st.session_state.items_factura:
                ali  = it['alicuota']
                neto = float(it.get('subtotal_neto', float(it['subtotal'])))  # compatibilidad ítems viejos
                tasa = {"21%": 0.21, "10.5%": 0.105, "27%": 0.27}.get(ali, 0.0)
                if ali == "No Gravado":
                    no_grav_calc += neto
                else:
                    neto_calc += neto
                    iva_calc  += neto * tasa
            total_calc = neto_calc + iva_calc + no_grav_calc

            if st.session_state.items_factura:
                st.markdown("---")
                st.markdown("##### 💰 Totales")
                tc1, tc2, tc3, tc4 = st.columns(4)
                tc1.metric("Neto", f"$ {neto_calc:,.2f}")
                tc2.metric("IVA", f"$ {iva_calc:,.2f}")
                tc3.metric("No Gravado", f"$ {no_grav_calc:,.2f}")
                tc4.metric("TOTAL", f"$ {total_calc:,.2f}")

            st.markdown("---")
            obs_fac = st.text_area("Observaciones", key="fac_obs", height=60, placeholder="Ej: Pago a 30 días. Servicio de traslado escolar.")

            if st.button("🧾 GENERAR COMPROBANTE", type="primary"):
                if not cliente_sel:
                    st.warning("Seleccioná un cliente.")
                elif not st.session_state.items_factura:
                    st.warning("Agregá al menos un ítem.")
                else:
                    responsable = st.session_state.nombre_usuario
                    nro_final = str(nro_comp).zfill(8)
                    pv_final  = str(pv_comp).zfill(4)

                    # Guardar en facturas
                    nueva_fac = pd.DataFrame([[
                        str(fecha_comp), tipo_comp, pv_final, nro_final,
                        cliente_sel, cuit_comp, cond_iva_comp,
                        " / ".join([it['descripcion'] for it in st.session_state.items_factura]),
                        round(neto_calc, 2), round(iva_calc, 2), round(no_grav_calc, 2),
                        round(total_calc, 2), "EMITIDA",
                        tipo_comp if comp_asoc else "",
                        comp_asoc,
                        obs_fac.strip()
                    ]], columns=COL_FACTURAS)
                    st.session_state.facturas = pd.concat([st.session_state.facturas, nueva_fac], ignore_index=True)
                    guardar_datos("facturas", st.session_state.facturas)

                    # Mover cuenta corriente cliente
                    signo = -1 if tipo_comp == "NOTA DE CREDITO" else 1
                    tipo_mov_cc = {"FACTURA": "FACTURA", "NOTA DE CREDITO": "NOTA DE CREDITO", "NOTA DE DEBITO": "NOTA DE DEBITO"}.get(tipo_comp, "FACTURA")
                    mov_cc = pd.DataFrame([[
                        str(fecha_comp), tipo_mov_cc, "CTA CTE",
                        "CTA CTE", f"{tipo_comp} {letra_comp} {pv_final}-{nro_final}",
                        cliente_sel, signo * round(total_calc, 2),
                        f"{pv_final}-{nro_final}"
                    ]], columns=COL_TESORERIA)
                    st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, mov_cc], ignore_index=True)
                    guardar_datos("tesoreria", st.session_state.tesoreria)

                    # Generar HTML
                    html_fac = generar_html_factura({
                        "tipo":              tipo_comp,
                        "letra":             letra_comp,
                        "punto_venta":       pv_final,
                        "numero":            nro_final,
                        "fecha":             str(fecha_comp),
                        "cliente":           cliente_sel,
                        "cuit_cliente":      cuit_comp,
                        "condicion_iva":     cond_iva_comp,
                        "direccion_cliente": dir_comp,
                        "items":             st.session_state.items_factura,
                        "neto":              round(neto_calc, 2),
                        "iva_monto":         round(iva_calc, 2),
                        "no_gravado":        round(no_grav_calc, 2),
                        "total":             round(total_calc, 2),
                        "observaciones":     obs_fac.strip(),
                        "comp_asoc":         comp_asoc,
                        "logo_b64":          LOGO_B64,
                        "responsable":       responsable,
                    })
                    st.session_state.html_factura_ready = html_fac
                    st.session_state.ultimo_tipo_fac = tipo_comp.replace(" ","_")
                    st.session_state.ultimo_nro_fac  = f"{pv_final}-{nro_final}"
                    st.rerun()

    # ─────────────────────────────────────────────
    # TAB 2 — HISTORIAL
    # ─────────────────────────────────────────────
    with tab_historial:
        st.markdown("##### 📂 Comprobantes Emitidos")
        df_h = st.session_state.facturas.copy()
        if df_h.empty:
            st.info("No hay comprobantes registrados.")
        else:
            # Filtros
            fh1, fh2, fh3 = st.columns(3)
            tipos_hist = ["Todos"] + sorted(df_h['Tipo'].unique().tolist())
            tipo_filt = fh1.selectbox("Tipo", tipos_hist, key="hist_tipo")
            clientes_hist = ["Todos"] + sorted(df_h['Cliente'].unique().tolist())
            cli_filt = fh2.selectbox("Cliente", clientes_hist, key="hist_cli")
            try:
                fecha_filt = fh3.date_input("Desde", value=date.today().replace(day=1), key="hist_fecha")
                df_h['Fecha_dt'] = pd.to_datetime(df_h['Fecha'], errors='coerce')
                df_h = df_h[df_h['Fecha_dt'].dt.date >= fecha_filt]
            except:
                pass
            if tipo_filt != "Todos":
                df_h = df_h[df_h['Tipo'] == tipo_filt]
            if cli_filt != "Todos":
                df_h = df_h[df_h['Cliente'] == cli_filt]

            total_fac = df_h[df_h['Tipo'] == 'FACTURA']['Total'].sum()
            total_nc  = df_h[df_h['Tipo'] == 'NOTA DE CREDITO']['Total'].sum()
            total_nd  = df_h[df_h['Tipo'] == 'NOTA DE DEBITO']['Total'].sum()
            hm1, hm2, hm3, hm4 = st.columns(4)
            hm1.metric("Facturas", f"$ {total_fac:,.2f}")
            hm2.metric("NC emitidas", f"$ {total_nc:,.2f}")
            hm3.metric("ND emitidas", f"$ {total_nd:,.2f}")
            hm4.metric("Neto facturado", f"$ {(total_fac - total_nc + total_nd):,.2f}")
            st.markdown("---")

            cols_show = ['Fecha','Tipo','Punto Venta','Numero','Cliente','Total','Estado']
            cols_show = [c for c in cols_show if c in df_h.columns]
            st.dataframe(
                df_h[cols_show].drop(columns=['Fecha_dt'], errors='ignore'),
                use_container_width=True, hide_index=True
            )

            # Reemitir cualquier comprobante del historial
            st.markdown("---")
            st.markdown("##### 🔁 Reemitir comprobante")
            idx_opts = df_h.index.tolist()
            if idx_opts:
                idx_sel = st.selectbox(
                    "Seleccioná fila",
                    idx_opts,
                    format_func=lambda i: f"{df_h.loc[i,'Tipo']} {df_h.loc[i,'Punto Venta']}-{df_h.loc[i,'Numero']} | {df_h.loc[i,'Cliente']} | $ {float(df_h.loc[i,'Total']):,.2f}",
                    key="reemitir_sel"
                )
                if st.button("🖨️ Reemitir", key="btn_reemitir"):
                    row_r = st.session_state.facturas.loc[idx_sel]
                    detalle_items = [{"descripcion": row_r['Detalle'], "cantidad": 1,
                                      "precio_unitario": float(row_r['Total']), "alicuota": "0%",
                                      "subtotal": float(row_r['Total'])}]
                    html_rei = generar_html_factura({
                        "tipo": row_r['Tipo'], "letra": "B",
                        "punto_venta": row_r['Punto Venta'], "numero": row_r['Numero'],
                        "fecha": row_r['Fecha'], "cliente": row_r['Cliente'],
                        "cuit_cliente": row_r['CUIT Cliente'], "condicion_iva": row_r['Condicion IVA'],
                        "direccion_cliente": "", "items": detalle_items,
                        "neto": float(row_r['Neto']), "iva_monto": float(row_r['IVA']),
                        "no_gravado": float(row_r['No Gravado']), "total": float(row_r['Total']),
                        "observaciones": row_r['Observaciones'], "comp_asoc": row_r['Comp Asoc Nro'],
                        "logo_b64": LOGO_B64, "responsable": st.session_state.nombre_usuario,
                    })
                    st.download_button("⬇️ Descargar reemisión", html_rei,
                                       file_name=f"Reemision_{row_r['Punto Venta']}-{row_r['Numero']}.html",
                                       mime="text/html", key="dl_reemision")

    # ─────────────────────────────────────────────
    # TAB 3 — CUENTA CORRIENTE CLIENTES (unificada)
    # ─────────────────────────────────────────────
    with tab_ctacte:
        st.markdown("##### 📒 Cuenta Corriente por Cliente")
        st.info("💡 La cuenta corriente unificada (viajes + facturas + cobros) está disponible en **VENTAS → CTA CTE INDIVIDUAL**.")
        if st.session_state.clientes.empty:
            st.info("No hay clientes cargados.")
        else:
            clientes_cc = sorted(st.session_state.clientes['Razón Social'].tolist())
            cli_cc = st.selectbox("Cliente", clientes_cc, key="cc_cliente_fac")

            filas_unif_fac = []

            # Viajes del cliente
            df_v_fac = st.session_state.viajes[st.session_state.viajes['Cliente'] == cli_cc].copy()
            for _, r in df_v_fac.iterrows():
                imp = float(r['Importe'])
                origen = str(r.get('Origen', '-'))
                destino = str(r.get('Destino', '-'))
                comp_nro = str(r.get('Nro Comp Asoc', '-'))
                if origen in ("AJUSTE", "SALDO INICIAL"):
                    tipo_mov = "SALDO INICIAL"
                    concepto = str(r.get('Nro Comp Asoc', 'Saldo inicial'))
                elif imp < 0:
                    tipo_mov = "COBRANZA VIAJE"
                    concepto = f"Cobro viaje | Ref: {comp_nro}"
                else:
                    tipo_mov = "VIAJE"
                    concepto = f"{origen} → {destino}" + (f" | {comp_nro}" if comp_nro not in ['-','','nan'] else "")
                filas_unif_fac.append({
                    "Fecha": str(r.get('Fecha Viaje', r.get('Fecha Carga', '-'))),
                    "Tipo": tipo_mov, "Comprobante": concepto,
                    "Debe": imp if imp > 0 else 0.0,
                    "Haber": abs(imp) if imp < 0 else 0.0,
                })

            # Facturas y cobros desde tesorería
            df_t_fac = st.session_state.tesoreria[
                (st.session_state.tesoreria['Cliente/Proveedor'] == cli_cc) &
                (st.session_state.tesoreria['Tipo'].isin(['FACTURA','NOTA DE CREDITO','NOTA DE DEBITO','COBRO','COBRANZA','COBRANZA FACTURA']))
            ].copy()
            for _, r in df_t_fac.iterrows():
                monto = float(r['Monto'])
                tipo_m = r['Tipo'] if r['Tipo'] in ['FACTURA','NOTA DE CREDITO','NOTA DE DEBITO'] else 'COBRO'
                filas_unif_fac.append({
                    "Fecha": str(r['Fecha']), "Tipo": tipo_m,
                    "Comprobante": str(r['Concepto']),
                    "Debe": monto if monto > 0 else 0.0,
                    "Haber": abs(monto) if monto < 0 else 0.0,
                })

            if not filas_unif_fac:
                st.info(f"No hay movimientos para {cli_cc}.")
            else:
                df_unif_fac = pd.DataFrame(filas_unif_fac)
                df_unif_fac['Fecha_dt'] = pd.to_datetime(df_unif_fac['Fecha'], errors='coerce')
                df_unif_fac = df_unif_fac.sort_values('Fecha_dt', na_position='last').reset_index(drop=True)
                saldo_a = 0.0
                filas_show = []
                for _, r in df_unif_fac.iterrows():
                    saldo_a += r['Debe'] - r['Haber']
                    filas_show.append({
                        "Fecha": r['Fecha'], "Tipo": r['Tipo'], "Comprobante": r['Comprobante'],
                        "Debe":  f"$ {r['Debe']:,.2f}"  if r['Debe']  > 0 else "",
                        "Haber": f"$ {r['Haber']:,.2f}" if r['Haber'] > 0 else "",
                        "Saldo": f"$ {saldo_a:,.2f}",
                    })
                df_show = pd.DataFrame(filas_show)
                total_d = df_unif_fac['Debe'].sum()
                total_h = df_unif_fac['Haber'].sum()
                sf = total_d - total_h
                cm1, cm2, cm3 = st.columns(3)
                cm1.metric("Total Facturado / Viajes", f"$ {total_d:,.2f}")
                cm2.metric("Total Cobrado",             f"$ {total_h:,.2f}")
                cm3.metric("Saldo Pendiente",            f"$ {sf:,.2f}")
                st.markdown("---")
                st.dataframe(df_show, use_container_width=True, hide_index=True)
                html_r = generar_html_resumen(cli_cc, df_show, sf)
                st.download_button("📄 Descargar resumen", html_r,
                                   file_name=f"CuentaCorriente_{cli_cc.replace(' ','_')}.html",
                                   mime="text/html", key="dl_cc_fac")



# =============================================================
# CHEQUES
# =============================================================
elif sel == "CHEQUES":
    st.header("🏦 Gestión de Cheques")

    hoy = date.today()

    # ── Helper: badge de estado con color ──
    def badge_estado(estado):
        colores = {
            "PENDIENTE":    ("#fff3cd", "#856404", "⏳"),
            "CONCILIADO":   ("#d1e7dd", "#0f5132", "✅"),
            "RECHAZADO":    ("#f8d7da", "#842029", "❌"),
            "EN CARTERA":   ("#cff4fc", "#055160", "📂"),
            "DEPOSITADO":   ("#d1e7dd", "#0f5132", "🏦"),
            "APLICADO PAGO":("#e2d9f3", "#4a235a", "💸"),
            "VENCIDO":      ("#f8d7da", "#842029", "⚠️"),
        }
        bg, fg, ico = colores.get(estado, ("#f0f2f6","#333","•"))
        return f"<span style='background:{bg};color:{fg};padding:3px 10px;border-radius:12px;font-size:12px;font-weight:bold;'>{ico} {estado}</span>"

    # ── Helper: días para vencer ──
    def dias_vencer(fecha_str):
        """Devuelve días hasta la fecha de cobro (Fecha Vencimiento guardada)."""
        try:
            fv = pd.to_datetime(fecha_str).date()
            return (fv - hoy).days
        except:
            return None

    def estado_cheque_cartera(fecha_str):
        """
        Lógica real de cheques:
        - d > 0: aún no llegó la fecha de cobro → 'PENDIENTE'
        - -30 <= d <= 0: fecha de cobro pasó pero dentro de los 30 días de gracia → 'LISTO PARA COBRAR'
        - d < -30: superó los 30 días de gracia → 'VENCIDO'
        """
        try:
            fv = pd.to_datetime(fecha_str).date()
            d = (fv - hoy).days
            if d > 0:
                return "pendiente", d
            elif d >= -30:
                return "listo", d
            else:
                return "vencido", d
        except:
            return None, None

    # ── Alertas globales: cheques próximos a vencer (≤7 días) ──
    alertas = []
    if not st.session_state.cheques_emitidos.empty:
        for _, r in st.session_state.cheques_emitidos[st.session_state.cheques_emitidos['Estado'] == 'PENDIENTE'].iterrows():
            est_e, d = estado_cheque_cartera(r['Fecha Vencimiento'])
            if est_e == "pendiente" and d is not None and d <= 7:
                alertas.append(f"⚠️ **Cheque emitido #{r['Nro Cheque']}** a {r['Beneficiario']} vence en **{d} día(s)** — $ {float(r['Importe']):,.2f}")
            elif est_e == "listo":
                dias_rest = 30 + d
                alertas.append(f"🟡 **Cheque emitido #{r['Nro Cheque']}** a {r['Beneficiario']} — pendiente de cobro, puede presentarse hasta en **{dias_rest} día(s)** — $ {float(r['Importe']):,.2f}")
            elif est_e == "vencido":
                alertas.append(f"🔴 **Cheque emitido VENCIDO #{r['Nro Cheque']}** a {r['Beneficiario']} — pasaron los 30 días de gracia — $ {float(r['Importe']):,.2f}")
    if not st.session_state.cheques_cartera.empty:
        for _, r in st.session_state.cheques_cartera[st.session_state.cheques_cartera['Estado'] == 'EN CARTERA'].iterrows():
            estado_c, d = estado_cheque_cartera(r['Fecha Vencimiento'])
            if estado_c == "pendiente" and d <= 7:
                alertas.append(f"⚠️ **Cheque en cartera #{r['Nro Cheque']}** de {r['Librador']} — fecha de cobro en **{d} día(s)** — tener fondos disponibles — $ {float(r['Importe']):,.2f}")
            elif estado_c == "listo":
                dias_restantes = 30 + d  # d es negativo, ej: -5 → quedan 25 días
                alertas.append(f"✅ **Cheque LISTO PARA COBRAR #{r['Nro Cheque']}** de {r['Librador']} — puede cobrarse hasta en {dias_restantes} día(s) — $ {float(r['Importe']):,.2f}")
            elif estado_c == "vencido":
                alertas.append(f"🔴 **Cheque VENCIDO #{r['Nro Cheque']}** de {r['Librador']} — venció el plazo de 30 días — $ {float(r['Importe']):,.2f}")

    if alertas:
        with st.expander(f"🚨 {len(alertas)} alerta(s) de vencimiento", expanded=True):
            for a in alertas:
                st.warning(a)

    tab_emit, tab_cart, tab_venc, tab_buscar, tab_export = st.tabs(["📤 CHEQUES EMITIDOS", "📂 CHEQUES EN CARTERA", "📅 PRÓXIMOS VENCIMIENTOS", "🔍 BUSCADOR", "📥 EXPORTAR A EXCEL"])

    # ══════════════════════════════════════════════════════
    # TAB 1 — CHEQUES EMITIDOS
    # ══════════════════════════════════════════════════════
    with tab_emit:
        if st.session_state.get("msg_cheq_emit"):
            st.success(st.session_state.msg_cheq_emit)
            st.session_state.msg_cheq_emit = None

        with st.expander("➕ REGISTRAR NUEVO CHEQUE EMITIDO", expanded=False):
            with st.form("f_cheq_emit", clear_on_submit=True):
                ce1, ce2, ce3 = st.columns(3)
                nro_ch   = ce1.text_input("Nro de Cheque")
                tipo_ch  = ce2.selectbox("Tipo", ["FÍSICO", "ECHEQ"])
                banco_ch = ce3.text_input("Banco Emisor")
                ce4, ce5 = st.columns(2)
                benef    = ce4.text_input("Beneficiario (Proveedor / Persona)")
                imp_ch   = ce5.number_input("Importe $", min_value=0.0, step=0.01)
                ce6, ce7 = st.columns(2)
                f_emis   = ce6.date_input("Fecha de Emisión", value=hoy)
                f_venc   = ce7.date_input("Fecha de Vencimiento", value=f_emis + timedelta(days=30),
                                          help="Fecha límite de pago = fecha de emisión + 30 días")
                obs_ch   = st.text_input("Observaciones (opcional)")
                if st.form_submit_button("✅ REGISTRAR CHEQUE"):
                    if nro_ch and benef and imp_ch > 0:
                        nueva_fila = pd.DataFrame([[
                            str(f_emis), nro_ch, tipo_ch, banco_ch, benef,
                            imp_ch, str(f_venc), "PENDIENTE", "-", obs_ch
                        ]], columns=COL_CHEQ_EMITIDOS)
                        st.session_state.cheques_emitidos = pd.concat([st.session_state.cheques_emitidos, nueva_fila], ignore_index=True)
                        guardar_datos("cheques_emitidos", st.session_state.cheques_emitidos)
                        # Registrar egreso en tesorería
                        mov = pd.DataFrame([[
                            str(f_emis), "CHEQUE EMITIDO", "CAJA GENERAL", f"CHEQUE {tipo_ch}",
                            f"Cheque #{nro_ch} a {benef}", benef, -imp_ch, nro_ch
                        ]], columns=COL_TESORERIA)
                        st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, mov], ignore_index=True)
                        guardar_datos("tesoreria", st.session_state.tesoreria)
                        st.session_state.msg_cheq_emit = f"✅ Cheque #{nro_ch} emitido a {benef} por $ {imp_ch:,.2f} registrado."
                        st.rerun()
                    else:
                        st.warning("Completá Nro de Cheque, Beneficiario e Importe.")

        st.markdown("---")

        # Filtro por estado
        fe_col1, fe_col2 = st.columns([2, 2])
        filtro_emit = fe_col1.radio("Mostrar", ["TODOS", "PENDIENTES", "CONCILIADOS"], horizontal=True, key="filtro_emit")
        df_emit = st.session_state.cheques_emitidos.copy()
        lista_benef = ["Todos"] + sorted(df_emit['Beneficiario'].dropna().unique().tolist())
        filtro_benef = fe_col2.selectbox("Beneficiario", lista_benef, key="filtro_benef_emit")
        if filtro_emit == "PENDIENTES":
            df_emit = df_emit[df_emit['Estado'] == 'PENDIENTE']
        elif filtro_emit == "CONCILIADOS":
            df_emit = df_emit[df_emit['Estado'] == 'CONCILIADO']
        if filtro_benef != "Todos":
            df_emit = df_emit[df_emit['Beneficiario'] == filtro_benef]

        if df_emit.empty:
            st.info("No hay cheques en esta categoría.")
        else:
            for i, row in df_emit.iterrows():
                est_e, d_venc = estado_cheque_cartera(row['Fecha Vencimiento']) if row['Estado'] == 'PENDIENTE' else (None, None)
                alerta_color = ""
                if est_e == "vencido":    alerta_color = "border-left:4px solid #e74c3c;"
                elif est_e == "listo":    alerta_color = "border-left:4px solid #27ae60;"
                elif est_e == "pendiente" and d_venc is not None and d_venc <= 7: alerta_color = "border-left:4px solid #f39c12;"
                elif est_e == "pendiente": alerta_color = "border-left:4px solid #2ecc71;"

                if est_e == "vencido":
                    badge_venc = "&nbsp;·&nbsp; <b style=color:#e74c3c>⛔ VENCIDO (pasaron 30 días)</b>"
                elif est_e == "listo":
                    dias_rest = 30 + d_venc
                    badge_venc = f"&nbsp;·&nbsp; <b style=color:#f39c12>⚠️ PENDIENTE DE COBRO — {dias_rest}d restantes</b>"
                elif est_e == "pendiente" and d_venc is not None and d_venc <= 7:
                    badge_venc = f"&nbsp;·&nbsp; <b style=color:#f39c12>Vence en {d_venc}d</b>"
                else:
                    badge_venc = ""

                with st.container():
                    col_inf, col_acc = st.columns([0.82, 0.18])
                    col_inf.markdown(
                        f"<div style='background:#f8f9fa;border-radius:8px;padding:12px 16px;{alerta_color}'>"
                        f"<b>#{row['Nro Cheque']}</b> — {row['Tipo']} — {row['Banco']} &nbsp;|&nbsp; "
                        f"Beneficiario: <b>{row['Beneficiario']}</b><br>"
                        f"Emisión: {row['Fecha Emisión']} &nbsp;·&nbsp; Vencimiento: <b>{row['Fecha Vencimiento']}</b>"
                        f"{badge_venc}"
                        f"&nbsp;&nbsp; {badge_estado(row['Estado'])}<br>"
                        f"<b style='font-size:17px;color:#5e2d61;'>$ {float(row['Importe']):,.2f}</b>"
                        f"{'&nbsp;&nbsp;<small>' + str(row['Observaciones']) + '</small>' if str(row['Observaciones']) not in ['-',''] else ''}"
                        f"</div>", unsafe_allow_html=True
                    )
                    if row['Estado'] == 'PENDIENTE':
                        if col_acc.button("✅ Conciliar", key=f"conc_{i}"):
                            st.session_state.cheques_emitidos.loc[i, 'Estado'] = 'CONCILIADO'
                            st.session_state.cheques_emitidos.loc[i, 'Fecha Conciliación'] = str(hoy)
                            guardar_datos("cheques_emitidos", st.session_state.cheques_emitidos)
                            st.session_state.msg_cheq_emit = f"✅ Cheque #{row['Nro Cheque']} conciliado."
                            st.rerun()
                        if col_acc.button("❌ Rechazar", key=f"rech_{i}"):
                            st.session_state.cheques_emitidos.loc[i, 'Estado'] = 'RECHAZADO'
                            guardar_datos("cheques_emitidos", st.session_state.cheques_emitidos)
                            st.rerun()
                    st.divider()

    # ══════════════════════════════════════════════════════
    # TAB 2 — CHEQUES EN CARTERA (de terceros)
    # ══════════════════════════════════════════════════════
    with tab_cart:
        if st.session_state.get("msg_cheq_cart"):
            st.success(st.session_state.msg_cheq_cart)
            st.session_state.msg_cheq_cart = None

        with st.expander("➕ INGRESAR CHEQUE DE TERCERO A CARTERA", expanded=False):
            with st.form("f_cheq_cart", clear_on_submit=True):
                cc1, cc2, cc3 = st.columns(3)
                nro_cc   = cc1.text_input("Nro de Cheque")
                tipo_cc  = cc2.selectbox("Tipo", ["FÍSICO", "ECHEQ"])
                banco_cc = cc3.text_input("Banco Librador")
                cc4, cc5 = st.columns(2)
                librador = cc4.text_input("Librador (quien entregó el cheque)")
                imp_cc   = cc5.number_input("Importe $", min_value=0.0, step=0.01)
                cc6, cc7 = st.columns(2)
                f_rec    = cc6.date_input("Fecha de Recepción", value=hoy)
                f_venc_c = cc7.date_input("Fecha de Vencimiento", value=f_rec + timedelta(days=30),
                                          help="Fecha límite de cobro = fecha del cheque + 30 días")
                obs_cc   = st.text_input("Observaciones (opcional)")
                if st.form_submit_button("✅ AGREGAR A CARTERA"):
                    if nro_cc and librador and imp_cc > 0:
                        nueva_cc = pd.DataFrame([[
                            str(f_rec), nro_cc, tipo_cc, banco_cc, librador,
                            imp_cc, str(f_venc_c), "EN CARTERA", "-", "-", obs_cc
                        ]], columns=COL_CHEQ_CARTERA)
                        st.session_state.cheques_cartera = pd.concat([st.session_state.cheques_cartera, nueva_cc], ignore_index=True)
                        guardar_datos("cheques_cartera", st.session_state.cheques_cartera)
                        st.session_state.msg_cheq_cart = f"✅ Cheque #{nro_cc} de {librador} ingresado a cartera por $ {imp_cc:,.2f}."
                        st.rerun()
                    else:
                        st.warning("Completá Nro de Cheque, Librador e Importe.")

        with st.expander("📋 CARGA MASIVA DE CHEQUES (varios a la vez)", expanded=False):
            st.markdown("Completá los datos de cada cheque. Dejá en blanco las filas que no uses.")
            if "cheques_masivos" not in st.session_state:
                st.session_state.cheques_masivos = 5
            cant_filas = st.number_input("Cantidad de cheques a ingresar", min_value=1, max_value=50, value=st.session_state.cheques_masivos, step=1, key="cant_cheq_masivos")
            st.session_state.cheques_masivos = int(cant_filas)

            librador_global = st.text_input("Librador (se aplica a todos si completás aquí)", key="librador_global_masivo")
            banco_global     = st.text_input("Banco Librador (se aplica a todos si completás aquí)", key="banco_global_masivo")
            tipo_global      = st.selectbox("Tipo (se aplica a todos)", ["FÍSICO", "ECHEQ"], key="tipo_global_masivo")
            f_rec_global     = st.date_input("Fecha de Recepción (para todos)", value=hoy, key="frec_global_masivo")

            st.markdown("##### Detalle de cheques")
            cheques_ingresados = []
            for idx_m in range(int(cant_filas)):
                cm1, cm2, cm3, cm4 = st.columns([2, 2, 2, 2])
                nro_m  = cm1.text_input(f"Nro Cheque #{idx_m+1}", key=f"mnro_{idx_m}")
                imp_m  = cm2.number_input(f"Importe #{idx_m+1}", min_value=0.0, step=0.01, key=f"mimp_{idx_m}")
                fv_m   = cm3.date_input(f"Vencimiento #{idx_m+1}", value=f_rec_global + timedelta(days=30), key=f"mfv_{idx_m}")
                obs_m  = cm4.text_input(f"Obs #{idx_m+1}", key=f"mobs_{idx_m}")
                if nro_m and imp_m > 0:
                    cheques_ingresados.append({
                        "nro": nro_m,
                        "importe": imp_m,
                        "vencimiento": fv_m,
                        "obs": obs_m
                    })

            if st.button("✅ GUARDAR TODOS LOS CHEQUES", key="btn_masivo_cartera", type="primary"):
                if not cheques_ingresados:
                    st.warning("No hay cheques válidos para guardar. Completá al menos Nro y Importe.")
                else:
                    librador_uso = librador_global.strip() if librador_global.strip() else "SIN DATOS"
                    banco_uso    = banco_global.strip()    if banco_global.strip()    else "-"
                    nuevos = []
                    for ch in cheques_ingresados:
                        nuevos.append([
                            str(f_rec_global), ch["nro"], tipo_global, banco_uso,
                            librador_uso, ch["importe"], str(ch["vencimiento"]),
                            "EN CARTERA", "-", "-", ch["obs"]
                        ])
                    df_nuevos = pd.DataFrame(nuevos, columns=COL_CHEQ_CARTERA)
                    st.session_state.cheques_cartera = pd.concat([st.session_state.cheques_cartera, df_nuevos], ignore_index=True)
                    guardar_datos("cheques_cartera", st.session_state.cheques_cartera)
                    st.session_state.msg_cheq_cart = f"✅ {len(nuevos)} cheques de {librador_uso} ingresados a cartera."
                    st.rerun()

        st.markdown("---")

        filtro_cart = st.radio("Mostrar", ["EN CARTERA", "APLICADOS/DEPOSITADOS", "TODOS"], horizontal=True, key="filtro_cart")
        df_cart = st.session_state.cheques_cartera.copy()
        if filtro_cart == "EN CARTERA":
            df_cart = df_cart[df_cart['Estado'] == 'EN CARTERA']
        elif filtro_cart == "APLICADOS/DEPOSITADOS":
            df_cart = df_cart[df_cart['Estado'].isin(['DEPOSITADO', 'APLICADO PAGO'])]

        if df_cart.empty:
            st.info("No hay cheques en esta categoría.")
        else:
            for i, row in df_cart.iterrows():
                d_venc_c = dias_vencer(row['Fecha Vencimiento'])
                alerta_c = ""
                estado_c, d_venc_c = estado_cheque_cartera(row['Fecha Vencimiento']) if row['Estado'] == 'EN CARTERA' else (None, None)
                alerta_c = ""
                if estado_c == "vencido":   alerta_c = "border-left:4px solid #e74c3c;"
                elif estado_c == "listo":   alerta_c = "border-left:4px solid #2ecc71;"
                elif estado_c == "pendiente" and d_venc_c is not None and d_venc_c <= 7: alerta_c = "border-left:4px solid #f39c12;"
                else:                       alerta_c = "border-left:4px solid #aaa;"

                with st.container():
                    col_ci, col_ca = st.columns([0.75, 0.25])
                    col_ci.markdown(
                        f"<div style='background:#f8f9fa;border-radius:8px;padding:12px 16px;{alerta_c}'>"
                        f"<b>#{row['Nro Cheque']}</b> — {row['Tipo']} — {row['Banco Librador']} &nbsp;|&nbsp; "
                        f"Librador: <b>{row['Librador']}</b><br>"
                        f"Recibido: {row['Fecha Recepción']} &nbsp;·&nbsp; Vencimiento: <b>{row['Fecha Vencimiento']}</b>"
                        f"{'&nbsp;·&nbsp; <b style=color:#e74c3c>⛔ VENCIDO</b>' if estado_c == 'vencido' else ('&nbsp;·&nbsp; <b style=color:#27ae60>✅ LISTO PARA COBRAR</b>' if estado_c == 'listo' else (f'&nbsp;·&nbsp; <b style=color:#f39c12>⚠️ Cobra en {d_venc_c}d</b>' if estado_c == 'pendiente' and d_venc_c is not None and d_venc_c <= 7 else ''))}"
                        f"&nbsp;&nbsp; {badge_estado(row['Estado'])}<br>"
                        f"<b style='font-size:17px;color:#5e2d61;'>$ {float(row['Importe']):,.2f}</b>"
                        f"{'&nbsp;&nbsp; → ' + str(row['Destino']) if str(row['Destino']) not in ['-',''] else ''}"
                        f"</div>", unsafe_allow_html=True
                    )
                    if row['Estado'] == 'EN CARTERA':
                        # Depositar en banco
                        if col_ca.button("🏦 Depositar", key=f"dep_{i}"):
                            st.session_state[f"accion_cart_{i}"] = "depositar"
                        # Aplicar como pago a proveedor
                        if col_ca.button("💸 Pagar c/cheque", key=f"pag_{i}"):
                            st.session_state[f"accion_cart_{i}"] = "pagar"
                        # Editar fecha de vencimiento
                        if col_ca.button("📅 Editar fecha", key=f"edit_{i}"):
                            st.session_state[f"accion_cart_{i}"] = "editar_fecha"

                        accion = st.session_state.get(f"accion_cart_{i}")
                        if accion == "depositar":
                            with st.form(f"f_dep_{i}"):
                                banco_dep = st.selectbox("Banco donde depositar", ["BANCO GALICIA", "BANCO PROVINCIA", "BANCO SUPERVIELLE", "OTRO"])
                                if st.form_submit_button("✅ CONFIRMAR DEPÓSITO"):
                                    st.session_state.cheques_cartera.loc[i, 'Estado']          = 'DEPOSITADO'
                                    st.session_state.cheques_cartera.loc[i, 'Destino']         = banco_dep
                                    st.session_state.cheques_cartera.loc[i, 'Fecha Aplicación']= str(hoy)
                                    guardar_datos("cheques_cartera", st.session_state.cheques_cartera)
                                    # Ingreso en tesorería del banco
                                    mov_dep = pd.DataFrame([[
                                        str(hoy), "DEPÓSITO CHEQUE", banco_dep, "CHEQUE TERCERO",
                                        f"Depósito cheque #{row['Nro Cheque']} de {row['Librador']}",
                                        row['Librador'], float(row['Importe']), row['Nro Cheque']
                                    ]], columns=COL_TESORERIA)
                                    st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, mov_dep], ignore_index=True)
                                    guardar_datos("tesoreria", st.session_state.tesoreria)
                                    st.session_state[f"accion_cart_{i}"] = None
                                    st.session_state.msg_cheq_cart = f"✅ Cheque #{row['Nro Cheque']} depositado en {banco_dep}."
                                    st.rerun()
                                if st.form_submit_button("❌ Cancelar"):
                                    st.session_state[f"accion_cart_{i}"] = None; st.rerun()

                        elif accion == "editar_fecha":
                            with st.form(f"f_edit_fecha_{i}"):
                                try:
                                    fecha_actual = pd.to_datetime(row['Fecha Vencimiento']).date()
                                except:
                                    fecha_actual = hoy + timedelta(days=30)
                                nueva_fecha = st.date_input(
                                    "Nueva fecha de vencimiento",
                                    value=fecha_actual,
                                    help="Recordá: la fecha límite de cobro es fecha del cheque + 30 días"
                                )
                                st.caption(f"💡 Fecha del cheque + 30 días = **{pd.to_datetime(row['Fecha Recepción']).date() + timedelta(days=30)}**")
                                if st.form_submit_button("✅ GUARDAR FECHA"):
                                    st.session_state.cheques_cartera.loc[i, 'Fecha Vencimiento'] = str(nueva_fecha)
                                    guardar_datos("cheques_cartera", st.session_state.cheques_cartera)
                                    st.session_state[f"accion_cart_{i}"] = None
                                    st.session_state.msg_cheq_cart = f"✅ Fecha de vencimiento del cheque #{row['Nro Cheque']} actualizada a {nueva_fecha}."
                                    st.rerun()
                                if st.form_submit_button("❌ Cancelar"):
                                    st.session_state[f"accion_cart_{i}"] = None; st.rerun()

                        elif accion == "pagar":
                            with st.form(f"f_pag_{i}"):
                                prov_pag = st.selectbox("Proveedor a pagar", st.session_state.proveedores['Razón Social'].unique() if not st.session_state.proveedores.empty else [""])
                                if st.form_submit_button("✅ CONFIRMAR PAGO"):
                                    st.session_state.cheques_cartera.loc[i, 'Estado']          = 'APLICADO PAGO'
                                    st.session_state.cheques_cartera.loc[i, 'Destino']         = prov_pag
                                    st.session_state.cheques_cartera.loc[i, 'Fecha Aplicación']= str(hoy)
                                    guardar_datos("cheques_cartera", st.session_state.cheques_cartera)
                                    # Egreso en tesorería
                                    mov_pag = pd.DataFrame([[
                                        str(hoy), "PAGO CHEQUE TERCERO", "CARTERA", "CHEQUE TERCERO",
                                        f"Pago a {prov_pag} con cheque #{row['Nro Cheque']} de {row['Librador']}",
                                        prov_pag, -float(row['Importe']), row['Nro Cheque']
                                    ]], columns=COL_TESORERIA)
                                    st.session_state.tesoreria = pd.concat([st.session_state.tesoreria, mov_pag], ignore_index=True)
                                    guardar_datos("tesoreria", st.session_state.tesoreria)
                                    st.session_state[f"accion_cart_{i}"] = None
                                    st.session_state.msg_cheq_cart = f"✅ Cheque #{row['Nro Cheque']} aplicado como pago a {prov_pag}."
                                    st.rerun()
                                if st.form_submit_button("❌ Cancelar"):
                                    st.session_state[f"accion_cart_{i}"] = None; st.rerun()
                    st.divider()

    # ══════════════════════════════════════════════════════
    # TAB 3 — PRÓXIMOS VENCIMIENTOS (ambos tipos)
    # ══════════════════════════════════════════════════════
    with tab_venc:
        st.markdown("##### Cheques con vencimiento en los próximos 30 días")
        dias_filtro = st.slider("Días hacia adelante", 7, 60, 30, key="dias_venc_slider")

        filas_venc = []

        if not st.session_state.cheques_emitidos.empty:
            for _, r in st.session_state.cheques_emitidos[st.session_state.cheques_emitidos['Estado'] == 'PENDIENTE'].iterrows():
                d = dias_vencer(r['Fecha Vencimiento'])
                if d is not None and d <= dias_filtro:
                    filas_venc.append({
                        "Tipo": "📤 EMITIDO",
                        "Nro Cheque": r['Nro Cheque'],
                        "Modalidad": r['Tipo'],
                        "Banco": r['Banco'],
                        "Contraparte": r['Beneficiario'],
                        "Importe": float(r['Importe']),
                        "Vencimiento": r['Fecha Vencimiento'],
                        "Días": d,
                        "Estado": r['Estado']
                    })

        if not st.session_state.cheques_cartera.empty:
            for _, r in st.session_state.cheques_cartera[st.session_state.cheques_cartera['Estado'] == 'EN CARTERA'].iterrows():
                est_c, d = estado_cheque_cartera(r['Fecha Vencimiento'])
                # Mostrar si: está próximo a su fecha de cobro, listo para cobrar, o vencido
                if est_c in ("pendiente", "listo", "vencido"):
                    dias_display = d if est_c == "pendiente" else (30 + d if est_c == "listo" else d)
                    label_est = "LISTO PARA COBRAR" if est_c == "listo" else ("VENCIDO" if est_c == "vencido" else f"Cobra en {d}d")
                    filas_venc.append({
                        "Tipo": "📂 CARTERA",
                        "Nro Cheque": r['Nro Cheque'],
                        "Modalidad": r['Tipo'],
                        "Banco": r['Banco Librador'],
                        "Contraparte": r['Librador'],
                        "Importe": float(r['Importe']),
                        "Vencimiento": r['Fecha Vencimiento'],
                        "Días": d,
                        "Estado real": label_est
                    })

        if filas_venc:
            df_venc = pd.DataFrame(filas_venc).sort_values("Días")
            total_emit_pend = df_venc[df_venc['Tipo'] == '📤 EMITIDO']['Importe'].sum()
            total_cart_pend = df_venc[df_venc['Tipo'] == '📂 CARTERA']['Importe'].sum()

            kv1, kv2, kv3 = st.columns(3)
            kv1.metric("Total cheques", len(df_venc))
            kv2.metric("📤 A pagar (emitidos)", f"$ {total_emit_pend:,.2f}")
            kv3.metric("📂 A cobrar/depositar", f"$ {total_cart_pend:,.2f}")
            st.markdown("---")

            for _, r in df_venc.iterrows():
                estado_real = r.get("Estado real", "")
                d_val = int(r["Días"])
                if r["Tipo"] == "📂 CARTERA":
                    if estado_real == "VENCIDO":
                        color_d = "#e74c3c"; label_d = "⛔ VENCIDO (pasaron 30 días)"
                    elif estado_real == "LISTO PARA COBRAR":
                        dias_rest = 30 + d_val
                        color_d = "#27ae60"; label_d = f"✅ LISTO PARA COBRAR — {dias_rest}d restantes"
                    elif d_val == 0:
                        color_d = "#f39c12"; label_d = "⚠️ Fecha de cobro: HOY"
                    elif d_val <= 3:
                        color_d = "#f39c12"; label_d = f"⚠️ Cobra en {d_val}d — tener fondos"
                    else:
                        color_d = "#3498db"; label_d = f"Cobra en {d_val}d"
                else:
                    # Cheques emitidos: lógica original
                    if d_val < 0:
                        color_d = "#e74c3c"; label_d = f"VENCIDO hace {abs(d_val)}d"
                    elif d_val == 0:
                        color_d = "#e74c3c"; label_d = "VENCE HOY"
                    elif d_val <= 3:
                        color_d = "#e74c3c"; label_d = f"Vence en {d_val}d"
                    elif d_val <= 7:
                        color_d = "#f39c12"; label_d = f"Vence en {d_val}d"
                    else:
                        color_d = "#2ecc71"; label_d = f"Vence en {d_val}d"

                st.markdown(
                    f"<div style='background:#f8f9fa;border-radius:8px;padding:12px 16px;margin-bottom:8px;"
                    f"border-left:5px solid {color_d};display:flex;justify-content:space-between;align-items:center;'>"
                    f"<div>{r['Tipo']} &nbsp; <b>#{r['Nro Cheque']}</b> &nbsp;·&nbsp; {r['Modalidad']} &nbsp;·&nbsp; {r['Banco']}<br>"
                    f"Contraparte: <b>{r['Contraparte']}</b> &nbsp;·&nbsp; Vence: <b>{r['Vencimiento']}</b></div>"
                    f"<div style='text-align:right;'>"
                    f"<div style='font-size:18px;font-weight:bold;color:#5e2d61;'>$ {r['Importe']:,.2f}</div>"
                    f"<div style='font-size:13px;font-weight:bold;color:{color_d};'>{label_d}</div>"
                    f"</div></div>",
                    unsafe_allow_html=True
                )
        else:
            st.success(f"✅ No hay cheques con vencimiento en los próximos {dias_filtro} días.")

    # ══════════════════════════════════════════════════════
    # TAB 4 — BUSCADOR DE CHEQUES
    # ══════════════════════════════════════════════════════
    with tab_buscar:
        st.markdown("##### 🔍 Buscador de Cheques")
        st.markdown("Ingresá el número de cheque (total o parcial) para encontrarlo en cualquier cartera.")

        busq_col1, busq_col2 = st.columns([2, 2])
        nro_buscar   = busq_col1.text_input("Número de Cheque", placeholder="Ej: 001234", key="busq_nro")
        tipo_buscar  = busq_col2.selectbox("Buscar en", ["Todos", "📤 Emitidos", "📂 Cartera"], key="busq_tipo")

        resultados_emit = []
        resultados_cart = []

        if nro_buscar.strip():
            nro_norm = nro_buscar.strip()

            # ── Buscar en cheques emitidos ──
            if tipo_buscar in ("Todos", "📤 Emitidos") and not st.session_state.cheques_emitidos.empty:
                mask_e = st.session_state.cheques_emitidos['Nro Cheque'].astype(str).str.contains(nro_norm, case=False, na=False)
                resultados_emit = st.session_state.cheques_emitidos[mask_e].copy()

            # ── Buscar en cheques en cartera ──
            if tipo_buscar in ("Todos", "📂 Cartera") and not st.session_state.cheques_cartera.empty:
                mask_c = st.session_state.cheques_cartera['Nro Cheque'].astype(str).str.contains(nro_norm, case=False, na=False)
                resultados_cart = st.session_state.cheques_cartera[mask_c].copy()

            total_encontrados = (len(resultados_emit) if hasattr(resultados_emit, '__len__') else 0) + \
                                (len(resultados_cart) if hasattr(resultados_cart, '__len__') else 0)

            if total_encontrados == 0:
                st.warning(f"⚠️ No se encontró ningún cheque con el número **{nro_norm}**.")
            else:
                st.success(f"✅ Se encontraron **{total_encontrados}** resultado(s) para el número **{nro_norm}**.")
                st.markdown("---")

                # ── Resultados EMITIDOS ──
                if hasattr(resultados_emit, '__len__') and len(resultados_emit) > 0:
                    st.markdown("#### 📤 Cheques Emitidos")
                    for _, r in resultados_emit.iterrows():
                        est_r, d_r = estado_cheque_cartera(r['Fecha Vencimiento']) if r['Estado'] == 'PENDIENTE' else (None, None)
                        if r['Estado'] == 'CONCILIADO':
                            color_b = "#2ecc71"; badge_txt = "✅ CONCILIADO"
                        elif est_r == "vencido":
                            color_b = "#e74c3c"; badge_txt = "⛔ VENCIDO"
                        elif est_r == "listo":
                            color_b = "#f39c12"; badge_txt = f"⚠️ LISTO PARA COBRAR"
                        else:
                            color_b = "#3498db"; badge_txt = f"🕐 PENDIENTE"

                        st.markdown(
                            f"<div style='background:#f0f7ff;border-radius:10px;padding:16px 20px;"
                            f"border-left:6px solid {color_b};margin-bottom:12px;'>"
                            f"<div style='display:flex;justify-content:space-between;align-items:flex-start;'>"
                            f"<div>"
                            f"<span style='font-size:18px;font-weight:bold;color:#5e2d61;'>Cheque #{r['Nro Cheque']}</span>"
                            f"&nbsp;&nbsp;<span style='background:{color_b};color:white;padding:3px 10px;"
                            f"border-radius:12px;font-size:12px;font-weight:bold;'>{badge_txt}</span><br><br>"
                            f"<table style='font-size:13px;border-collapse:collapse;'>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>🏦 Banco emisor</td>"
                            f"<td style='font-weight:bold;'>{r.get('Banco','-')}</td></tr>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>🏷️ Tipo</td>"
                            f"<td>{r.get('Tipo','-')}</td></tr>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>👤 Beneficiario</td>"
                            f"<td style='font-weight:bold;'>{r.get('Beneficiario','-')}</td></tr>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>📅 Fecha Emisión</td>"
                            f"<td>{r.get('Fecha Emisión','-')}</td></tr>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>📅 Fecha Vencimiento</td>"
                            f"<td><b>{r.get('Fecha Vencimiento','-')}</b></td></tr>"
                            f"{'<tr><td style=color:#888;padding:3px 16px 3px 0;>📅 Fecha Conciliación</td><td>' + str(r.get('Fecha Conciliación','-')) + '</td></tr>' if str(r.get('Fecha Conciliación','-')) not in ['-',''] else ''}"
                            f"{'<tr><td style=color:#888;padding:3px 16px 3px 0;>📝 Observaciones</td><td>' + str(r.get('Observaciones','-')) + '</td></tr>' if str(r.get('Observaciones','-')) not in ['-','','nan'] else ''}"
                            f"</table></div>"
                            f"<div style='text-align:right;'>"
                            f"<div style='font-size:28px;font-weight:bold;color:#5e2d61;'>$ {float(r['Importe']):,.2f}</div>"
                            f"<div style='font-size:11px;color:#888;margin-top:4px;'>IMPORTE</div>"
                            f"</div></div></div>",
                            unsafe_allow_html=True
                        )

                # ── Resultados CARTERA ──
                if hasattr(resultados_cart, '__len__') and len(resultados_cart) > 0:
                    st.markdown("#### 📂 Cheques en Cartera")
                    for _, r in resultados_cart.iterrows():
                        est_r, d_r = estado_cheque_cartera(r['Fecha Vencimiento']) if r['Estado'] == 'EN CARTERA' else (None, None)
                        if r['Estado'] == 'DEPOSITADO':
                            color_b = "#2ecc71"; badge_txt = "✅ DEPOSITADO"
                        elif r['Estado'] == 'APLICADO PAGO':
                            color_b = "#9b59b6"; badge_txt = "💸 APLICADO A PAGO"
                        elif est_r == "vencido":
                            color_b = "#e74c3c"; badge_txt = "⛔ VENCIDO"
                        elif est_r == "listo":
                            color_b = "#27ae60"; badge_txt = "✅ LISTO PARA COBRAR"
                        else:
                            color_b = "#3498db"; badge_txt = "🕐 EN CARTERA"

                        destino_txt = ""
                        if str(r.get('Destino', '-')) not in ['-', '', 'nan']:
                            destino_txt = f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>🏁 Destino / Aplicado a</td><td style='font-weight:bold;'>{r['Destino']}</td></tr>"
                        fecha_aplic_txt = ""
                        if str(r.get('Fecha Aplicación', '-')) not in ['-', '', 'nan']:
                            fecha_aplic_txt = f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>📅 Fecha Aplicación</td><td>{r['Fecha Aplicación']}</td></tr>"

                        st.markdown(
                            f"<div style='background:#f0fff8;border-radius:10px;padding:16px 20px;"
                            f"border-left:6px solid {color_b};margin-bottom:12px;'>"
                            f"<div style='display:flex;justify-content:space-between;align-items:flex-start;'>"
                            f"<div>"
                            f"<span style='font-size:18px;font-weight:bold;color:#5e2d61;'>Cheque #{r['Nro Cheque']}</span>"
                            f"&nbsp;&nbsp;<span style='background:{color_b};color:white;padding:3px 10px;"
                            f"border-radius:12px;font-size:12px;font-weight:bold;'>{badge_txt}</span><br><br>"
                            f"<table style='font-size:13px;border-collapse:collapse;'>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>🏦 Banco librador</td>"
                            f"<td style='font-weight:bold;'>{r.get('Banco Librador','-')}</td></tr>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>🏷️ Tipo</td>"
                            f"<td>{r.get('Tipo','-')}</td></tr>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>✍️ Librador</td>"
                            f"<td style='font-weight:bold;'>{r.get('Librador','-')}</td></tr>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>📥 Recibido de</td>"
                            f"<td>{r.get('Observaciones','-')}</td></tr>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>📅 Fecha Recepción</td>"
                            f"<td>{r.get('Fecha Recepción','-')}</td></tr>"
                            f"<tr><td style='color:#888;padding:3px 16px 3px 0;'>📅 Fecha Vencimiento</td>"
                            f"<td><b>{r.get('Fecha Vencimiento','-')}</b></td></tr>"
                            f"{destino_txt}{fecha_aplic_txt}"
                            f"</table></div>"
                            f"<div style='text-align:right;'>"
                            f"<div style='font-size:28px;font-weight:bold;color:#5e2d61;'>$ {float(r['Importe']):,.2f}</div>"
                            f"<div style='font-size:11px;color:#888;margin-top:4px;'>IMPORTE</div>"
                            f"</div></div></div>",
                            unsafe_allow_html=True
                        )
        else:
            # Estado inicial — mostrar stats rápidos
            total_emit = len(st.session_state.cheques_emitidos) if not st.session_state.cheques_emitidos.empty else 0
            total_cart = len(st.session_state.cheques_cartera)  if not st.session_state.cheques_cartera.empty  else 0
            emit_pend  = len(st.session_state.cheques_emitidos[st.session_state.cheques_emitidos['Estado'] == 'PENDIENTE']) if not st.session_state.cheques_emitidos.empty else 0
            cart_disp  = len(st.session_state.cheques_cartera[st.session_state.cheques_cartera['Estado'] == 'EN CARTERA'])  if not st.session_state.cheques_cartera.empty  else 0
            sb1, sb2, sb3, sb4 = st.columns(4)
            sb1.metric("📤 Emitidos (total)", total_emit)
            sb2.metric("📤 Pendientes",        emit_pend)
            sb3.metric("📂 En cartera (total)", total_cart)
            sb4.metric("📂 Disponibles",        cart_disp)
            st.info("💡 Escribí el número de cheque arriba para buscarlo.")

    # ══════════════════════════════════════════════════════
    # TAB 5 — EXPORTAR CHEQUES A EXCEL
    # ══════════════════════════════════════════════════════
    with tab_export:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        st.markdown("##### 📥 Exportar cheques a Excel")
        st.markdown("Elegí el rango de fechas y qué cheques exportar.")

        ex1, ex2, ex3 = st.columns(3)
        fecha_desde = ex1.date_input("Desde", value=date.today().replace(day=1), key="exp_desde")
        fecha_hasta = ex2.date_input("Hasta", value=date.today(), key="exp_hasta")
        tipo_export = ex3.multiselect(
            "Incluir",
            ["📤 Emitidos", "📂 Cartera"],
            default=["📤 Emitidos", "📂 Cartera"],
            key="exp_tipo"
        )

        if st.button("📊 GENERAR EXCEL", key="btn_generar_excel", type="primary"):
            wb = Workbook()
            wb.remove(wb.active)  # quitar hoja vacía por defecto

            # Estilos
            color_header  = "4B1A6B"   # violeta empresa
            color_total   = "F39C12"   # naranja total
            color_emitido = "EAF4FB"   # celeste suave
            color_cartera = "F0FFF4"   # verde suave
            color_vencido = "FDEDEC"   # rojo suave
            color_listo   = "EAFAF1"   # verde brillante

            thin = Side(style="thin", color="CCCCCC")
            border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

            def header_style(cell, bg=color_header):
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
                cell.fill = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_all

            def data_style(cell, bg="FFFFFF", bold=False, color="000000", num_fmt=None):
                cell.font = Font(bold=bold, color=color, name="Arial", size=10)
                cell.fill = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(vertical="center")
                cell.border = border_all
                if num_fmt:
                    cell.number_format = num_fmt

            def calc_estado(fecha_str):
                try:
                    fv = pd.to_datetime(fecha_str).date()
                    d  = (fv - date.today()).days
                    if d > 0:   return f"Cobra en {d}d"
                    elif d >= -30: return "LISTO PARA COBRAR"
                    else:       return "VENCIDO"
                except:
                    return "-"

            # ── HOJA EMITIDOS ──
            if "📤 Emitidos" in tipo_export and not st.session_state.cheques_emitidos.empty:
                ws_e = wb.create_sheet("CHEQUES EMITIDOS")
                ws_e.sheet_view.showGridLines = False

                # Título
                ws_e.merge_cells("A1:J1")
                ws_e["A1"] = f"CHEQUES EMITIDOS — {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
                ws_e["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
                ws_e["A1"].fill = PatternFill("solid", start_color=color_header)
                ws_e["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws_e.row_dimensions[1].height = 28

                headers_e = ["Nro Cheque", "Tipo", "Banco", "Beneficiario", "Fecha Emisión", "Fecha Vencimiento", "Importe $", "Estado", "Situación", "Observaciones"]
                for col, h in enumerate(headers_e, 1):
                    cell = ws_e.cell(row=2, column=col, value=h)
                    header_style(cell)
                ws_e.row_dimensions[2].height = 22

                df_e = st.session_state.cheques_emitidos.copy()
                df_e['_fecha'] = pd.to_datetime(df_e['Fecha Emisión'], errors='coerce').dt.date
                df_e = df_e[(df_e['_fecha'] >= fecha_desde) & (df_e['_fecha'] <= fecha_hasta)]

                data_row = 3
                for _, r in df_e.iterrows():
                    sit = calc_estado(r['Fecha Vencimiento'])
                    bg = color_vencido if sit == "VENCIDO" else (color_listo if sit == "LISTO PARA COBRAR" else color_emitido)
                    vals = [r['Nro Cheque'], r['Tipo'], r['Banco'], r['Beneficiario'],
                            r['Fecha Emisión'], r['Fecha Vencimiento'], float(r['Importe']),
                            r['Estado'], sit, str(r['Observaciones']) if str(r['Observaciones']) != '-' else '']
                    for col, v in enumerate(vals, 1):
                        cell = ws_e.cell(row=data_row, column=col, value=v)
                        num_fmt = '$#,##0.00' if col == 7 else None
                        data_style(cell, bg=bg, num_fmt=num_fmt)
                    ws_e.row_dimensions[data_row].height = 18
                    data_row += 1

                # Fila TOTAL con fórmula Excel
                total_row = data_row
                ws_e.merge_cells(f"A{total_row}:F{total_row}")
                cell_lbl = ws_e[f"A{total_row}"]
                cell_lbl.value = "TOTAL"
                cell_lbl.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                cell_lbl.fill = PatternFill("solid", start_color=color_total)
                cell_lbl.alignment = Alignment(horizontal="right", vertical="center")
                cell_lbl.border = border_all

                cell_tot = ws_e[f"G{total_row}"]
                cell_tot.value = f"=SUM(G3:G{total_row-1})"
                cell_tot.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
                cell_tot.fill = PatternFill("solid", start_color=color_total)
                cell_tot.number_format = '$#,##0.00'
                cell_tot.alignment = Alignment(horizontal="right", vertical="center")
                cell_tot.border = border_all
                ws_e.row_dimensions[total_row].height = 24

                for col_idx in range(8, 11):
                    cell = ws_e.cell(row=total_row, column=col_idx)
                    cell.fill = PatternFill("solid", start_color=color_total)
                    cell.border = border_all

                # Anchos columnas
                anchos_e = [14, 10, 22, 35, 16, 18, 18, 14, 22, 25]
                for i, w in enumerate(anchos_e, 1):
                    ws_e.column_dimensions[get_column_letter(i)].width = w

            # ── HOJA CARTERA ──
            if "📂 Cartera" in tipo_export and not st.session_state.cheques_cartera.empty:
                ws_c = wb.create_sheet("CHEQUES EN CARTERA")
                ws_c.sheet_view.showGridLines = False

                ws_c.merge_cells("A1:K1")
                ws_c["A1"] = f"CHEQUES EN CARTERA — {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"
                ws_c["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
                ws_c["A1"].fill = PatternFill("solid", start_color=color_header)
                ws_c["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws_c.row_dimensions[1].height = 28

                headers_c = ["Nro Cheque", "Tipo", "Banco Librador", "Librador", "Fecha Recepción", "Fecha Vencimiento", "Importe $", "Estado", "Situación", "Destino", "Observaciones"]
                for col, h in enumerate(headers_c, 1):
                    cell = ws_c.cell(row=2, column=col, value=h)
                    header_style(cell)
                ws_c.row_dimensions[2].height = 22

                df_c = st.session_state.cheques_cartera.copy()
                df_c['_fecha'] = pd.to_datetime(df_c['Fecha Recepción'], errors='coerce').dt.date
                df_c = df_c[(df_c['_fecha'] >= fecha_desde) & (df_c['_fecha'] <= fecha_hasta)]

                data_row_c = 3
                for _, r in df_c.iterrows():
                    sit = calc_estado(r['Fecha Vencimiento'])
                    bg = color_vencido if sit == "VENCIDO" else (color_listo if sit == "LISTO PARA COBRAR" else color_cartera)
                    vals = [r['Nro Cheque'], r['Tipo'], r['Banco Librador'], r['Librador'],
                            r['Fecha Recepción'], r['Fecha Vencimiento'], float(r['Importe']),
                            r['Estado'], sit,
                            str(r['Destino']) if str(r['Destino']) not in ['-',''] else '',
                            str(r['Observaciones']) if str(r['Observaciones']) != '-' else '']
                    for col, v in enumerate(vals, 1):
                        cell = ws_c.cell(row=data_row_c, column=col, value=v)
                        num_fmt = '$#,##0.00' if col == 7 else None
                        data_style(cell, bg=bg, num_fmt=num_fmt)
                    ws_c.row_dimensions[data_row_c].height = 18
                    data_row_c += 1

                total_row_c = data_row_c
                ws_c.merge_cells(f"A{total_row_c}:F{total_row_c}")
                cell_lbl_c = ws_c[f"A{total_row_c}"]
                cell_lbl_c.value = "TOTAL"
                cell_lbl_c.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                cell_lbl_c.fill = PatternFill("solid", start_color=color_total)
                cell_lbl_c.alignment = Alignment(horizontal="right", vertical="center")
                cell_lbl_c.border = border_all

                cell_tot_c = ws_c[f"G{total_row_c}"]
                cell_tot_c.value = f"=SUM(G3:G{total_row_c-1})"
                cell_tot_c.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
                cell_tot_c.fill = PatternFill("solid", start_color=color_total)
                cell_tot_c.number_format = '$#,##0.00'
                cell_tot_c.alignment = Alignment(horizontal="right", vertical="center")
                cell_tot_c.border = border_all
                ws_c.row_dimensions[total_row_c].height = 24

                for col_idx in range(8, 12):
                    cell = ws_c.cell(row=total_row_c, column=col_idx)
                    cell.fill = PatternFill("solid", start_color=color_total)
                    cell.border = border_all

                anchos_c = [14, 10, 22, 30, 16, 18, 18, 14, 22, 20, 25]
                for i, w in enumerate(anchos_c, 1):
                    ws_c.column_dimensions[get_column_letter(i)].width = w

            # ── HOJA RESUMEN ──
            ws_r = wb.create_sheet("RESUMEN", 0)
            ws_r.sheet_view.showGridLines = False
            ws_r.column_dimensions['A'].width = 35
            ws_r.column_dimensions['B'].width = 22

            ws_r.merge_cells("A1:B1")
            ws_r["A1"] = "RESUMEN DE CHEQUES"
            ws_r["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=14)
            ws_r["A1"].fill = PatternFill("solid", start_color=color_header)
            ws_r["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws_r.row_dimensions[1].height = 32

            ws_r["A2"] = "Período"
            ws_r["B2"] = f"{fecha_desde.strftime('%d/%m/%Y')} — {fecha_hasta.strftime('%d/%m/%Y')}"
            ws_r["A3"] = "Generado"
            ws_r["B3"] = date.today().strftime('%d/%m/%Y')
            for r_idx in [2, 3]:
                for c_idx in ['A', 'B']:
                    ws_r[f"{c_idx}{r_idx}"].font = Font(name="Arial", size=10)
                    ws_r[f"{c_idx}{r_idx}"].border = border_all

            resumen_data = []
            if "📤 Emitidos" in tipo_export and "CHEQUES EMITIDOS" in [s.title for s in wb.worksheets]:
                ws_ref_e = wb["CHEQUES EMITIDOS"]
                last_e = ws_ref_e.max_row
                resumen_data.append(("Total Cheques Emitidos (cant.)", f"=COUNTA('CHEQUES EMITIDOS'!A3:A{last_e-1})"))
                resumen_data.append(("Total Cheques Emitidos ($)", f"='CHEQUES EMITIDOS'!G{last_e}"))
            if "📂 Cartera" in tipo_export and "CHEQUES EN CARTERA" in [s.title for s in wb.worksheets]:
                ws_ref_c = wb["CHEQUES EN CARTERA"]
                last_c = ws_ref_c.max_row
                resumen_data.append(("Total Cheques en Cartera (cant.)", f"=COUNTA('CHEQUES EN CARTERA'!A3:A{last_c-1})"))
                resumen_data.append(("Total Cheques en Cartera ($)", f"='CHEQUES EN CARTERA'!G{last_c}"))

            for idx, (label, formula) in enumerate(resumen_data, 5):
                ws_r[f"A{idx}"] = label
                ws_r[f"B{idx}"] = formula
                ws_r[f"A{idx}"].font = Font(bold=True, name="Arial", size=10)
                ws_r[f"B{idx}"].font = Font(bold=True, name="Arial", size=11, color="4B1A6B")
                ws_r[f"B{idx}"].number_format = '$#,##0.00' if '$' in label else '#,##0'
                ws_r[f"A{idx}"].border = border_all
                ws_r[f"B{idx}"].border = border_all
                ws_r.row_dimensions[idx].height = 22

            # Guardar en buffer y ofrecer descarga
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            nombre_archivo = f"Cheques_{fecha_desde.strftime('%d%m%Y')}_{fecha_hasta.strftime('%d%m%Y')}.xlsx"
            st.success(f"✅ Excel generado con {len(tipo_export)} hoja(s) de datos + resumen.")
            st.download_button(
                label="⬇️ DESCARGAR EXCEL",
                data=buf.getvalue(),
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )
